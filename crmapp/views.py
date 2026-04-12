import csv
import json
import random
import datetime
from io import BytesIO
from decimal import Decimal
import base64
from time import time
import traceback
import matplotlib.pyplot as plt
import openpyxl

from django import contrib
from django.conf import settings
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q, Sum, Count, Max
from django.db.models.functions import Lower
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User

from amc.models import AMCContract, AMCServiceSchedule

from .forms import (
    InventoryServiceForm,
    InventoryAddForm,
    AddProductForm,
    UpdateProductForm,
    LeadImportForm
)

from .models import (
    ServiceProduct,
    customer_details,
    service_management,
    quotation,
    invoice,
    Product,
    Inventory_summary,
    Inventory_add,
    
    UserProfile,
    SalesPerson,
    TechnicianProfile,
    lead_management,
    firstfollowup,
    secondfollowup,
    thirdfollowup,
    finalfollowup,
    BranchManager,
    OperationPerson,
    ProductRequiredItem,
    ServiceProductItem,
)

from .decorators import role_required
from schedule_meetings.models import Meeting

from django.shortcuts import render, redirect



state_map = {
        'Andaman and Nicobar Islands': {'code': 35, 'shortcut': 'AN'},
        'Andhra Pradesh': {'code': 37, 'shortcut': 'AP'},
        'Arunachal Pradesh': {'code': 12, 'shortcut': 'AR'},
        'Assam': {'code': 18, 'shortcut': 'AS'},
        'Bihar': {'code': 10, 'shortcut': 'BR'},
        'Chandigarh': {'code': 4, 'shortcut': 'CH'},
        'Chhattisgarh': {'code': 22, 'shortcut': 'CG'},
        'Dadra and Nagar Haveli and Daman and Diu': {'code': 26, 'shortcut': 'DNHDD'},
        'Delhi': {'code': 7, 'shortcut': 'DL'},
        'Goa': {'code': 30, 'shortcut': 'GA'},
        'Gujarat': {'code': 24, 'shortcut': 'GJ'},
        'Haryana': {'code': 6, 'shortcut': 'HR'},
        'Himachal Pradesh': {'code': 2, 'shortcut': 'HP'},
        'Jammu and Kashmir': {'code': 1, 'shortcut': 'JK'},
        'Jharkhand': {'code': 20, 'shortcut': 'JH'},
        'Karnataka': {'code': 29, 'shortcut': 'KA'},
        'Kerala': {'code': 32, 'shortcut': 'KL'},
        'Ladakh': {'code': 38, 'shortcut': 'LA'},
        'Lakshadweep': {'code': 31, 'shortcut': 'LD'},
        'Madhya Pradesh': {'code': 23, 'shortcut': 'MP'},
        'Maharashtra': {'code': 27, 'shortcut': 'MH'},
        'Manipur': {'code': 14, 'shortcut': 'MN'},
        'Meghalaya': {'code': 17, 'shortcut': 'ML'},
        'Mizoram': {'code': 15, 'shortcut': 'MZ'},
        'Nagaland': {'code': 13, 'shortcut': 'NL'},
        'Odisha': {'code': 21, 'shortcut': 'OD'},
        'Other Country': {'code': 99, 'shortcut': 'OC'},
        'Other Territory': {'code': 97, 'shortcut': 'OT'},
        'Puducherry': {'code': 34, 'shortcut': 'PY'},
        'Punjab': {'code': 3, 'shortcut': 'PB'},
        'Rajasthan': {'code': 8, 'shortcut': 'RJ'},
        'Sikkim': {'code': 11, 'shortcut': 'SK'},
        'Tamil Nadu': {'code': 33, 'shortcut': 'TN'},
        'Telangana': {'code': 36, 'shortcut': 'TS'},
        'Tripura': {'code': 16, 'shortcut': 'TR'},
        'Uttar Pradesh': {'code': 9, 'shortcut': 'UP'},
        'Uttarakhand': {'code': 5, 'shortcut': 'UK'},
        'West Bengal': {'code': 19, 'shortcut': 'WB'}
        } 



def landing_page(request):
    return render(request , 'landing_page.html')


def login_view(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user:
            try:
                role = user.userprofile.role

                if role in ["admin", "sales","branch_manager","operation_person","HO_operation","HO_manager"]:
                    login(request, user)
                    return redirect("index")   # Admin/Sales Dashboard
                elif role == "technician":
                    login(request, user)
                    return redirect("technician_dashboard")  # Technician Dashboard
                else:
                    messages.error(request, "Access denied: Invalid role.")
            except UserProfile.DoesNotExist:
                messages.error(request, "User profile not found.")
        else:
            messages.error(request, "Invalid username or password.")
    
    return render(request, "landing_page.html")

@login_required
@role_required(['admin', 'sales', 'branch_manager','operation_person',"HO_operation","HO_manager"])
def index(request):
        # Fetch data for Service Management
        service_data = service_management.objects.values('selected_services').annotate(total_charges=Sum('total_charges'))

        # Fetch data for Quotations
        quotation_data = quotation.objects.values('quotation_date').annotate(total_amount=Sum('total_amount'))

        # Fetch data for Invoices
        invoice_data = invoice.objects.values('company_name').annotate(total_amount=Sum('total_amount'))

        # Fetch data for Lead Management
        lead_data1 = lead_management.objects.values('typeoflead').annotate(count=Count('id'))

        # Lead type chart data (filtered by enquirydate range)
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            filtered_leads = lead_management.objects.filter(enquirydate__range=[start_date, end_date])
        else:
            filtered_leads = lead_management.objects.all()

        lead_data = {
            "labels": ["Hot", "Warm", "Cold", "NotInterested", "LossOfOrder"],
            "datasets": [{
                "data": [
                    filtered_leads.filter(typeoflead='Hot').count(),
                    filtered_leads.filter(typeoflead='Warm').count(),
                    filtered_leads.filter(typeoflead='Cold').count(),
                    filtered_leads.filter(typeoflead='NotInterested').count(),
                    filtered_leads.filter(typeoflead='LossofOrder').count()
                ],
                "backgroundColor": ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF'],
            }]
        }
        lead_data_json = json.dumps(lead_data)

    

        # Follow-up remark chart data (from main_followup model)
        start_date_followup = request.GET.get('start_date_followup')
        end_date_followup = request.GET.get('end_date_followup')

        # Filter main_followup records by enquirydate if date range is provided
         # Adjust import according to your app structure

        followup_queryset = main_followup.objects.all()
        if start_date_followup and end_date_followup:
            try:
                start_dt_followup = datetime.strptime(start_date_followup, "%Y-%m-%d")
                end_dt_followup = datetime.strptime(end_date_followup, "%Y-%m-%d")
                # Assuming main_followup has a date field like enquirydate or followup_date to filter by
                followup_queryset = followup_queryset.filter(enquirydate__range=[start_dt_followup, end_dt_followup])
            except ValueError:
                pass  # Ignore if date parsing fails

        # Count by followup_remark choices
        followup_counts = followup_queryset.values('followup_remark').annotate(count=Count('id'))
        # Prepare labels and counts based on your choices order
        followup_labels = [
            'Call not received',
            'Give next Follow up date',
            'Call Out of Coverage Area',
        ]
        # Create data list aligned with labels order
        followup_data_list = []
        for label in followup_labels:
            entry = next((item for item in followup_counts if item['followup_remark'] == label), None)
            followup_data_list.append(entry['count'] if entry else 0)

        # Bar chart data for products by category
        pest_control_count = Product.objects.filter(category='Pest Control').count()
        fumigation_count = Product.objects.filter(category='Fumigation').count()
        product_sell_count = Product.objects.filter(category='Product Sell').count()

        bar_chart_data = {
            "labels": ['Pest Control', 'Fumigation', 'Product Sell'],
            "datasets": [{
                "label": "Number of Products per Category",
                "data": [pest_control_count, fumigation_count, product_sell_count],
                "backgroundColor": ['#FF6384', '#36A2EB', '#FFCE56'],
            }]
        }

        # Contract Type Distribution chart filtering
        start_date_service = request.GET.get('start_date_service')
        end_date_service = request.GET.get('end_date_service')

        if start_date_service and end_date_service:
            start_date_service_obj = datetime.strptime(start_date_service, "%Y-%m-%d")
            end_date_service_obj = datetime.strptime(end_date_service, "%Y-%m-%d")
            service_data = service_management.objects.filter(
                Q(service_date__gte=start_date_service_obj) & Q(service_date__lte=end_date_service_obj)
            ).values("contract_type").annotate(count=Count("id")).order_by("contract_type")
        else:
            service_data = service_management.objects.values("contract_type").annotate(count=Count("id")).order_by("contract_type")

        labellist = [entry["contract_type"] for entry in service_data]
        countlist = [entry["count"] for entry in service_data]

        # Quotation and Invoice counts filtered by date and type
        start_date_qo = request.GET.get('start_date_qo')
        end_date_qo = request.GET.get('end_date_qo')
        filter_type = request.GET.get('filter_type')

        quotations_count = 0
        orders_count = 0

        if start_date_qo and end_date_qo:
            start_date_obj = datetime.strptime(start_date_qo, "%Y-%m-%d")
            end_date_obj = datetime.strptime(end_date_qo, "%Y-%m-%d")
            if filter_type == 'quotation':
                quotations_count = quotation.objects.filter(
                    Q(quotation_date__gte=start_date_obj) & Q(quotation_date__lte=end_date_obj)
                ).count()
                orders_count = invoice.objects.count()
            elif filter_type == 'invoice':
                orders_count = invoice.objects.filter(
                    Q(date__gte=start_date_obj) & Q(date__lte=end_date_obj)
                ).count()
                quotations_count = quotation.objects.count()
        else:
            quotations_count = quotation.objects.count()
            orders_count = invoice.objects.count()

        start_date = request.GET.get("start_date_order")
        end_date = request.GET.get("end_date_order")
        followups = main_followup.objects.all()
        # print("followups",followups)

        if start_date and end_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d").date()
                end = datetime.strptime(end_date, "%Y-%m-%d")+ timedelta(days=1) 
                followups = followups.filter(created_at__range=(start, end))
                # print("followup_filter", followups)
            except ValueError:
                pass  

        # Create labels and data
        order_status_labels = ['Close Win', 'Close Loss', 'Not Closed']
        queryset = followups.values('order_status').annotate(count=Count('id'))

        order_status_data = []
        for label in order_status_labels:
            match = next((item for item in queryset if item['order_status'] == label), None)
            order_status_data.append(match['count'] if match else 0)


        context = {
            'lead_data': lead_data_json,
            'service_data': service_data,
            'quotation_data': quotation_data,
            'invoice_data': invoice_data,
            'labellist': json.dumps(labellist),
            'countlist': json.dumps(countlist),
            "quotationlist": json.dumps(["Quotations", "Orders"]),
            "order": json.dumps([quotations_count, orders_count]),
            'bar_chart_data': json.dumps(bar_chart_data),


            "order_status_labels": order_status_labels,
            "order_status_data": order_status_data,
            # Follow-up chart context
            'follow_up_labels': json.dumps(followup_labels),
            'follow_up_data': json.dumps(followup_data_list),
        }

        return render(request, 'index.html', context)

def generate_customerid(fullname):
    names = fullname.strip().split()

    # Extract first and last name safely
    firstname = names[0] if names else ''
    lastname = names[-1] if len(names) > 1 else ''

    # Get first 3 letters (or less if name is short)
    first_part = firstname[:3].upper().ljust(3, 'X')
    last_part = lastname[:3].upper().ljust(3, 'X')

    # Generate 4-digit random number
    random_number = str(random.randint(1000, 9999))

    return first_part + last_part + random_number


@role_required(['admin', 'sales'])
def get_customer_name(request):
    customer_id = request.GET.get('customer_id', None)
    if customer_id:
        try:
            customer = customer_details.objects.get(customerid=customer_id)
            customer_name = f"{customer.firstname} {customer.lastname}"
            return JsonResponse({'customer_name': customer_name})
        except customer_details.DoesNotExist:
            return JsonResponse({'error': 'Customer not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)



def signup(request):
    if request.method == "GET":
        return render(request, 'signup.html')
    else:
        uname = request.POST['uname']
        uemail = request.POST['uemail']
        upass = request.POST['upass']
        cpass = request.POST['cpass']
        security_key = request.POST.get('security_key', '')  # Retrieve the security key from the form

        # Check if fields are empty
        if uname == "" or upass == "" or cpass == "" or security_key == "":
            context = {'msg1': 'Field can not be empty'}
            return render(request, "signup.html", context)

        # Check if passwords match
        elif upass != cpass:
            context = {'msg2': 'Password and Confirm Password should be same'}
            return render(request, "signup.html", context)
        
        # Check if the security key is correct
        elif security_key != settings.SECURITY_KEY:
            context = {'msg2': 'Invalid security key'}
            return render(request, "signup.html", context)
        
        else:
            # Create the user
            u = User.objects.create(username=uname, email=uemail)
            u.set_password(upass)
            u.save()
            context = {'msg3': 'User Registered Successfully'}
            return render(request, "signup.html", context)


# def user_login(request):
#     if request.method == "GET":
#         return render(request, 'login.html')
    
#     else:
#         uname = request.POST.get('uname') 
#         upass = request.POST.get('upass')

#         u = authenticate(username=uname, password=upass,  )

#         if u is not None:
#             login(request, u)
#             start_date = request.GET.get('start_date')
#             end_date = request.GET.get('end_date')

#             latest_followups_qs = main_followup.objects.values('lead').annotate(
#                 latest_id=Max('id')
#             ).values_list('latest_id', flat=True)

#             if start_date and end_date:
#                 filtered_leads = main_followup.objects.filter(
#                     id__in=latest_followups_qs,
#                     lead__enquirydate__range=[start_date, end_date]
#                 )
#             else:
#                 filtered_leads = main_followup.objects.filter(id__in=latest_followups_qs)

#             # Prepare lead type chart data
#             lead_data = {
#             "labels": ["Hot", "Warm", "Cold", "NotInterested", "LossOfOrder"],
#             "datasets": [{
#             "data": [
#             filtered_leads.filter(typeoflead='Hot').count(),
#             filtered_leads.filter(typeoflead='Warm').count(),
#             filtered_leads.filter(typeoflead='Cold').count(),
#             filtered_leads.filter(typeoflead='NotInterested').count(),
#             filtered_leads.filter(typeoflead='LossofOrder').count()
#              ],
#                 "backgroundColor": ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF'],
#             }]
#             }


#             # Serialize the data to JSON
#             lead_data_json = json.dumps(lead_data)

#             # Retrieve the start and end dates from query parameters
#             start_date_followup = request.GET.get('start_date_followup')
#             end_date_followup = request.GET.get('end_date_followup')

#             # Default to today's date if no date range is provided
#             if start_date_followup:
#                 start_date_followup = datetime.strptime(start_date_followup, "%Y-%m-%d")
#             if end_date_followup:
#                 end_date_followup = datetime.strptime(end_date_followup, "%Y-%m-%d")

#             # Define the stages
#             stages = {
#                 1: "No Follow-Up Done",
#                 2: "First Follow-Up Done",
#                 3: "Second Follow-Up Done",
#                 4: "Third Follow-Up Done",
#                 5: "Final Follow-Up Done"
#             }

#             # Filter the lead_management table by the date range if provided
#             if start_date_followup and end_date_followup:
#                 stage_counts = lead_management.objects.filter(enquirydate__range=[start_date_followup, end_date_followup]) \
#                     .values('stage') \
#                     .annotate(count=Count('id')) \
#                     .order_by('stage')
#             else:
#                 # If no date range, fetch all data
#                 stage_counts = lead_management.objects.values('stage') \
#                     .annotate(count=Count('id')) \
#                     .order_by('stage')

#             # Prepare the labels and data for the chart


#             pest_control_count = Product.objects.filter(category='Pest Control').count()
#             fumigation_count = Product.objects.filter(category='Fumigation').count()
#             product_sell_count = Product.objects.filter(category='Product Sell').count()

#             # Bar chart data
#             bar_chart_data = {
#                 "labels": ['Pest Control', 'Fumigation', 'Product Sell'],
#                 "datasets": [{
#                     "label": "Number of Products per Category",
#                     "data": [pest_control_count, fumigation_count, product_sell_count],
#                     "backgroundColor": ['#FF6384', '#36A2EB', '#FFCE56'],
#                 }]
#             }

    
#             # Extract service-specific date filters
#             start_date_service = request.GET.get('start_date_service')
#             end_date_service = request.GET.get('end_date_service')

#             # Filter service management data by service_date range if present
#             if start_date_service and end_date_service:
#                 start_date_service_obj = datetime.strptime(start_date_service, "%Y-%m-%d")
#                 end_date_service_obj = datetime.strptime(end_date_service, "%Y-%m-%d")

#                 # Apply date filter on service_date field for contract type distribution
#                 service_data = service_management.objects.filter(
#                     Q(service_date__gte=start_date_service_obj) & Q(service_date__lte=end_date_service_obj)
#                 ).values("contract_type").annotate(count=Count("id")).order_by("contract_type")
#             else:
#                 # Default data if no service date filter is applied
#                 service_data = service_management.objects.values("contract_type").annotate(count=Count("id")).order_by("contract_type")

#             # Prepare data for the Contract Type Distribution chart
#             labellist = [entry["contract_type"] for entry in service_data]
#             countlist = [entry["count"] for entry in service_data]

#             # Extract query parameters
#             start_date_qo = request.GET.get('start_date_qo')
#             end_date_qo = request.GET.get('end_date_qo')
#             filter_type = request.GET.get('filter_type')

#             # Initialize counts
#             quotations_count = 0
#             orders_count = 0

#             # Apply date filters if present
#             if start_date_qo and end_date_qo:
#                 # Parse dates
#                 start_date_obj = datetime.strptime(start_date_qo, "%Y-%m-%d")
#                 end_date_obj = datetime.strptime(end_date_qo, "%Y-%m-%d")

#                 if filter_type == 'quotation':
#                     # Filter quotations by date range
#                     quotations_count = quotation.objects.filter(
#                         Q(quotation_date__gte=start_date_obj) & Q(quotation_date__lte=end_date_obj)
#                     ).count()
#                     orders_count = invoice.objects.count()  # Unfiltered
#                 elif filter_type == 'invoice':
#                     # Filter invoices by date range
#                     orders_count = invoice.objects.filter(
#                         Q(date__gte=start_date_obj) & Q(date__lte=end_date_obj)
#                     ).count()
#                     quotations_count = quotation.objects.count()  # Unfiltered
#             else:
#                 # Default counts without filters
#                 quotations_count = quotation.objects.count()
#                 orders_count = invoice.objects.count()
            
            
#             # Prepare context
#             context = {
#                 # 'total_leads': leads.count(),
#                 # 'hot_leads': hot_leads,
#                 # 'warm_leads': warm_leads,
#                 # 'cold_leads': cold_leads,
#                 # 'not_interested': not_interested,
#                 # 'loss_of_order': loss_of_order,
#                 'lead_data': lead_data_json,
#                 'service_data': service_data,
#                 # 'quotation_data': quotation_data,
#                 # 'invoice_data': invoice_data,
#                 # 'lead_data1': lead_data1,
#                 'labellist': json.dumps(labellist),  # Serialize labels
#                 'countlist': json.dumps(countlist),  # Serialize counts
#                 "quotationlist": json.dumps(["Quotations", "Orders"]),
#                 "order": json.dumps([quotations_count, orders_count]),
#                 'lead_data': json.dumps(lead_data),
#                 'bar_chart_data': json.dumps(bar_chart_data),
                
#             }
            
#             # return render(request, 'index.html', context)
#             return redirect('index')


#             # return render(request, "index.html")
        

#         else:
#             context = {'msg1': 'Wrong Username Or Password'}
#             return render(request, "login.html", context)

def user_login(request):
    if request.method == "GET":
        return render(request, 'login.html')

    uname = request.POST.get('uname') 
    upass = request.POST.get('upass')
    user = authenticate(username=uname, password=upass)

    if user:
        try:
            role = user.userprofile.role
            if role not in ['admin', 'sales']:
                messages.error(request, "Only admin and sales users are allowed to log in.")
                return render(request, 'login.html')

            # Login only if role is valid
            login(request, user)


            # return render(request, "index.html", context)
            return redirect('index')

        except UserProfile.DoesNotExist:
            messages.error(request, "User profile not found.")
            return render(request, 'login.html')
    else:
        return render(request, "login.html", {'msg1': 'Wrong Username Or Password'})


@login_required
def user_logout(request):
    logout(request)
    return redirect("/")

# Add Sales Person
@login_required
@role_required(['admin','branch_manager'])
def add_sales_person(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        date_of_birth = parse_date(request.POST.get('date_of_birth'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(id = branch_id)
        co_ordinator = request.POST.get('co_ordinator') == True
        # Create Django User (username as email or phone, password default or random)
        username = mobile_no
        user = User.objects.create_user(username=username, password=password, email=email, first_name=full_name)
        user.is_staff = True
        user.save()
        

        # Assign role to user via UserProfile
        user_profile = user.userprofile  # auto-created by the signal
        user_profile.role = 'sales'
        user_profile.phone = mobile_no
        user_profile.save()

        # Create SalesPerson record
        SalesPerson.objects.create(
            full_name=full_name,
            date_of_joining=date_of_joining,
            mobile_no=mobile_no,
            email=email,
            date_of_birth=date_of_birth,
            branch = branch,
            co_ordinator = co_ordinator
            
        )

        return redirect('sales_person_list')
    branches = Branch.objects.all()
    return render(request, 'add_sales_person.html',{"branches":branches})

# List Sales Persons
@login_required
@role_required(['admin', 'branch_manager'])
def sales_person_list(request):
    if request.user.userprofile.role == 'admin':
        sales_persons = SalesPerson.objects.all()
    elif request.user.userprofile.role == 'branch_manager':
        branch_manager = BranchManager.objects.get(mobile_no = request.user.username)
        sales_persons = SalesPerson.objects.filter(branch=branch_manager.branch)
    
    return render(request, 'sales_person_list.html', {'sales_persons': sales_persons})

@login_required
@role_required(['admin', 'branch_manager'])
def export_sales_person_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sales_persons.csv"'

    writer = csv.writer(response)
    
    # Header row
    writer.writerow(['Full Name', 'Date of Joining', 'Mobile No', 'Email', 'Date of Birth'])

    # Data rows
    for sp in SalesPerson.objects.all():
        writer.writerow([
            sp.full_name,
            sp.date_of_joining.strftime('%Y-%m-%d'),
            sp.mobile_no,
            sp.email,
            sp.date_of_birth.strftime('%Y-%m-%d'),
        ])

    return response


# Edit Sales Person
@login_required
@role_required(['admin','branch_manager'])
def edit_sales_person(request, pk):
    person = get_object_or_404(SalesPerson, pk=pk)

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        date_of_birth = parse_date(request.POST.get('date_of_birth'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(id = branch_id)
        co_ordinator = request.POST.get('co_ordinator')
        # Update SalesPerson
        person.full_name = full_name
        person.date_of_joining = date_of_joining
        person.mobile_no = mobile_no
        person.email = email
        person.date_of_birth = date_of_birth
        person.branch = branch
        person.co_ordinator = co_ordinator
        person.save()

        # Check for existing user
        user = User.objects.filter(email=email).first()

        if user:
            # Update existing user
            user.first_name = full_name
            user.username = mobile_no
            user.email = email
            if password:
                user.set_password(password)
            user.save()
        else:
            # Create new user
            user = User.objects.create_user(
                username=mobile_no,
                email=email,
                password=password or User.objects.make_random_password(),
                first_name=full_name
            )
        user.is_staff = True
        user.save()

        # Update or create UserProfile
        user_profile, created = UserProfile.objects.get_or_create(user=user)
        user_profile.role = 'sales'
        user_profile.phone = mobile_no
        user_profile.save()

        return redirect('sales_person_list')
    branches = Branch.objects.all()
    return render(request, 'edit_sales_person.html', {'person': person, 'branches':branches})

@login_required
@role_required(['admin'])
def delete_sales_person(request, pk):
    person = get_object_or_404(SalesPerson, pk=pk)

    if request.method == 'POST':
         # Delete the User whose username is the BranchManager's mobile number
        try:
            user = User.objects.get(username=person.mobile_no)
            user.delete()  # This will also delete UserProfile if CASCADE
        except User.DoesNotExist:
            pass  # User not found, just continue
        person.delete()
        return redirect('sales_person_list')

    return render(request, 'delete_sales_person.html', {'person': person})


@login_required
@role_required(['admin'])
def add_branch_manager(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        date_of_birth = parse_date(request.POST.get('date_of_birth'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch_instance = Branch.objects.get(pk=branch_id)

        # Create Django User (username as email or phone, password default or random)
        username = mobile_no
        user = User.objects.create_user(username=username, password=password, email=email, first_name=full_name)
        user.is_staff = True
        user.save()
        

        # Assign role to user via UserProfile
        user_profile = user.userprofile  # auto-created by the signal
        user_profile.role = 'branch_manager'
        user_profile.phone = mobile_no
        user_profile.save()

        # Create SalesPerson record
        BranchManager.objects.create(
            full_name=full_name,
            date_of_joining=date_of_joining,
            mobile_no=mobile_no,
            email=email,
            date_of_birth=date_of_birth,
            branch=branch_instance
            
        )

        return redirect('branch_manager_list')
    
    branch = Branch.objects.all()
    return render(request, 'add_branch_manager.html',{'branch': branch})


@login_required
@role_required(['admin'])
def branch_manager_list(request):
    persons = BranchManager.objects.all()
    return render(request, 'branch_manager_list.html', {'persons': persons})


@login_required
@role_required(['admin'])
def edit_branch_manager(request, pk):
    person = get_object_or_404(BranchManager, pk=pk)

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        date_of_birth = parse_date(request.POST.get('date_of_birth'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch_instance = Branch.objects.get(pk=branch_id)
        # Update SalesPerson
        person.full_name = full_name
        person.date_of_joining = date_of_joining
        person.mobile_no = mobile_no
        person.email = email
        person.date_of_birth = date_of_birth
        person.branch = branch_instance
        person.save()

        # Check for existing user
        user = User.objects.filter(email=email).first()

        if user:
            # Update existing user
            user.first_name = full_name
            user.username = mobile_no
            user.email = email
            if password:
                user.set_password(password)
            user.save()
        else:
            # Create new user
            user = User.objects.create_user(
                username=mobile_no,
                email=email,
                password=password or User.objects.make_random_password(),
                first_name=full_name
            )
        user.is_staff = True
        user.save()

        # Update or create UserProfile
        user_profile, created = UserProfile.objects.get_or_create(user=user)
        user_profile.role = 'branch_manager'
        user_profile.phone = mobile_no
        user_profile.save()

        return redirect('branch_manager_list')
    branch = Branch.objects.all()
    return render(request, 'edit_branch_manager.html', {'person': person, 'branch':branch})

@login_required
@role_required(['admin'])
def delete_branch_manager(request, pk):
    person = get_object_or_404(BranchManager, pk=pk)

    if request.method == 'POST':
        # Delete the User whose username is the BranchManager's mobile number
        try:
            user = User.objects.get(username=person.mobile_no)
            user.delete()  # This will also delete UserProfile if CASCADE
        except User.DoesNotExist:
            pass  # User not found, just continue

        # Then delete the BranchManager object
        person.delete()

        return redirect('branch_manager_list')

    return redirect('branch_manager_list')

@login_required
@role_required(['admin','branch_manager'])
def add_operation_person(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        date_of_birth = parse_date(request.POST.get('date_of_birth'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(id = branch_id)
        # Create Django User (username as email or phone, password default or random)
        username = mobile_no
        user = User.objects.create_user(username=username, password=password, email=email, first_name=full_name)
        user.is_staff = True
        user.save()
        

        # Assign role to user via UserProfile
        user_profile = user.userprofile  # auto-created by the signal
        user_profile.role = 'operation_person'
        user_profile.phone = mobile_no
        user_profile.save()

        # Create SalesPerson record
        OperationPerson.objects.create(
            user = user,
            full_name=full_name,
            date_of_joining=date_of_joining,
            mobile_no=mobile_no,
            email=email,
            date_of_birth=date_of_birth,
            branch = branch,
           
            
        )

        return redirect('operation_person_list')
    branches = Branch.objects.all()
    return render(request, 'add_operation_person.html',{"branches":branches})

@login_required
@role_required(['admin','branch_manager'])
def operation_person_list(request):
    u = request.user
    if request.user.userprofile.role == 'admin':
        persons = OperationPerson.objects.all()
    elif request.user.userprofile.role == 'branch_manager':
        branch_manager = BranchManager.objects.get(mobile_no = request.user.username)
        persons = OperationPerson.objects.filter(branch=branch_manager.branch)
    return render(request, 'operation_person_list.html', {'persons': persons})

@login_required
@role_required(['admin','branch_manager'])
def edit_operation_person(request, pk):
    person = get_object_or_404(OperationPerson, pk=pk)

    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        date_of_birth = parse_date(request.POST.get('date_of_birth'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch_instance = Branch.objects.get(pk=branch_id)
        # Update SalesPerson
        person.full_name = full_name
        person.date_of_joining = date_of_joining
        person.mobile_no = mobile_no
        person.email = email
        person.date_of_birth = date_of_birth
        person.branch = branch_instance
        person.save()

        # Check for existing user
        user = User.objects.filter(email=email).first()

        if user:
            # Update existing user
            user.first_name = full_name
            user.username = mobile_no
            user.email = email
            if password:
                user.set_password(password)
            user.save()
        else:
            # Create new user
            user = User.objects.create_user(
                username=mobile_no,
                email=email,
                password=password or User.objects.make_random_password(),
                first_name=full_name
            )
        user.is_staff = True
        user.save()
        person.user = user

        # Update or create UserProfile
        user_profile, created = UserProfile.objects.get_or_create(user=user)
        user_profile.role = 'operation_person'
        user_profile.phone = mobile_no
        user_profile.save()

        return redirect('operation_person_list')
    branch = Branch.objects.all()
    return render(request, 'edit_operation_person.html', {'person': person, 'branches':branch})

@login_required
@role_required(['admin','branch_manager'])
def delete_operation_person(request, pk):
    operation_person = get_object_or_404(OperationPerson, pk=pk)

    # Also delete linked User (to avoid orphan user accounts)
    user = operation_person.user

    # Delete OperationPerson first
    operation_person.delete()

    # Then delete user account
    user.delete()

    messages.success(request, "Operation Person deleted successfully.")
    return redirect('operation_person_list')

@login_required
def customer_details_create(request):
    if request.method == 'GET':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            primarycontact = request.GET.get('primarycontact')
            lead = lead_management.objects.filter(primarycontact=primarycontact).order_by('-enquirydate').first()
            if lead:
                data = {
                    'fullname': lead.customername or '',
                    'primaryemail': lead.customeremail or '',
                    'secondarycontact': lead.secondarycontact or '',
                    'contactperson': (
                        lead.salesperson.full_name if lead.salesperson 
                        else lead.branch_manager.full_name if getattr(lead, 'branch_manager', None)
                        else ''
                    ),
                    'contactedby': lead.contactedby or '',
                    'shifttopartyaddress': lead.customeraddress or '',
                    'shifttopartycity': lead.city or '',
                    'soldtopartyaddress': lead.customeraddress or '',
                    'soldtopartycity': lead.city or '',
                    'customer_type':lead.customer_type or '',
                    'or_name':lead.or_name or '',
                    'or_contact': lead.or_contact or None,
                    'branch': lead.branch_id or None,
                }
                print(data)

                return JsonResponse({'status': 'exists', 'data': data})
            return JsonResponse({'status': 'not_found'})
        branches = Branch.objects.all()
        return render(request, 'customer_details.html',{'branches':branches})
    elif request.method == 'POST':
            fullname = request.POST.get('fullname', '').strip()
            primarycontact = request.POST.get('primarycontact', '').strip()

            # fallback email if empty
            primaryemail = request.POST.get('primaryemail') or "test@sample.com"
            secondaryemail = request.POST.get('secondaryemail', '')
            secondarycontact = request.POST.get('secondarycontact') or None
            contactperson = request.POST.get('contactperson', '')
            designstion = request.POST.get('designstion', '')
            shifttopartyaddress = request.POST.get('shifttopartyaddress', '')
            shifttopartycity = request.POST.get('shifttopartycity', '')
            shifttopartystate = request.POST.get('shifttopartystate', '')
            shifttopartypostal = request.POST.get('shifttopartypostal', '')
            soldtopartyaddress = request.POST.get('soldtopartyaddress', '')
            soldtopartycity = request.POST.get('soldtopartycity', '')
            soldtopartystate = request.POST.get('soldtopartystate', '')
            soldtopartypostal = request.POST.get('soldtopartypostal', '')
            customer_type = request.POST.get('customer_type', 'Individual')
            or_name = request.POST.get('or_name', '')
            or_contact = request.POST.get('or_contact') or None
            branch = Branch.objects.get(id = request.POST.get("branch"))
            # validate required fields
            if not fullname or not primarycontact:
                return render(request, "customer_details.html", {
                    'msg1': 'Full Name and Primary Contact are required'
                })

            customerid = generate_customerid(fullname)

            customer_details.objects.create(
                fullname=fullname,
                primaryemail=primaryemail,
                secondaryemail=secondaryemail,
                primarycontact=primarycontact,
                secondarycontact=secondarycontact,
                contactperson=contactperson,
                designation=designstion,
                shifttopartyaddress=shifttopartyaddress,
                shifttopartycity=shifttopartycity,
                shifttopartystate=shifttopartystate,
                shifttopartypostal=shifttopartypostal,
                soldtopartyaddress=soldtopartyaddress,
                soldtopartycity=soldtopartycity,
                soldtopartystate=soldtopartystate,
                soldtopartypostal=soldtopartypostal,
                customerid=customerid,
                customer_type=customer_type,
                or_name=or_name,
                or_contact=or_contact,
                branch = branch
            )

            # Conditional redirect
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('/display_customer')
    

@login_required
def export_customer_excel(request):

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="customer_list.csv"'

    writer = csv.writer(response)

    # Get all model field names
    model_fields = [field.name for field in customer_details._meta.fields]

    # Write header row
    writer.writerow([field.replace('_', ' ').title() for field in model_fields])

    # Write data rows
    for obj in customer_details.objects.all().values_list(*model_fields):
        writer.writerow(obj)

    return response


@login_required
def product_list(request):
    # 🏷️ Get filters
    category_filter = request.GET.get('category', 'all')
    search_query = request.GET.get('search', '').strip()

    # 🧭 Base queryset
    products = Product.objects.all()

    # 🔍 Search filter
    if search_query:
        products = products.filter(
            Q(product_name__icontains=search_query) |
            Q(hsn_code__icontains=search_query) |
            Q(category__icontains=search_query)
        )

    # 🏷️ Category filter
    if category_filter and category_filter != 'all':
        products = products.filter(category=category_filter)

    # 📄 Pagination
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 🗂️ Category list
    categories = [choice[0] for choice in Product.CATEGORY_CHOICES]

    # 🧩 Context
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'selected_category': category_filter,
        'search_query': search_query,
        'product_count': paginator.count,
    }
    return render(request, 'product_list.html', context)

@login_required
def export_product_list_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_list.csv"'

    writer = csv.writer(response)
    
    # Header row
    writer.writerow(['Name', 'Category'])

    # Data rows
    for p in Product.objects.all():
        writer.writerow([
            p.product_name,
            p.category
        ])

    return response

@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)
    product.delete()
    return redirect('/products')


# views.py - Add this function
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import ProductRequiredItem

@login_required
def get_product_items(request):
    """API to get items for a specific product"""
    product_id = request.GET.get('product_id')
    
    if not product_id:
        return JsonResponse({'error': 'Product ID required', 'success': False}, status=400)
    
    try:
        items = ProductRequiredItem.objects.filter(product_id=product_id).values('id', 'item_name')
        return JsonResponse({
            'items': list(items),
            'success': True
        })
    except Exception as e:
        return JsonResponse({'error': str(e), 'success': False}, status=500)

@login_required
@role_required(['admin','sales', 'branch_manager'])
def service_management_create(request):
    customers = customer_details.objects.all()
    category_choices = Product.CATEGORY_CHOICES
    products = Product.objects.all()
    sales_persons = list(SalesPerson.objects.all()) + list(BranchManager.objects.all())
    branch = Branch.objects.all()
    frequency_choices = [str(i) for i in range(1, 13)] + ['Fortnight', 'Weekly', 'Daily']
    segments = service_management._meta.get_field('segment').choices
    if request.method == 'POST':
        try:
            customer_contact = request.POST['customer_contact']
            customer = customer_details.objects.get(primarycontact=customer_contact)
            address = request.POST.get('address', 'Null')
            lead_date = request.POST.get('lead_date')
            service_date = request.POST.get('service_date')

            lead_date = datetime.strptime(lead_date, '%Y-%m-%d').date() if lead_date else None
            service_date = datetime.strptime(service_date, '%Y-%m-%d').date() if service_date else None

            total_price = float(request.POST.get('total_price', 0) or 0)
            total_with_gst = float(request.POST.get('total_with_gst', 0) or 0)
            total_gst = float(request.POST.get('gst_price', 0) or 0)
            latitude = request.POST.get('latitude')
            longitude = request.POST.get('longitude')

            apply_gst = request.POST.get('apply_gst') == 'on'
            gst_status = 'GST' if apply_gst else 'NON-GST'
            if not apply_gst:
                total_with_gst = total_price

            delivery_time = request.POST.get('delivery_time', timezone.now().time())
            branch_id = request.POST.get('branch')
            branch_instance = Branch.objects.get(id=branch_id) if branch_id else None
            # Create service instance
            instance = service_management.objects.create(
                status='pending',
                customer=customer,
                address=address,
                total_price=total_price,
                total_price_with_gst=total_with_gst,
                total_charges = total_gst,
                service_subject = request.POST.get('subject'),
                contract_type=request.POST.get('contract_type', 'NOT SELECTED'),
                contract_status=request.POST.get('contract_status', 'NOT SELECTED'),
                segment = request.POST.get('segments'),
                property_type=request.POST.get('property_type',''),
                warranty_period=request.POST.get('warranty_period',''),
                state=request.POST.get('state', 'Null'),
                city=request.POST.get('city', 'Null'),
                pincode=request.POST.get('pincode', '000000'),
                latitude=latitude,
                longitude=longitude,
                gps_location=request.POST.get('gps_location'),
                
                frequency_count=request.POST.get('frequency_count', 'NOT SELECTED'),
                payment_terms=request.POST.get('payment_terms', '100% Advance payment OR Whatever mutually Decided'),
                sales_person_name=request.POST.get('sales_person_name'),
                sales_person_contact_no=request.POST.get('sales_person_contact_no'),
                sales_person_email=request.POST.get('sales_person_email'),
                delivery_time=delivery_time,
                lead_date=lead_date,
                service_date=service_date,
                gst_status=gst_status,
                branch=branch_instance,
            )

            # Get products from JSON string
            # Get products from JSON string
            # Get products from JSON string
            selected_products_json = request.POST.get('selected_products_json', '[]')
            print("=" * 50)
            print("RAW selected_products_json:", selected_products_json)
            print("Type of selected_products_json:", type(selected_products_json))
            
            try:
                selected_products = json.loads(selected_products_json)
                print("Parsed selected_products:", selected_products)
                print("Number of products:", len(selected_products))
                print("Type of selected_products:", type(selected_products))
            except json.JSONDecodeError as e:
                print(f"JSON Decode Error: {e}")
                selected_products = []
            
            # Loop through product entries
            for idx, item in enumerate(selected_products):
                print(f"\n--- Processing Product {idx + 1} ---")
                print(f"Item data: {item}")
                
                product_id = item.get('p_id')
                price = item.get('price')
                quantity = item.get('quantity')
                gst_percentage = item.get('gst', 0)
                description = item.get('description', '')
                
                print(f"Product ID: {product_id}")
                print(f"Price: {price}")
                print(f"Quantity: {quantity}")
                print(f"GST: {gst_percentage}")
                print(f"Description: {description}")
                
                if not product_id or price is None or quantity is None:
                    print(f"Skipping - Missing data: product_id={product_id}, price={price}, quantity={quantity}")
                    continue
                
                try:
                    product = Product.objects.get(product_id=product_id)
                    print(f"Found product: {product.product_name}")
                except Product.DoesNotExist:
                    print(f"Product with ID {product_id} not found!")
                    continue
                
                # Create ServiceProduct
                service_product = ServiceProduct.objects.create(
                    service=instance,
                    product=product,
                    price=Decimal(price),
                    quantity=Decimal(quantity),
                    gst_percentage=Decimal(gst_percentage) or Decimal('0.00'),
                    description=description,
                )
                print(f"Created ServiceProduct with ID: {service_product.id}")
                
                # Save selected items for this service product
                selected_items = item.get('items', [])
                print(f"Number of selected items: {len(selected_items)}")
                
                for s_idx, selected_item in enumerate(selected_items):
                    print(f"  Item {s_idx + 1}: {selected_item}")
                    required_item_id = selected_item.get('required_item_id')
                    if required_item_id:
                        try:
                            required_item = ProductRequiredItem.objects.get(id=required_item_id)
                            print(f"    Found required item: {required_item.item_name}")
                            item_obj = ServiceProductItem.objects.create(
                                service_product=service_product,
                                required_item=required_item,
                                quantity=Decimal(selected_item.get('quantity', 1)),
                                notes=selected_item.get('notes', '')
                            )
                            print(f"    Created ServiceProductItem with ID: {item_obj.id}")
                        except ProductRequiredItem.DoesNotExist:
                            print(f"    Item with ID {required_item_id} not found")
                            continue
                        except Exception as e:
                            print(f"    Error creating ServiceProductItem: {e}")
                            continue
            
            print("\n" + "=" * 50)
            print("All products processed successfully")
            print("=" * 50)
            
            return redirect('/display_service_management')

        except Exception as e:
            return render(request, 'service_management.html', {
                'error': str(e),
                'category_choices': category_choices,
                'products': products,
                'customers': customers,
                'sales_persons': sales_persons,
                'frequency_choices': frequency_choices,
            })

    return render(request, 'service_management.html', {
        'category_choices': category_choices,
        'products': products,
        'customers': customers,
        'sales_persons': sales_persons,
        'frequency_choices': frequency_choices,
        'segments':segments,
        'branches':branch
    })




# views.py - COMPLETE VERSION WITH ALL ACTIONS

# views.py - COMPLETE VERSION WITH AMC PRODUCTS API

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .models import service_management
from amc.models import AMCContract, AMCServiceSchedule, AMCServiceVisit
import json
from datetime import datetime

@login_required
def approve_service(request, id):
    """Approve main service - sets status to Approved"""
    service = get_object_or_404(service_management, id=id)
    service.status = 'Approved'
    service.save()
    return JsonResponse({'status': 'approved', 'success': True})

@login_required
def reject_service(request, id):
    """Reject main service - sets status to Rejected"""
    service = get_object_or_404(service_management, id=id)
    service.status = 'Rejected'
    service.save()
    return JsonResponse({'status': 'rejected', 'success': True})

@login_required
def complete_service(request, id):
    """Complete main service - sets status to Completed"""
    service = get_object_or_404(service_management, id=id)
    service.status = 'Completed'
    service.save()
    return JsonResponse({'status': 'completed', 'success': True})

@login_required
@csrf_exempt
def approve_amc_visit(request, schedule_id):
    """Approve individual AMC visit"""
    if request.method == 'POST':
        schedule = get_object_or_404(AMCServiceSchedule, id=schedule_id)
        schedule.is_completed = True
        schedule.save()
        
        visit = AMCServiceVisit.objects.filter(amc=schedule.amc, service_date=schedule.service_date).first()
        if visit:
            visit.allocation_status = 'ALLOCATED'
            visit.save()
        
        return JsonResponse({'success': True, 'message': 'AMC visit approved'})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)

@login_required
@csrf_exempt
def reject_amc_visit(request, schedule_id):
    """Reject individual AMC visit"""
    if request.method == 'POST':
        schedule = get_object_or_404(AMCServiceSchedule, id=schedule_id)
        schedule.is_completed = False
        schedule.save()
        
        visit = AMCServiceVisit.objects.filter(amc=schedule.amc, service_date=schedule.service_date).first()
        if visit:
            visit.allocation_status = 'CANCELLED'
            visit.save()
        
        return JsonResponse({'success': True, 'message': 'AMC visit rejected'})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)

@login_required
@csrf_exempt
def reschedule_amc_visit(request, schedule_id):
    """Reschedule an AMC visit to a new date"""
    if request.method == 'POST':
        data = json.loads(request.body)
        new_date = data.get('new_date')
        
        if new_date:
            new_date_obj = datetime.strptime(new_date, '%Y-%m-%d').date()
            schedule = get_object_or_404(AMCServiceSchedule, id=schedule_id)
            old_date = schedule.service_date
            schedule.service_date = new_date_obj
            schedule.save()
            
            visit = AMCServiceVisit.objects.filter(amc=schedule.amc, service_date=old_date).first()
            if visit:
                visit.service_date = new_date_obj
                visit.rescheduled_from = old_date
                visit.save()
            
            return JsonResponse({'success': True, 'new_date': str(new_date_obj)})
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required
@csrf_exempt
def get_amc_visit_details(request, schedule_id):
    """Get details of a specific AMC visit including products and required items"""
    schedule = get_object_or_404(AMCServiceSchedule, id=schedule_id)
    visit = AMCServiceVisit.objects.filter(amc=schedule.amc, service_date=schedule.service_date).first()
    
    # Get the main service (service_management) that this AMC belongs to
    main_service = schedule.amc.service
    
    # Get products with their required items for this AMC
    products = []
    
    # Get regular products from ServiceProduct (if any)
    for sp in main_service.service_products.all():
        product_data = {
            "name": sp.product.product_name,
            "type": "NORMAL",
            "quantity": float(sp.quantity) if sp.quantity else 1,
            "price": float(sp.price) if sp.price else 0,
            "gst_percentage": float(sp.gst_percentage) if sp.gst_percentage else 0,
            "items": []
        }
        
        for item in sp.selected_items.all():
            product_data["items"].append({
                "name": item.required_item.item_name,
                "qty": float(item.quantity),
                "notes": item.notes or ""
            })
        
        products.append(product_data)
    
    # Add AMC schedule as a product with its required items
    # Get any required items associated with this AMC visit
    amc_product_data = {
        "name": f"AMC Service - {schedule.amc.contract_number}",
        "type": "AMC",
        "service_date": schedule.service_date.strftime('%Y-%m-%d'),
        "schedule_id": schedule.id,
        "is_completed": schedule.is_completed,
        "items": []
    }
    
    # You can add custom items for AMC visits here if needed
    # For now, add the visit date as an item
    amc_product_data["items"].append({
        "name": f"Scheduled Visit on {schedule.service_date.strftime('%d %b, %Y')}",
        "qty": 1,
        "notes": "Regular AMC service visit"
    })
    
    products.append(amc_product_data)
    
    data = {
        'id': schedule.id,
        'contract_number': schedule.amc.contract_number,
        'service_date': schedule.service_date.strftime('%Y-%m-%d'),
        'service_date_display': schedule.service_date.strftime('%b %d, %Y'),
        'is_completed': schedule.is_completed,
        'amc_type': schedule.amc.amc_type,
        'customer_name': schedule.amc.customer.fullname if schedule.amc.customer else '',
        'customer_mobile': schedule.amc.customer.primarycontact if schedule.amc.customer else '',
        'customer_address': schedule.amc.default_customer_address or '',
        'technicians': [t.user.get_full_name() or t.user.username for t in schedule.amc.technicians.all()],
        'products': products,
        'visit_details': {
            'allocation_status': visit.allocation_status if visit else 'PENDING',
            'remarks': visit.remarks if visit else '',
            'crm_service_id': visit.crm_service.id if visit and visit.crm_service else None,
            'auto_allocation_done': visit.auto_allocation_done if visit else False
        } if visit else None,
        'contract_details': {
            'start_date': schedule.amc.start_date.strftime('%Y-%m-%d'),
            'end_date': schedule.amc.end_date.strftime('%Y-%m-%d'),
            'frequency': schedule.amc.frequency,
            'total_amount': float(schedule.amc.total_amount) if schedule.amc.total_amount else 0,
            'per_visit_amount': float(schedule.amc.per_visit_amount) if schedule.amc.per_visit_amount else 0,
            'service_description': schedule.amc.service_description or '',
            'notes': schedule.amc.notes or ''
        }
    }
    return JsonResponse({'success': True, 'data': data})

@login_required
@csrf_exempt
def get_amc_service_products(request, schedule_id):
    """Get products and required items for a specific AMC schedule"""
    schedule = get_object_or_404(AMCServiceSchedule, id=schedule_id)
    main_service = schedule.amc.service
    
    products = []
    
    # Get regular products
    for sp in main_service.service_products.all():
        product_data = {
            "name": sp.product.product_name,
            "type": "NORMAL",
            "quantity": float(sp.quantity) if sp.quantity else 1,
            "price": float(sp.price) if sp.price else 0,
            "gst_percentage": float(sp.gst_percentage) if sp.gst_percentage else 0,
            "items": []
        }
        
        for item in sp.selected_items.all():
            product_data["items"].append({
                "name": item.required_item.item_name,
                "qty": float(item.quantity),
                "notes": item.notes or ""
            })
        
        products.append(product_data)
    
    # Add AMC schedule product
    amc_product = {
        "name": f"AMC Visit - {schedule.amc.contract_number}",
        "type": "AMC",
        "service_date": schedule.service_date.strftime('%Y-%m-%d'),
        "schedule_id": schedule.id,
        "is_completed": schedule.is_completed,
        "items": [
            {
                "name": f"Visit Date: {schedule.service_date.strftime('%d %b, %Y')}",
                "qty": 1,
                "notes": f"AMC Type: {schedule.amc.amc_type}"
            }
        ]
    }
    products.append(amc_product)
    
    return JsonResponse({'success': True, 'products': products, 'count': len(products)})

def get_products_with_amc_schedules(service):
    """Get products for a service, automatically including AMC schedules"""
    products = []
    
    # First, add regular products from ServiceProduct
    for sp in service.service_products.all():
        product_data = {
            "name": sp.product.product_name,
            "type": "NORMAL",
            "quantity": float(sp.quantity),
            "price": float(sp.price),
            "gst_percentage": float(sp.gst_percentage),
            "total_with_gst": float(sp.total_with_gst) if sp.total_with_gst else None,
            "items": []
        }
        
        for item in sp.selected_items.all():
            product_data["items"].append({
                "name": item.required_item.item_name,
                "qty": float(item.quantity),
                "notes": item.notes or ""
            })
        
        products.append(product_data)
    
    # Check if this service has any AMC contracts
    amc_contracts = AMCContract.objects.filter(service=service)
    
    for amc in amc_contracts:
        schedules = AMCServiceSchedule.objects.filter(amc=amc).order_by('service_date')
        
        for schedule in schedules:
            visit = AMCServiceVisit.objects.filter(amc=amc, service_date=schedule.service_date).first()
            
            products.append({
                "name": f"AMC Service - {amc.contract_number}",
                "type": "AMC",
                "service_date": schedule.service_date,
                "schedule_id": schedule.id,
                "visit_id": visit.id if visit else None,
                "is_completed": schedule.is_completed,
                "amc_contract": amc.contract_number,
                "items": [
                    {
                        "name": f"Visit Date: {schedule.service_date}",
                        "qty": 1,
                        "status": "Completed" if schedule.is_completed else "Pending"
                    }
                ]
            })
    
    return products

def bharat_page(request):
    filter_type = request.GET.get("type", "all")
    
    services = service_management.objects.select_related('customer').prefetch_related(
        'service_products__product',
        'service_products__selected_items__required_item'
    ).order_by("-id")
    
    if filter_type == "amc":
        service_ids_with_amc = AMCContract.objects.values_list('service_id', flat=True).distinct()
        services = services.filter(id__in=service_ids_with_amc)
    
    grouped = {}
    
    for service in services:
        key = service.customer.primarycontact if service.customer else f"no_{service.id}"
        
        products_with_amc = get_products_with_amc_schedules(service)
        
        amc_sub_services = []
        amc_contracts = AMCContract.objects.filter(service=service)
        
        for amc in amc_contracts:
            schedules = AMCServiceSchedule.objects.filter(amc=amc).order_by('service_date')
            for schedule in schedules:
                amc_sub_services.append({
                    "id": schedule.id,
                    "schedule_id": schedule.id,
                    "type": "AMC",
                    "product": f"AMC Visit ({amc.contract_number})",
                    "service_date": schedule.service_date,
                    "service_date_display": schedule.service_date.strftime('%b %d, %Y'),
                    "status": "Completed" if schedule.is_completed else "Pending",
                    "contract_number": amc.contract_number,
                    "is_completed": schedule.is_completed
                })
        
        has_amc = len(amc_contracts) > 0
        
        main_service_data = {
            "id": service.id,
            "service_no": f"SRV-{service.id:04d}",
            "customer": service.customer.fullname if service.customer else "",
            "mobile": service.customer.primarycontact if service.customer else "",
            "status": service.status.capitalize() if service.status else "Pending",
            "date": service.service_date,
            "contract": "AMC" if has_amc else service.contract_type,
            "products": products_with_amc,
            "sub_services": amc_sub_services
        }
        
        if key not in grouped:
            grouped[key] = {"main": main_service_data, "subs": []}
        else:
            grouped[key]["subs"].append(main_service_data)
    
    total_services = service_management.objects.count()
    completed_count = service_management.objects.filter(status='Completed').count()
    approved_count = service_management.objects.filter(status='Approved').count()
    pending_count = service_management.objects.filter(status='Pending').count()
    rejected_count = service_management.objects.filter(status='Rejected').count()
    
    context = {
        "groups": grouped.values(),
        "total_entries": total_services,
        "completed_count": completed_count,
        "approved_count": approved_count,
        "pending_count": pending_count,
        "rejected_count": rejected_count,
        "selected_type": filter_type
    }
    
    return render(request, 'bharat.html', context)



from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import service_management

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime
from .models import quotation_management, customer_details, Product, QuotationTerm
from django.shortcuts import render, redirect
from django.http import JsonResponse
from datetime import datetime
from django.utils import timezone
from .models import Product, QuotationTerm, quotation_management, customer_details

from django.shortcuts import render
from django.utils import timezone
from django.http import JsonResponse
from datetime import datetime
from .models import quotation_management, QuotationTerm
from .models import Product
from .models import customer_details
from .models import Branch  



# New----------------
@login_required
@role_required(['admin','sales', 'branch_manager'])
def quotation_management_create(request):
    category_choices = Product.CATEGORY_CHOICES
    terms = QuotationTerm.objects.all() 
    branches = Branch.objects.all()
    products = Product.objects.all()
    sales_person_list =  list(SalesPerson.objects.all())+ list(BranchManager.objects.all())
    thank_notes_qs = quotation_management.objects.values_list('thank_u_note', flat=True).distinct()
    thank_notes = [note for note in thank_notes_qs if note]  
    
    if request.method == 'POST':
        custom_terms = request.POST.get('add_terms_conditions') or None
        customer_id = request.POST.get('customer_id')
        customer = None
        data = request.POST.copy()
        data['terms_and_conditions'] = request.POST.getlist('terms_and_conditions')
        customer_id = request.POST.get('customer_id')
        if customer_id:
            data['customer_id'] = customer_id

        request.session['quotation_form_data'] = data
        print("Session stored terms:", request.session['quotation_form_data'].get('product_json_data'))
        request.session.modified = True
        if customer_id:
            try:
                customer = customer_details.objects.get(id=customer_id)
                contact_no = request.POST.get('contact_no')
                customer = customer_details.objects.filter(primarycontact=contact_no).first()

                if customer:
                    # Update customer details
                    customer_full_name = request.POST.get('customer_full_name')
                    secondary_contact_no = request.POST.get('secondary_contact_no') or None
                    customer_email = request.POST.get('customer_email')
                    secondary_email = request.POST.get('secondary_email') or None
                    customer_type = request.POST.get('customer_type')
                    or_name = request.POST.get('or_name') or None
                    or_contact = request.POST.get('or_contact') or None
                    c_branch_id = request.POST.get('branch')
                    # branch = Branch.objects.get(id = branch_id)
                    # Assign values to the existing instance
                    customer.fullname = customer_full_name
                    customer.secondarycontact = secondary_contact_no
                    customer.primaryemail = customer_email
                    customer.secondaryemail = secondary_email
                    customer.customer_type = customer_type
                    customer.or_name = or_name
                    customer.or_contact = or_contact
                    customer.branch = c_branch_id
                    
                    # Save changes
                    customer.save(update_fields=[
                        "fullname",
                        "secondarycontact",
                        "primaryemail",
                        "secondaryemail",
                        "customer_type",
                        "or_name",
                        "or_contact",

                    ])
            except customer_details.DoesNotExist:
                raise ValueError("Invalid customer ID.")


        try:
            # Core data
            # customer_full_name = request.POST.get('customer_full_name')
            # contact_no = request.POST.get('contact_no')
            # secondary_contact_no = request.POST.get('secondary_contact_no')
            # customer_email = request.POST.get('customer_email')
            # secondary_email = request.POST.get('secondary_email')
            contact_by = request.POST.get('sales_person_list')
            contact_by_no = request.POST.get('contact_by_no')
            address = request.POST.get('address')
            # city = request.POST.get('city')
            # state = request.POST.get('state')
            # gps_location = request.POST.get('gps_location')
            # pincode = request.POST.get('pincode', '000000')
            subject = request.POST.get('subject')
            branch_id = request.POST.get('branch_id')
            product_details_json = request.POST.get('product_details_json')
            or_name = request.POST.get('or_name')
            or_contact = request.POST.get('or_contact')
            thank_u_note = request.POST.get('thank_u_note')
            ordered_term_ids_str = request.POST.get('terms_and_conditions_ordered', '')
            
            # Handle quotation date
            date_str = request.POST.get('quotation_date')
            if date_str:
                try:
                    quotation_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    quotation_date = timezone.now().date()
            else:
                quotation_date = timezone.now().date()

            selected_service_names_list = request.POST.get('selected_services_names', '').split(',')
            selected_service_names_list = list(set(name.strip() for name in selected_service_names_list if name.strip()))
            if not selected_service_names_list:
                raise ValueError("No services selected. Please select at least one service.")

            selected_services = Product.objects.filter(product_name__in=selected_service_names_list).distinct()
            if not selected_services.exists():
                raise ValueError("Selected services are invalid or not found.")

            # Prices and GST
            total_price = request.POST.get('total_price', '').strip()
            total_gst = request.POST.get('total_gst', '0').strip()
            total_price_with_gst = request.POST.get('total_with_gst', '').strip()
            
            # Validate prices
            try:
                total_price = float(total_price)
                total_gst = float(total_gst) if total_gst else 0
                total_price_with_gst = float(total_price_with_gst) if total_price_with_gst else total_price
    
        # Ensure total_with_gst is at least total_price
                if total_price_with_gst < total_price:
                    total_price_with_gst = total_price + total_gst
            except (ValueError, TypeError):
                raise ValueError("Invalid price values")

            enable_gst = request.POST.get('enable_gst') == 'on'
            gst_type = request.POST.get('gst_type', 'cgst_sgst')
            
            # Calculate CGST, SGST, IGST based on GST type
            if enable_gst:
                gst_status = 'GST'
                if gst_type == 'cgst_sgst':
                    cgst = total_gst / 2
                    sgst = total_gst / 2
                    igst = 0
                else:
                    cgst = 0
                    sgst = 0
                    igst = total_gst
                total_price_with_gst = total_price + total_gst    
            else:
                gst_status = 'NON-GST'
                cgst = 0
                sgst = 0
                igst = 0
                total_gst = 0

                total_price_with_gst = total_price

            # Create the quotation
            quotation = quotation_management.objects.create(
                customer=customer,
                contact_by = contact_by,
                contact_by_no = contact_by_no,
                address=address,
                subject=subject,
                quotation_date=quotation_date,
                apply_gst=enable_gst,
                gst_status=gst_status,
                total_charges=total_price,
                total_price=total_price,
                total_price_with_gst=total_price_with_gst,
                cgst=cgst,
                sgst=sgst,
                igst=igst,
                gst_total=total_gst,
                branch_id=branch_id if branch_id else None,
                product_details_json=json.loads(product_details_json),
                custom_terms = custom_terms,
                or_name = or_name,
                or_contact = or_contact,
                thank_u_note = thank_u_note,
                terms_order = ordered_term_ids_str
            )

            # Ensure it's safe
            custom_terms_list = [term.strip() for term in (custom_terms or '').split('\n') if term.strip()]

            quotation_terms_to_add = []

            for term_text in custom_terms_list:
                # Check if term already exists
                term_obj, created = QuotationTerm.objects.get_or_create(description=term_text)
                quotation_terms_to_add.append(term_obj)

            ordered_term_ids = [int(tid) for tid in ordered_term_ids_str.split(',') if tid.isdigit()]
            print('term', ordered_term_ids)
            quotation.terms_and_conditions.set(ordered_term_ids)
    
            quotation.terms_order = ordered_term_ids
            quotation.save()
            # Add terms to the quotation's terms_and_conditions field
            quotation.terms_and_conditions.add(*quotation_terms_to_add)

            quotation.selected_services.set(selected_services)
            selected_term_ids = request.POST.getlist('terms_and_conditions[]')
            quotation.terms_and_conditions.set(selected_term_ids)
            request.session.pop('quotation_form_data', None)
            request.session.modified = True

           

            return redirect(f'/create_quotation/?pdf={quotation.id}')


        except Exception as e:
            print(f"Error saving quotation: {e}")
            return render(request, 'quotation_create_new.html', {
                'error': str(e),
                'products': products,
                'category_choices': category_choices,
                'terms': terms,
                'branches': branches
            })
    form_data = request.session.get('quotation_form_data', {})
    print('form_data',form_data)
    print('product', form_data.get('product_details_json'))
    product_json = form_data.get('product_details_json', '[]')
    form_data['terms_and_conditions_ordered_list'] = form_data.get('terms_and_conditions_ordered', '').split(',')
    order_list = form_data.get('terms_and_conditions_ordered_list', [])
    terms = sorted(
        terms,
        key=lambda t: order_list.index(str(t.id)) if str(t.id) in order_list else 999
    )

    pdf_id = request.GET.get("pdf")

    context = {
        'products': products,
        'category_choices': category_choices,
        'sales_person_list': sales_person_list,
        'terms': terms,
        'branches': Branch.objects.all(),
        'form': AddProductForm(),   # always fresh
        'form_data':form_data,            # reset session cleared already
        'thank_notes': json.dumps(thank_notes),
        "product_details_json": product_json
    }

    if pdf_id:  # just submitted
        context['pdf_url'] = f'/generate_quotation/quotation/pdf/{pdf_id}/view'
        context['show_pdf_script'] = True

    return render(request, 'quotation_create_new.html', context)

    


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def save_quotation_session(request):
    if request.method == 'POST':
        data = request.POST.dict()

        # Manually fix the multi-select checkbox field
        data['terms_and_conditions'] = request.POST.getlist('terms_and_conditions[]')

        request.session['quotation_form_data'] = data
        request.session.modified = True
        return JsonResponse({'status': 'success'})
    
    return JsonResponse({'status': 'invalid request'}, status=400)
# @csrf_exempt
# def clear_quotation_session(request):
#     request.session.pop('quotation_form_data', None)
#     return JsonResponse({'status': 'success'})

from django.http import JsonResponse
from .models import quotation_management
@login_required
@role_required(['admin','sales'])
def export_quotation_excel(request):

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Quotation_list.csv"'

    writer = csv.writer(response)

    # Get all model field names
    model_fields = [field.name for field in quotation_management._meta.fields]

    # Write header row
    writer.writerow([field.replace('_', ' ').title() for field in model_fields])

    # Write data rows
    for obj in quotation_management.objects.all().values_list(*model_fields):
        writer.writerow(obj)

    return response



def get_quotation_details_by_no(request):
    quotation_no = request.GET.get('quotation_no')
    if not quotation_no:
        return JsonResponse({'error': 'Quotation number is required'}, status=400)

    try:
        quotation = quotation_management.objects.select_related('branch').get(quotation_no=quotation_no)
        branch = quotation.branch
        product = quotation.product_details_json
        aplly_gst = quotation.apply_gst
        cgst = quotation.cgst
        sgst = quotation.sgst
        igst = quotation.igst
        total = quotation.total_price_with_gst
        bank_accounts = BankAccounts.objects.all().values('id', 'bank_name', 'account_number', 'ifs_code', 'branch')
        # Fetch customer manually by customer_id string field
        customer_contact = quotation.customer.primarycontact
        customer = customer_details.objects.get(primarycontact=customer_contact)

        print(customer.customer_type)
        print(customer.or_name)
        print(customer.or_contact)
        return JsonResponse({
            # Branch data
            'branch_name':branch.branch_name if branch else '',
            'branch_contact_1': branch.contact_1 if branch else '',
            'branch_email_1': branch.email_1 if branch else '',
            'branch_gst': branch.gst_number if branch else '',
            'branch_pan': branch.pan_number if branch else '',
            'branch_address': branch.full_address if branch else '',
            'state':branch.state if branch else '',
            'code':branch.code if branch else '',

            # Customer data
            'fullname': customer.fullname,
            'primarycontact': customer.primarycontact,
            'primaryemail': customer.primaryemail,
            'soldtopartyaddress': customer.soldtopartyaddress,
            'soldtopartycity': customer.soldtopartycity,
            'soldtopartystate': customer.soldtopartystate,
            'soldtopartypostal': customer.soldtopartypostal,
            'shifttopartyaddress': customer.shifttopartyaddress,
            'shifttopartycity': customer.shifttopartycity,
            'shifttopartystate': customer.shifttopartystate,
            'shifttopartypostal': customer.shifttopartypostal,
            'customer_type': customer.customer_type,
            'or_name': customer.or_name,
            'or_contact':customer.or_contact,
            
            # Product data
            'product':product,
            'apply_gst':aplly_gst,
            'cgst':cgst,
            'sgst':sgst,
            'igst':igst,
            'total':total,
            

            # Bank data
            'bank': list(bank_accounts),
            'state_map': state_map,
        })
       

    except quotation_management.DoesNotExist:
        return JsonResponse({'error': 'Quotation not found'}, status=404)
    except customer_details.DoesNotExist:
        return JsonResponse({'error': 'Customer not found for this quotation'}, status=404)
    
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from .models import quotation_management

def generate_quotation_pdf_download(request, id):
    quotation = quotation_management.objects.get(id=id)
    template_path = 'pdf_template.html'
    context = {'quotation': quotation}
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.id}.pdf"'
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response




from django.http import HttpResponse
from django.template.loader import get_template
from crmapp.custom_filters import price_in_words
from xhtml2pdf import pisa
import json
import os


def get_products_by_category(request):
    categories = request.GET.get('categories', '')
    category_list = categories.split(',') if categories else []

    products = Product.objects.filter(category__in=category_list).values('product_id', 'product_name','hsn_code')
    product_list = [{'product_id': product['product_id'], 'product_name': product['product_name'], 'hsn_code':product['hsn_code']} for product in products]

    return JsonResponse({'products': product_list})

from django.http import JsonResponse

def get_customer_fullname(request, customer_id):
    customer = customer_details.objects.get(id=customer_id)
    full_name = f"{customer.firstname} {customer.lastname}"
    return JsonResponse({'full_name': full_name})
  



from django.shortcuts import render, redirect, get_object_or_404
from .models import QuotationTerm, InvoiceTerm

# --- QUOTATION ---
def add_quotation_term(request):
    if request.method == 'POST':
        QuotationTerm.objects.create(description=request.POST.get('description'))
        return redirect('view_quotation_terms')
    return render(request, 'add_quotation_term.html')

def edit_quotation_term(request, id):
    term = get_object_or_404(QuotationTerm, id=id)
    
    if request.method == 'POST':
        term.description = request.POST.get('description')
        term.save()
        return redirect('view_quotation_terms')
    return render(request, 'edit_quotation_term.html', {'term': term})

def view_quotation_terms(request):
    terms = QuotationTerm.objects.all()
    return render(request, 'view_quotation_terms.html', {'terms': terms})

def delete_quotation_term(request, id):
    term = get_object_or_404(QuotationTerm, id=id)
    term.delete()
    return redirect('view_quotation_terms')


# --- INVOICE ---
def add_invoice_term(request):
    if request.method == 'POST':
        InvoiceTerm.objects.create(description=request.POST.get('description'))
        return redirect('view_invoice_terms')
    return render(request, 'add_invoice_term.html')

def edit_invoice_term(request, id):
    term = get_object_or_404(InvoiceTerm, id=id)
    if request.method == 'POST':
        term.description = request.POST.get('description')
        term.save()
        return redirect('view_invoice_terms')
    return render(request, 'edit_invoice_term.html', {'term': term})

def view_invoice_terms(request):
    terms = InvoiceTerm.objects.all()
    return render(request, 'view_invoice_terms.html', {'terms': terms})

def delete_invoice_term(request, id):
    term = get_object_or_404(InvoiceTerm, id=id)
    term.delete()
    return redirect('view_invoice_terms')




def quotation_history(request, customer_id):
    customer = customer_details.objects.get(id=customer_id)
    print(customer, customer_id)
    quotations = quotation_management.objects.filter(customer=customer).order_by('-quotation_date')  
    context = {
        'customer': customer,
        'quotations': quotations,
    }
    return render(request, 'quotation_history.html', context)



import random
from django.utils.timezone import now
import string

def generate_invoice_number():
    # Use "INV" as the first 3 characters
    alphanumeric_part = "INV"

    # Generate remaining 7 digits
    numeric_part = ''.join(random.choices(string.digits, k=7))

    # Combine both parts
    invoice_no = alphanumeric_part + numeric_part

    return invoice_no



def invoice_create(request):
    customers = customer_details.objects.all()  # Fetch all customers from the database

    if request.method == 'POST':
        modeofpayment = request.POST['modeofpayment']
        dispatchedthrough = request.POST['dispatchedthrough']
        termofdelivery = request.POST['termofdelivery']
        termsandcondition = request.POST['termsandcondition']
        company_name = request.POST['company_name']
        company_email = request.POST['company_email']
        company_contact_no = request.POST['company_contact_no']
        description_of_goods = request.POST['description_of_goods']
        hsn_sac_code = request.POST['hsn_sac_code']
        quantity = int(request.POST['quantity'])
        price = float(request.POST['price'])
        discount = float(request.POST['discount']) if request.POST['discount'] else None
        gst_checkbox = True if 'gst_checkbox' in request.POST else False
        pan_card_no = request.POST['pan_card_no']
        account_no = request.POST['account_no']
        branch = request.POST['branch']
        ifsc_code = request.POST['ifsc_code']
        delivery_date = request.POST['delivery_date']
        dispatched_date = request.POST['dispatched_date']
        designation = request.POST['designation']
        customer_id = request.POST['customer_id']
        customer = customer_details.objects.get(id=customer_id)
        invoice_no = generate_invoice_number()
        total_amount = quantity * price
        
        discounted_amount = total_amount - (total_amount * (discount / 100))
        total_amount_with_gst = discounted_amount * 1.18 if gst_checkbox else discounted_amount

        m = invoice.objects.create(
            modeofpayment=modeofpayment,
            dispatchedthrough=dispatchedthrough,
            termofdelivery=termofdelivery,
            termsandcondition=termsandcondition,
            company_name=company_name,
            company_email=company_email,
            company_contact_no=company_contact_no,
            description_of_goods=description_of_goods,
            hsn_sac_code=hsn_sac_code,
            quantity=quantity,
            price=price,
            discount=discount,
            gst_checkbox=gst_checkbox,
            pan_card_no=pan_card_no,
            account_no=account_no,
            branch=branch,
            ifsc_code=ifsc_code,
            delivery_date=delivery_date,
            dispatched_date=dispatched_date,
            designation=designation,
            customer=customer,
            total_amount_with_gst=total_amount_with_gst,
            total_amount=total_amount,
            invoice_no=invoice_no,

        )

        m.save()
        return redirect('/display_invoice') 
        
    context = {
        'customers': customers,
        }
    return render(request, 'invoice.html', context)

   
def firstfollowup_create(request,lead_id,next_stage):
    lead = get_object_or_404(lead_management, id=lead_id)
    lead.stage = next_stage
    lead.save()

    if request.method == 'POST':
        havedonepestcontrolearlier = request.POST['havedonepestcontrolearlier']
        agency = request.POST['agency']
        inspectiononsite = request.POST['inspectiononsite']
        levelofinspection = request.POST['levelofinspection']
        quotationgiven = request.POST['quotationgiven']
        quotationamount = float(request.POST['quotationamount'])
        mailsent = request.POST['mailsent']
        customermeeting = request.POST['customermeeting']
        firstremark = request.POST['firstremark']
        secondfollowupdate = request.POST['secondfollowupdate']

        m = firstfollowup.objects.create(
            havedonepestcontrolearlier=havedonepestcontrolearlier,
            agency=agency,
            inspectiononsite=inspectiononsite,
            levelofinspection=levelofinspection,
            quotationgiven=quotationgiven,
            quotationamount=quotationamount,
            mailsent=mailsent,
            customermeeting=customermeeting,
            firstremark=firstremark,
            secondfollowupdate=secondfollowupdate,
            lead=lead,
        )

        m.save()
        return redirect('/display_followup') 
        
    context = {
        'lead': lead,
    }
    return render(request, 'first_followup.html', context)

def secondfollowup_create(request,lead_id,next_stage):
    lead = get_object_or_404(lead_management, id=lead_id)
    lead.stage = next_stage
    lead.save()

    if request.method == 'POST':
        negotiationstage = request.POST['negotiationstage']
        mailsent2 = request.POST['mailsent2']
        secondremark = request.POST['secondremark']
        thirdfollowupdate = request.POST['thirdfollowupdate']

        m = secondfollowup.objects.create(
            negotiationstage=negotiationstage,
            mailsent2=mailsent2,
            secondremark=secondremark,
            thirdfollowupdate=thirdfollowupdate,
            lead=lead,
        )

        m.save()
        return redirect('/display_followup') 
        
    context = {
        'lead': lead,
        }
    return render(request, 'second_followup.html', context)

def thirdfollowup_create(request,lead_id,next_stage):
    lead = get_object_or_404(lead_management, id=lead_id)
    lead.stage = next_stage
    lead.save()

    if request.method == 'POST':
        thirdremark = request.POST['thirdremark']
        fourthfollowupdate = request.POST['fourthfollowupdate']

        m = thirdfollowup.objects.create(
            thirdremark=thirdremark,
            fourthfollowupdate=fourthfollowupdate,
            lead=lead,
        )

        m.save()
        return redirect('/display_followup') 
        
    context = {
        'lead': lead,
        }
    return render(request, 'third_followup.html', context)

def finalfollowup_create(request,lead_id,next_stage):
    lead = get_object_or_404(lead_management, id=lead_id)
    lead.stage = next_stage
    lead.save()
    
    if request.method == 'POST':
        fourthremark = request.POST['fourthremark']
        finalstatus = request.POST['finalstatus']
        contracttype = request.POST['contracttype']
        bookingamount = request.POST['bookingamount']

        m = finalfollowup.objects.create(
            fourthremark=fourthremark,
            finalstatus=finalstatus,
            contracttype=contracttype,
            bookingamount=bookingamount,
            lead=lead,
        )

        m.save()
        return redirect('/display_followup') 
        
    context = {
        'lead': lead,
        }
    return render(request, 'final_followup.html', context)




from django.http import JsonResponse
from .models import lead_management

def check_phone_number(request):
    phone = request.GET.get('primarycontact', None)
    exists = lead_management.objects.filter(primarycontact=phone).exists()
    return JsonResponse({'exists': exists})



from django.http import JsonResponse
from django.shortcuts import render, redirect
from .models import lead_management
from datetime import datetime
from django.utils import timezone
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import lead_management
from datetime import datetime

from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import lead_management
from datetime import datetime

# change

@login_required
@role_required(['admin','sales','branch_manager'])
def lead_management_create(request):
    salespersons = list(SalesPerson.objects.all())+ list(BranchManager.objects.all())
    
    branches = Branch.objects.all()
    if request.method == 'GET':
        # Handle AJAX GET for mobile number lookup
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' and 'primarycontact' in request.GET:
            primarycontact = request.GET.get('primarycontact')
            lead = lead_management.objects.filter(primarycontact=primarycontact).order_by('enquirydate').first()
            
            if lead:
                data = {
                    'sourceoflead': lead.sourceoflead,
                    'salesperson': lead.salesperson.id if lead.salesperson else '',
                    'customername': lead.customername,
                    'customeremail':lead.customeremail,
                    'customersegment': lead.customersegment,
                    'enquirydate': lead.enquirydate.strftime('%Y-%m-%d') if lead.enquirydate else '',
                    'contactedby': lead.contactedby,
                    'maincategory': lead.maincategory,
                    'subcategory': lead.subcategory,
                    'secondarycontact': lead.secondarycontact,
                    'customeremail': lead.customeremail,
                    'customeraddress': lead.customeraddress,
                    'location': lead.location,
                    'state': lead.state,
                    'city': lead.city,
                    'branch': lead.branch.id if lead.branch else '',
                    'typeoflead': lead.typeoflead,
                    'firstfollowupdate': lead.firstfollowupdate.strftime('%Y-%m-%d') if lead.firstfollowupdate else '',
                }
                return JsonResponse({'status': 'exists', 'data': data})
            else:
                return JsonResponse({'status': 'not_found'})

        return render(request, 'lead_management.html', {'salespersons': salespersons, 'branches': branches})

    else:  # POST request
        try:
            role = request.user.userprofile.role
            # Get all form data
            sourceoflead = request.POST.get('sourceoflead')
            # salesperson_id = request.POST.get("salesperson")
          
            customername = request.POST.get('customername')
            customersegment = request.POST.get('customersegment')
            
            # Handle date fields
            enquirydate_str = request.POST.get('enquirydate')
            enquirydate = datetime.strptime(enquirydate_str, '%Y-%m-%d').date() if enquirydate_str else None
            
            contactedby = request.POST.get('contactedby')
            maincategory = request.POST.get('maincategory')
            subcategory = request.POST.get('subcategory')

            # Handle phone numbers
            primarycontact = request.POST.get('primarycontact')
            primarycontact = int(primarycontact) if primarycontact and primarycontact.isdigit() else None
            
            secondarycontact = request.POST.get('secondarycontact')
            secondarycontact = int(secondarycontact) if secondarycontact and secondarycontact.isdigit() else None
            or_contact = request.POST.get('or_contact')
            or_contact = int(or_contact) if or_contact and or_contact.isdigit() else None
            or_name = request.POST.get('or_name')
            # Other fields
            customeremail = request.POST.get('customeremail')
            customeraddress = request.POST.get('customeraddress')
            location = request.POST.get('location')
            city = request.POST.get('city', 'Unknown City')
            state = request.POST.get('state')
            typeoflead = request.POST.get('typeoflead')
            customer_type = request.POST.get('customer_type')
            
            firstfollowupdate_str = request.POST.get('firstfollowupdate')
            firstfollowupdate = datetime.strptime(firstfollowupdate_str, '%Y-%m-%d').date() if firstfollowupdate_str else None

            branch_id = request.POST.get('branch')
            branch = branches.get(id = int(branch_id))
            sp = None
            bm = None
            ad = None
            if role == "sales":
                sp = SalesPerson.objects.get(mobile_no = request.user.username)
            elif role == "branch_manager":
                bm = BranchManager.objects.get(mobile_no = request.user.username)
            else: 
                ad = request.user
            # Create new lead (duplicates allowed)
            lead = lead_management.objects.create(
                sourceoflead=sourceoflead,
                salesperson=sp,
                branch_manager = bm,
                admin = ad,
                customername=customername,
                customersegment=customersegment,
                enquirydate=enquirydate,
                contactedby=contactedby,
                maincategory=maincategory,
                subcategory=subcategory,
                primarycontact=primarycontact,
                secondarycontact=secondarycontact,
                customeremail=customeremail,
                customeraddress=customeraddress,
                location=location,
                state=state,
                city=city,
                branch=branch,
                typeoflead=typeoflead,
                firstfollowupdate=firstfollowupdate,
                or_contact = or_contact,
                or_name = or_name,
                customer_type = customer_type,
            )

            return redirect('/display_lead_management')

        except Exception as e:
            # Log the error for debugging
            print(f"Error creating lead: {str(e)}")
            # Return the user to the form with error message
            return render(request, 'lead_management.html', {
                'salespersons': salespersons,
                'error': 'There was an error submitting the form. Please try again.'
            })























import json
from datetime import date
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

@csrf_exempt
def landing_page_lead_create_api(request):
    if request.method != "POST":
        return JsonResponse({
            "status": False,
            "message": "Only POST method allowed"
        }, status=405)

    try:
        data = json.loads(request.body)

        # Defaults
        sourceoflead = "Landing Page"
        enquirydate = date.today()
        firstfollowupdate = date.today()
        typeoflead = "Hot"

        # Basic fields
        customername = data.get("name")
        primarycontact = data.get("contact")
        customeremail = data.get("email")
        city = data.get("city")

        # Convert service type
        service_type = data.get("service_type")
        if service_type == "Individual":
            customer_type = "Residential"
        else:
            customer_type = "Commercial"

        # Categories
        maincategory = data.get("interested_in")          # Pest Control / Fumigation
        subcategory = data.get("issue")                   # Description

        lead = lead_management.objects.create(
            sourceoflead=sourceoflead,
            customername=customername,
            enquirydate=enquirydate,
            primarycontact=int(primarycontact) if primarycontact else None,
            customeremail=customeremail,
            city=city,
            customer_type=customer_type,
            maincategory=maincategory,
            subcategory=subcategory,
            typeoflead=typeoflead,
            firstfollowupdate=firstfollowupdate,
        )

        return JsonResponse({
            "status": True,
            "message": "Lead submitted successfully",
            "lead_id": lead.id
        }, status=201)

    except Exception as e:
        print("Landing Page API Error:", e)
        return JsonResponse({
            "status": False,
            "message": "Something went wrong",
            "error": str(e)
        }, status=500)




























































from django.shortcuts import render, get_object_or_404, redirect
from .models import lead_management, main_followup
from django.utils import timezone

def main_followup_view(request, lead_id):
    lead = get_object_or_404(lead_management, id=lead_id)
    followups = main_followup.objects.filter(lead=lead).order_by('created_at')
    followup_count = followups.count() + 1

    latest_followup = followups.last()

    if request.method == "POST":
        done_pest_control = request.POST.get('done_pest_control', 'No')
        agency_name = request.POST.get('agency_name') if done_pest_control == 'Yes' else None
        onsite_infestation = request.POST.get('onsite_infestation', '')
        infestation_level = request.POST.get('infestation_level', '')
        typeoflead = request.POST.get('typeoflead', '')
        followup_remark = request.POST.get('followup_remark', '')
        followup_comment = request.POST.get('followup_comment', '')
        order_status = request.POST.get('order_status', 'Not Closed')
        next_followup_date = (
            request.POST.get('next_followup_date')
            if order_status not in ['Close Win', 'Close Loss'] else None
        )

        if latest_followup:
            # Update existing follow-up
            latest_followup.done_pest_control = done_pest_control
            latest_followup.agency_name = agency_name
            latest_followup.onsite_infestation = onsite_infestation
            latest_followup.infestation_level = infestation_level
            latest_followup.typeoflead = typeoflead
            latest_followup.followup_remark = followup_remark
            latest_followup.followup_comment = followup_comment
            latest_followup.next_followup_date = next_followup_date
            latest_followup.order_status = order_status
            latest_followup.save()
        else:
            # Create new follow-up if none exists
            main_followup.objects.create(
                lead=lead,
                done_pest_control=done_pest_control,
                agency_name=agency_name,
                onsite_infestation=onsite_infestation,
                infestation_level=infestation_level,
                typeoflead=typeoflead,
                followup_remark=followup_remark,
                followup_comment=followup_comment,
                next_followup_date=next_followup_date,
                order_status=order_status,
                created_at=timezone.now(),
            )

        # Update the lead's status
       # Update the lead's status
        lead.typeoflead = typeoflead
        main_followup.order_status = order_status  
        if order_status in ['Close Win', 'Close Loss']:
            lead.stage = 0  
            lead.save()
            if order_status == 'Close Win':
                return redirect('service_management_create')
        else:
            lead.save()  
        query_string = request.GET.urlencode()
        redirect_url = reverse('pending_followups')
        if query_string:
            redirect_url += f"?{query_string}"
        return redirect(redirect_url)

    context = {
        'lead': lead,
        'followup_count': followup_count,
        'followups': followups,
        'previous_followup': latest_followup,
        
    }
    return render(request, 'main_followup.html', context)


from django.shortcuts import render
from .models import main_followup, lead_management
from django.utils import timezone
from datetime import date
from django.core.paginator import Paginator
from itertools import chain

@login_required
@role_required(['admin','sales','branch_manager'])
def today_work(request):
    today = date.today()

    salesperson_filter = request.GET.get('salesperson')

    # Filter today's follow-ups
    if request.user.userprofile.role =='admin':
        lead_folloup = lead_management.objects.filter(firstfollowupdate = today)
        followups = main_followup.objects.filter(next_followup_date=today).select_related('lead')
    elif request.user.userprofile.role == 'sales': 
        lead_folloup = lead_management.objects.filter(firstfollowupdate = today ,
                                                    salesperson__mobile_no = request.user.username )
        followups = main_followup.objects.filter(next_followup_date=today, 
                                                 lead__salesperson__mobile_no=request.user.username ).select_related('lead')
    elif request.user.userprofile.role == 'branch_manager': 
        branch = BranchManager.objects.get(mobile_no = request.user.username ).branch
        lead_folloup = lead_management.objects.filter(firstfollowupdate = today ,
                                                    branch = branch)
        followups = main_followup.objects.filter(next_followup_date=today, 
                                                 lead__branch = branch ).select_related('lead')
            
    # lead_folloup = lead_management.objects.filter(firstfollowupdate = today)
    # followups = main_followup.objects.filter(next_followup_date=today).select_related('lead')

    if salesperson_filter:
        followups = followups.filter(lead__salesperson=salesperson_filter)
        lead_folloup = lead_folloup.filter(salesperson = salesperson_filter)

    # Combine followups and lead_folloup
    combined = list(chain(
        followups,  # main_followup objects (with .lead field)
        [lead for lead in lead_folloup if not main_followup.objects.filter(lead=lead).exists()]  # avoid duplication
    ))
    
    count = len(combined)
    # Pagination
    paginator = Paginator(combined, 10)  # Show 10 records per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get all unique salespersons
    salespersons = SalesPerson.objects.all()

    # Used for correct indexing
    start_index = (page_obj.number - 1) * paginator.per_page


    return render(request, 'today_work.html', {
        'page_obj': page_obj,
        'salespersons': salespersons,
        'selected_salesperson': salesperson_filter,
        'start_index': start_index,
        'count': count,
    })




from itertools import chain
from datetime import date
from django.core.paginator import Paginator
from django.shortcuts import render
from django.db.models import Q
from .models import lead_management, main_followup, SalesPerson

from itertools import chain
from django.db.models import Q
from django.core.paginator import Paginator
from datetime import date

@login_required
@role_required(['admin','sales', 'branch_manager'])
def pending_followups(request):
    today = date.today()
    branches = Branch.objects.all()
    salespersons = []

    # Get filters from request
    search_query = request.GET.get('search', '').strip()
    typeoflead_filter = request.GET.get('typeoflead')
    source_filter = request.GET.get('sourceoflead')
    salesperson_filter = request.GET.get('salesperson')
    branch_filter = request.GET.get('branch')
    enquiry_from = request.GET.get('enquiry_from')
    enquiry_to = request.GET.get('enquiry_to')
    followup_from = request.GET.get('followup_from')
    followup_to = request.GET.get('followup_to')
    sort_by = request.GET.get('sort', 'customername')
    order = request.GET.get('order', 'asc')
    segment_filter = request.GET.get('segments')

    # Base queryset filtered by role
    if request.user.userprofile.role == 'admin':
        lead_folloup = lead_management.objects.filter(firstfollowupdate__lt=today)
        salespersons = list(SalesPerson.objects.all()) + list(BranchManager.objects.all())
    elif request.user.userprofile.role == 'sales':
        lead_folloup = lead_management.objects.filter(
            firstfollowupdate__lt=today,
            salesperson__mobile_no=request.user.username
        )
    else:  # branch_manager
        branch_manager = BranchManager.objects.get(mobile_no=request.user.username)
        lead_folloup = lead_management.objects.filter(
            firstfollowupdate__lt=today,
            branch=branch_manager.branch
        )
        salespersons = SalesPerson.objects.filter(branch=branch_manager.branch)

    # Overdue followups (main_followup instances referencing leads)
    followups = main_followup.objects.filter(next_followup_date__lt=today).select_related('lead')

    # Role-based restriction for followups (only once)
    if request.user.userprofile.role != 'admin':
        followups = followups.filter(lead__salesperson__mobile_no=request.user.username)

    # --- Simple search via DB icontains (applies before combining) ---
    if search_query:
        lead_search_q = (
            Q(primarycontact__icontains=search_query) |
            Q(customername__icontains=search_query) |
            Q(typeoflead__icontains=search_query)
        )
        lead_folloup = lead_folloup.filter(lead_search_q)

        followup_search_q = (
            Q(lead__primarycontact__icontains=search_query) |
            Q(lead__customername__icontains=search_query) |
            Q(lead__typeoflead__icontains=search_query)
        )
        followups = followups.filter(followup_search_q)

    # Apply other filters to both querysets where relevant
    if typeoflead_filter:
        lead_folloup = lead_folloup.filter(typeoflead=typeoflead_filter)
        followups = followups.filter(lead__typeoflead=typeoflead_filter)
    if source_filter:
        lead_folloup = lead_folloup.filter(sourceoflead=source_filter)
        followups = followups.filter(lead__sourceoflead=source_filter)
    if branch_filter:
        try:
            bf_id = int(branch_filter)
            lead_folloup = lead_folloup.filter(branch_id=bf_id)
            followups = followups.filter(lead__branch_id=bf_id)
        except (ValueError, TypeError):
            pass
    if segment_filter:
        lead_folloup = lead_folloup.filter(customersegment=segment_filter)
        followups = followups.filter(lead__customersegment=segment_filter)
    if enquiry_from and enquiry_to:
        lead_folloup = lead_folloup.filter(enquirydate__range=[enquiry_from, enquiry_to])
    if followup_from and followup_to:
        lead_folloup = lead_folloup.filter(firstfollowupdate__range=[followup_from, followup_to])

    if salesperson_filter:
        lead_folloup = lead_folloup.filter(
            Q(salesperson__full_name=salesperson_filter) |
            Q(branch_manager__full_name=salesperson_filter)
        )
        followups = followups.filter(
            Q(lead__salesperson__full_name=salesperson_filter) |
            Q(lead__branch_manager__full_name=salesperson_filter)
        )

    # Combine leads:
    # - include followup objects (they reference lead via .lead)
    # - include leads from lead_folloup that do NOT have any main_followup
    leads_without_followup = lead_folloup.exclude(id__in=main_followup.objects.values_list('lead_id', flat=True))
    combined = list(chain(followups, leads_without_followup))

    # Sorting: get attribute from underlying lead (for followup items)
    def get_sort_value(obj):
        lead = obj.lead if hasattr(obj, 'lead') else obj
        return getattr(lead, sort_by, '') or ''

    combined.sort(key=get_sort_value, reverse=(order == 'desc'))

    # Pagination on combined list
    paginator = Paginator(combined, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page

    # Dropdown options
    typeoflead_choices = [c[0] for c in lead_management._meta.get_field('typeoflead').choices if c[0]]
    source_choices = [c[0] for c in lead_management._meta.get_field('sourceoflead').choices if c[0]]
    segments = [c[0] for c in lead_management._meta.get_field('customersegment').choices if c[0] != "NOT SELECTED"]

    return render(request, 'pending_followups.html', {
        'page_obj': page_obj,
        'count_data': paginator.count,
        'search_query': search_query,
        'start_index': start_index,
        'lead_types': typeoflead_choices,
        'sources': source_choices,
        'branches': branches,
        'salespersons': salespersons,
        'selected_salesperson': salesperson_filter,
        'current_sort': sort_by,
        'current_order': order,
        'segments': segments,
        'selected_segment': segment_filter,
    })


# In crmapp/views.py

from django.shortcuts import render
from .models import lead_management, firstfollowup, secondfollowup, thirdfollowup, finalfollowup


from django.db.models import Q  # For complex queries

@login_required
@role_required(['admin','sales', 'branch_manager','operation_person'])
def display_followup(request):
    search_query = request.GET.get('q', '')  # Get the search query
    
    leads = lead_management.objects.all()
    if search_query:
        leads = leads.filter(
            Q(customername__icontains=search_query) |  # Search by customer name
            Q(customeremail__icontains=search_query)  # Optionally, search by email
        )
    
    followups = []
    for lead in leads:
        followups.append({
            "lead": lead,
            "firstfollowup": firstfollowup.objects.filter(lead=lead).first(),
            "secondfollowup": secondfollowup.objects.filter(lead=lead).first(),
            "thirdfollowup": thirdfollowup.objects.filter(lead=lead).first(),
            "finalfollowup": finalfollowup.objects.filter(lead=lead).first(),
        })
    
    context = {
        "followups": followups,
        "search_query": search_query,
    }
    return render(request, 'display_followup.html', context)

def fetch_customer_details(request, customer_id):
    customer = get_object_or_404(customer_details, customerid=customer_id)
    return render(request, 'customer_details_modal.html', {'customer': customer})


def display_customer(request):
    query = request.GET.get('search', '')
    sort_order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'customerid')
    customer_type = request.GET.get('customer_type')

    if request.user.userprofile.role == 'admin': 
        m = customer_details.objects.all()

    elif request.user.userprofile.role == 'sales':
        sales_person_name = SalesPerson.objects.get(full_name = SalesPerson.objects.get(mobile_no = request.user.username).full_name)
        m = customer_details.objects.filter(contactperson__iexact = sales_person_name)

    elif request.user.userprofile.role == 'branch_manager':
        branch = BranchManager.objects.get(mobile_no = request.user.username ).branch
        m = customer_details.objects.filter(branch_id = branch)
        
    # Base queryset

    # Apply search filters
    if query:
        m = m.filter(
            Q(customerid__icontains=query) |
            Q(primarycontact__icontains=query) |
            Q(fullname__icontains=query)
        )

    # Apply customer type filter
    if customer_type:
        m = m.filter(customer_type=customer_type)

    filter_count = m.count()
    # Apply sorting
    if sort_by == 'firstname':
        m = m.order_by('-firstname' if sort_order == 'desc' else 'firstname')
    else:
        m = m.order_by('-customerid' if sort_order == 'desc' else 'customerid')

    m = m.order_by('-id')
    # Pagination
    paginator = Paginator(m, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page

    context = {
        'current_order': sort_order,
        'current_sort_by': sort_by,
        'page_obj': page_obj,
        'start_index': start_index,
        'data': page_obj.object_list,
        'selected_type': customer_type,  
        'search_query': query,   
        'filter_count' : filter_count,         
    }

    return render(request, 'display_customer.html', context)

def import_customers(request):
    if request.method == 'POST':
        form = CustomerImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_type = file.name.split('.')[-1]

            if file_type == 'csv':
                handle_customer_csv(file)
                messages.success(request, "Customer data imported successfully.")
            else:
                messages.error(request, 'Only CSV files are supported.')
                return redirect('import_customers')

            return redirect('display_customer')  # Update as per your success page
    else:
        form = CustomerImportForm()

    return render(request, 'import_customer.html', {'form': form})


def handle_customer_csv(file):
    decoded_file = file.read().decode('utf-8').splitlines()
    reader = csv.reader(decoded_file)
    next(reader)  # Skip header

    for row in reader:
        try:
            if len(row) < 16:
                print(f"⚠️ Skipped incomplete row: {row}")
                continue

            customer_details.objects.create(
                fullname=row[1].strip(),
                primaryemail=row[2].strip(),
                secondaryemail=row[3].strip() if row[3].strip().lower() != 'null' else None,
                primarycontact=int(row[4].strip()),
                secondarycontact=int(row[5].strip()) if row[5].strip().lower() != 'null' else None,
                contactperson=row[6].strip(),
                customersegment=row[7].strip(),
                shifttopartyaddress=row[8].strip(),
                shifttopartycity=row[9].strip(),
                shifttopartystate=row[10].strip(),
                shifttopartypostal=row[11].strip(),
                soldtopartyaddress=row[12].strip(),
                soldtopartycity=row[13].strip(),
                soldtopartystate=row[14].strip(),
                soldtopartypostal=row[15].strip(),
                customerid=row[16].strip() if row[16].strip() else None,
            )

        except Exception as e:
            print(f"❌ Error importing row {row}: {e}")


from crmapp.models import Reschedule
from django.db.models import Prefetch
@login_required
@role_required(['admin','sales','branch_manager','operation_person'])
def display_reschedule(request):
    query = request.GET.get('search', '').strip()
    sort_order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'customerid')
    
    if query:
        # Filter results by customer ID or service ID
        m = Reschedule.objects.filter(
            service__customer__customerid__icontains=query
        ) | Reschedule.objects.filter(
            service__id__icontains=query
        )
    else:
        # Display all records if no search query
        m = Reschedule.objects.all()

    if sort_by == 'service_id':
        if sort_order == 'desc':
            m = m.order_by('-service__id')  
        else:
            m = m.order_by('service__id')  
    else:
        if sort_order == 'desc':
            m = m.order_by('-service__customer__customerid')  
        else:
            m = m.order_by('service__customer__customerid')

    paginator = Paginator(m, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page

    context = {
        'page_obj': page_obj,
        'start_index': start_index,
        'search_query': query,
        'current_order': sort_order,
        'current_sort_by': sort_by,
    }

    return render(request, 'display_reschedule.html', context)

@login_required
@role_required(['admin','sales', 'branch_manager'])
def display_service_management(request):
    query = request.GET.get('search', '')
    sort_order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'customerid')  
    selected_contract_type = request.GET.get('contract_type', '')

    selected_segment = request.GET.get('segments', '')
    selected_salesperson = request.GET.get('salesperson', '')
    service_from = request.GET.get('service_from', '')
    service_to = request.GET.get('service_to', '')
    customer_type = request.GET.get('customer_type')

    m = service_management.objects.all()
    if request.user.userprofile.role =='admin':
        m = service_management.objects.all()
    elif request.user.userprofile.role == 'sales':
        m = service_management.objects.filter(sales_person_contact_no = request.user.username)
    elif request.user.userprofile.role == 'branch_manager':
        m = service_management.objects.filter(branch = BranchManager.objects.get(mobile_no = request.user.username).branch)
    
    if query:
        m = m.filter(
            Q(customer__customerid__icontains=query) | 
            Q(customer__primarycontact__icontains=query)
        )

    # Apply Segment filter
    if selected_segment:
        m = m.filter(segment=selected_segment)

    # Apply Salesperson filter
    if selected_salesperson:
        m = m.filter(sales_person_name = selected_salesperson)

    # Filter by contract type if provided
    if selected_contract_type:
        m = m.filter(contract_type=selected_contract_type)

    # Filter by customer type 
    if customer_type:
        m=m.filter(customer__customer_type=customer_type)

    # Apply Service Date Range Filter
    if service_from:
        try:
            service_from_date = datetime.strptime(service_from, '%Y-%m-%d').date()
            m = m.filter(service_date__gte=service_from_date)
        except ValueError:
            pass  
        
    if service_to:
        try:
            service_to_date = datetime.strptime(service_to, '%Y-%m-%d').date()
            m = m.filter(service_date__lte=service_to_date)
        except ValueError:
            pass
    
    # Sorting logic
    if sort_by == 'firstname':
        m = m.order_by('-customer__firstname' if sort_order == 'desc' else 'customer__firstname')
    else:
        m = m.order_by('-customer__customerid' if sort_order == 'desc' else 'customer__customerid')

    m = m.order_by('-id')
    paginator = Paginator(m, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page

    count_data = m.count()

    # Get distinct segments and salespersons for dropdowns
    segments_choices = service_management._meta.get_field('segment').choices
    segments = [choice[0] for choice in segments_choices]
    salespersons = SalesPerson.objects.all()
    contract_types = ['One Time', 'AMC', 'Warranty']

    context = {
        'current_order': sort_order,
        'current_sort_by': sort_by,
        'page_obj': page_obj,
        'start_index': start_index,
        'contract_type': contract_types,
        'count_data':count_data,
        'segments': segments,
        'salespersons': salespersons,
        'selected_segment': selected_segment,
        
    }
    context['data'] = m
    return render(request, 'display_service_management.html', context)


def get_service_details(request, service_id):
    service = get_object_or_404(service_management, id=service_id)
    return render(request, 'service_details_modal.html', {'service': service})


from itertools import chain
from django.core.paginator import Paginator
from django.db.models import Q

def display_allocation(request):
    user_profile = request.user.userprofile
    query = request.GET.get('search', '')
    sort_order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'customerid')

    # ✅ Base queryset depending on role
    if user_profile.role == 'admin':
        base_qs = service_management.objects.all()
        print("Admin called")

    elif user_profile.role == 'branch_manager':
        branch_manager = BranchManager.objects.get( mobile_no = request.user.username)
        base_qs = service_management.objects.filter(branch = branch_manager.branch)
        print("Manager called")
    elif user_profile.role == 'operation_person':
        operation_person = OperationPerson.objects.get( user = request.user)
        base_qs = service_management.objects.filter(branch=operation_person.branch)
        print("Operation called")

    # ✅ Base queryset depending on search
    if query:
        base_qs = service_management.objects.filter(
            Q(customer__customerid__icontains=query) |
            Q(customer__primarycontact__icontains=query)
        )

    # ✅ Split into groups
    allocated_ids = WorkAllocation.objects.values_list('service_id', flat=True)

    unallocated = base_qs.exclude(id__in=allocated_ids)
    pending = base_qs.filter(work_allocations__status="Pending").distinct()
    completed = base_qs.filter(work_allocations__status="Completed").distinct()

    # ✅ Sorting field
    if sort_by == 'firstname':
        order_field = 'customer__firstname'
    else:
        order_field = 'customer__customerid'

    if sort_order == 'desc':
        order_field = f'-{order_field}'


    # After (recommended fix)
    sort_field_map = {
    'customerid': 'customer__customerid',
    'fullname':   'customer__fullname',     # ← add this line
        }

   
    # ✅ Apply sorting
    unallocated = unallocated.order_by(order_field)
    pending = pending.order_by(order_field)
    completed = completed.order_by(order_field)

    # ✅ Merge in required order
    m = list(chain(unallocated, pending, completed))

    # ✅ Pagination
    paginator = Paginator(m, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page

    context = {
        'unallocated_ids': list(unallocated.values_list('id', flat=True)),
        'pending_ids': list(pending.values_list('id', flat=True)),
        'completed_ids': list(completed.values_list('id', flat=True)),
        'current_order': sort_order,
        'current_sort_by': sort_by,
        'page_obj': page_obj,
        'start_index': start_index,
    }
    return render(request, 'display_allocation.html', context)

def get_allocation_details(request, service_id):
    allocation = get_object_or_404(service_management, id=service_id)
    return render(request, 'allocation_details_modal.html', {'allocation': allocation})



from django.core.paginator import Paginator
from django.shortcuts import render
from .models import quotation_management

from django.core.paginator import Paginator

def display_quotation(request):
    query = request.GET.get('search', '')
    sort_order = request.GET.get('order', 'asc')
    sort_order = sort_order if sort_order in ['asc', 'desc'] else 'asc'
    sort_by = request.GET.get('sort_by', 'customer__fullname')
    customer_type = request.GET.get('customer_type')
    valid_sort_fields = ['quotation_no','customer__fullname', 'quotation_date', 'total_price', 'total_price_with_gst']
    branch = request.GET.get('branch')
    sfs_representatives = request.GET.get('sfs_representatives')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    is_sales_coordinator = False

    if request.user.userprofile.role == 'admin':
        m = quotation_management.objects.all()

    elif request.user.userprofile.role == 'sales':
        sales = SalesPerson.objects.get(mobile_no = request.user.username)
        if sales.co_ordinator:
            is_sales_coordinator = True
            m = quotation_management.objects.all()
        else:
            m = quotation_management.objects.filter(contact_by_no = sales.mobile_no)

    elif request.user.userprofile.role == 'branch_manager':
        m = quotation_management.objects.filter(branch = BranchManager.objects.get(mobile_no = request.user.username).branch)
    if query:
        m = m.filter(
            Q(customer__fullname__icontains=query) |
            Q(quotation_no__icontains=query) |
            Q(customer__customerid__icontains=query) |
            Q(customer__primarycontact__icontains=query)
        )
    # Filter by customer type 
    if customer_type:
        m=m.filter(customer__customer_type=customer_type)

    if branch:
        m = m.filter(branch = branch)
    
    if sfs_representatives:
        m = m.filter(contact_by = sfs_representatives)

     # Date range filter
    if from_date:
        from_date_obj = parse_date(from_date)
        if from_date_obj:
            m = m.filter(quotation_date__gte=from_date_obj)
    if to_date:
        to_date_obj = parse_date(to_date)
        if to_date_obj:
            m = m.filter(quotation_date__lte=to_date_obj)

    filter_count = m.count()
    # if sort_by in valid_sort_fields:
    #     order_prefix = '-' if sort_order == 'desc' else ''
    #     m = m.order_by(f'{order_prefix}{sort_by}')
    # else:
    #     m = m.order_by('customer__fullname')
    m = m.order_by('-quotation_date')
    paginator = Paginator(m, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page
    branch_list = Branch.objects.all()
    sfs_representatives = quotation_management.objects.values_list("contact_by", flat=True).distinct()
    context = {
        'current_order': sort_order,
        'current_sort_by': sort_by,
        'page_obj': page_obj,
        'start_index': start_index,
        'search_query': query,
        'filter_count':filter_count,
        'branches':branch_list,
        'sfs_representatives':sfs_representatives,
        'from_date': from_date,
        'to_date': to_date,
        "querystring": request.GET.urlencode(),
        # "role": request.user.userprofile.role,
        'is_sales_coordinator': is_sales_coordinator, 
    }
    return render(request, 'display_quotation.html', context)


def get_quotation_details(request, quotation_id):
    quotation_var = get_object_or_404(quotation, id=quotation_id)
    return render(request, 'quotation_details_modal.html', {'quotation_var': quotation_var})

def display_invoice(request):
    query = request.GET.get('search', '')
    sort_order = request.GET.get('order', 'asc')
    sort_by = request.GET.get('sort_by', 'customerid')  
    
    if query:
        m = invoice.objects.filter(customer__customerid__icontains=query) & invoice.objects.filter(invoice_no__icontains=query)
    else:
        m = invoice.objects.all()
    
    if sort_by == 'firstname':
        if sort_order == 'desc':
            m = m.order_by('-customer__firstname')  
        else:
            m = m.order_by('customer__firstname')  
    else:
        if sort_order == 'desc':
            m = m.order_by('-customer__customerid')  
        else:
            m = m.order_by('customer__customerid')

    m = m.order_by('-id')
    paginator = Paginator(m, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    start_index = (page_obj.number - 1) * paginator.per_page

    context = {
        'current_order': sort_order,
        'current_sort_by': sort_by,
        'page_obj': page_obj,
        'start_index': start_index,
    }
    return render(request, 'display_invoice.html', context)


def get_invoice_details(request, invoice_id):
    invoice_var = get_object_or_404(invoice, id=invoice_id)
    return render(request, 'invoice_details_modal.html', {'invoice_var': invoice_var})





from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from django.db.models.functions import Lower
from .models import lead_management, SalesPerson
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render
from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator
from .models import lead_management, SalesPerson, main_followup
from django.contrib.auth import get_user_model

@login_required
def display_lead_management(request):
    # 1. Start with all leads
    salespersons = []
    User = get_user_model()
    if request.user.userprofile.role =='admin':
        filtered_leads = lead_management.objects.all()
        # salespersons = list(SalesPerson.objects.all())+ list(BranchManager.objects.all())

        salespersons = (
            list(SalesPerson.objects.all()) +
            list(BranchManager.objects.all()) 
        )
    elif request.user.userprofile.role == 'sales':
        salesperson = SalesPerson.objects.get(mobile_no =request.user.username)
        filtered_leads = lead_management.objects.filter(salesperson=salesperson)

    elif request.user.userprofile.role == 'branch_manager':
        branch_manager = BranchManager.objects.get(mobile_no =request.user.username)
        filtered_leads = lead_management.objects.filter(branch=branch_manager.branch)
        salespersons = SalesPerson.objects.filter(branch=branch_manager.branch)
    # 2. Get filters from request
    search_query = request.GET.get('search','').strip()
    typeoflead_filter = request.GET.get('typeoflead')
    source_filter = request.GET.get('sourceoflead')
    salesperson_filter = request.GET.get('salesperson')
    branch_filter = request.GET.get('branch')
    enquiry_from = request.GET.get('enquiry_from')
    enquiry_to = request.GET.get('enquiry_to')
    followup_from = request.GET.get('followup_from')
    followup_to = request.GET.get('followup_to')
    sort_by = request.GET.get('sort', 'customername')
    order = request.GET.get('order', 'asc')
    segment_filter = request.GET.get('segments')
    customer_type = request.GET.get('customer_type')
    followup_status = request.GET.get('followup_status', '').strip()

    # 3. Apply search filter
    if search_query:
        filtered_leads = filtered_leads.filter(
            Q(primarycontact__icontains=search_query) |
            Q(typeoflead__icontains=search_query) |
            Q(customername__icontains=search_query)
        )

    # 4. Apply other filters
    if typeoflead_filter:
        filtered_leads = filtered_leads.filter(typeoflead=typeoflead_filter)
    if source_filter:
        filtered_leads = filtered_leads.filter(sourceoflead=source_filter)
    if salesperson_filter:
        filtered_leads = filtered_leads.filter(
        Q(salesperson__full_name=salesperson_filter) |
        Q(branch_manager__full_name=salesperson_filter) |
        Q(admin__first_name=salesperson_filter)
    )
    if branch_filter:
        filtered_leads = filtered_leads.filter(branch=branch_filter)
    if enquiry_from and enquiry_to:
        filtered_leads = filtered_leads.filter(enquirydate__range=[enquiry_from, enquiry_to])
    if followup_from and followup_to:
        filtered_leads = filtered_leads.filter(firstfollowupdate__range=[followup_from, followup_to])
    if segment_filter:
        filtered_leads = filtered_leads.filter(customersegment=segment_filter)
    if customer_type:
        filtered_leads = filtered_leads.filter(customer_type=customer_type)

    if followup_status:
        # Get all follow-ups for the leads we currently have
        followups_qs = main_followup.objects.filter(
            lead__in=filtered_leads
        ).order_by('lead_id', '-created_at')

        latest_status_map = {}
        for f in followups_qs:
            if f.lead_id not in latest_status_map:  # first = latest because of ordering
                latest_status_map[f.lead_id] = f.order_status

        lead_ids_with_followup = list(latest_status_map.keys())

        if followup_status == 'Not Closed':
            # Leads with latest status = Not Closed
            leads_not_closed = [
                lid for lid, status in latest_status_map.items() if status == 'Not Closed'
            ]

            # Leads that have NO follow-up → also considered Not Closed
            leads_no_followup = list(
                filtered_leads.exclude(id__in=lead_ids_with_followup)
                            .values_list('id', flat=True)
            )

            combined_ids = set(leads_not_closed) | set(leads_no_followup)
            filtered_leads = filtered_leads.filter(id__in=combined_ids)

        else:
            # Leads whose latest followup exactly matches selected status
            leads_matching = [
                lid for lid, status in latest_status_map.items() if status == followup_status
            ]
            filtered_leads = filtered_leads.filter(id__in=leads_matching)
            
    # 5. Count for display
    branch_count = filtered_leads.count()

    # 6. Apply sorting
    leads = filtered_leads.order_by('-id')
    # 7. Apply pagination
    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    start_index = (page_obj.number - 1) * paginator.per_page

    # 8. Prepare followup data
    lead_ids = [lead.id for lead in page_obj]
    followups = main_followup.objects.filter(lead_id__in=lead_ids).order_by('lead_id', '-created_at')

    latest_followups = {}
    for followup in followups:
        if followup.lead_id not in latest_followups:
            latest_followups[followup.lead_id] = followup

    # Attach the latest follow-up to each lead
    for lead in page_obj:
        lead.latest_followup = latest_followups.get(lead.id)
        

    # 9. Get dropdown filter values
    typeoflead_choices = [choice[0] for choice in lead_management._meta.get_field('typeoflead').choices if choice[0]]
    lead_types = sorted(set(typeoflead_choices))
    source_choices = [choice[0] for choice in lead_management._meta.get_field('sourceoflead').choices if choice[0]]

    sources = sorted(set(source_choices))


    # branch_choices = [choice[0] for choice in lead_management._meta.get_field('branch').choices if choice[0]]
    # branch_used = lead_management.objects.values_list('branch', flat=True).distinct()
    # branches = sorted(set(branch_choices + list(branch_used)))

    # salespersons = SalesPerson.objects.values_list('full_name', flat=True).distinct()

    # segments = lead_management.objects.exclude( customersegment__in=["NOT SELECTED", "", None] ).values_list('customersegment', flat=True).distinct()
    segments = [ choice[0] for choice in lead_management._meta.get_field('customersegment').choices if choice[0] != "NOT SELECTED" ]

    c_types = ['Organization','Individual']
    


    # 10. No data message if needed
    no_data_message = ""
    if not leads.exists():
        if search_query:
            no_data_message = f"No data found for search: {search_query}"
        elif typeoflead_filter:
            no_data_message = f"No data found for Type of Lead: {typeoflead_filter}"
        elif source_filter:
            no_data_message = f"No data found for Source of Lead: {source_filter}"
        elif salesperson_filter:
            no_data_message = f"No data found for Salesperson: {salesperson_filter}"
        elif branch_filter:
            no_data_message = f"No data found for Branch: {branch_filter}"
        else:
            no_data_message = "No data found for the selected filters."

    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    # 11. Final context
    branches = Branch.objects.all()
    context = {
        'page_obj': page_obj,
        'start_index': start_index,
        'lead_types': lead_types,
        'sources': sources,
        'branches': branches,
        'salespersons': salespersons,
        'segments':segments,
        'search_query': search_query,
        'current_sort': sort_by,
        'current_order': order,
        'no_data_message': no_data_message,
        'branch_count': branch_count,
        'selected_segment': segment_filter,
        'c_types' : c_types,
        'query_params': query_params.urlencode(),
    }

    return render(request, 'display_lead_management.html', context)



from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
import xlwt
from .models import lead_management

def export_leads_excel(request):

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leads_full.csv"'

    writer = csv.writer(response)

    # Get all model field names
    model_fields = [field.name for field in lead_management._meta.fields]

    # Write header row
    writer.writerow([field.replace('_', ' ').title() for field in model_fields])

    # Write data rows
    for obj in lead_management.objects.all().values_list(*model_fields):
        writer.writerow(obj)

    return response




def get_lead_details(request, lead_id):
    lead = get_object_or_404(lead_management, id=lead_id)
    return render(request, 'lead_details_modal.html', {'lead': lead})

# Delete Section
# Delete Customer Details

def delete_customer(request , rid):
    
    if request.method == "POST":
        password = request.POST.get("password")
        correct_password = "seva123"  # Replace with the actual password you want to use

        if password == correct_password:
            try:
                m=customer_details.objects.filter(id=rid)
                m.delete()
                messages.success(request, "Record deleted successfully.")
            except customer_details.DoesNotExist:
                messages.error(request, "Record not found.")
        else:
            messages.error(request, "Invalid password. Deletion failed.")

    return redirect('/display_customer')

# Delete Service Management

# def delete_service_management(request , rid):
#     m=service_management.objects.filter(id=rid)

#     m.delete()

#     return redirect('/display_allocation')

# test 1
from django.contrib import messages

def delete_service_management(request, rid):
    if request.method == "POST":
        password = request.POST.get("password")
        correct_password = "seva123"  # Replace with the actual password you want to use

        if password == correct_password:
            try:
                m = service_management.objects.get(id=rid)
                m.delete()
                messages.success(request, "Record deleted successfully.")
            except service_management.DoesNotExist:
                messages.error(request, "Record not found.")
        else:
            messages.error(request, "Invalid password. Deletion failed.")

    return redirect('/display_allocation')


# Delete Quotation

def delete_quotation(request, rid):
    if request.method == "POST":
        password = request.POST.get("password")
        correct_password = "seva123"  # Ideally, move to settings or env variable for security

        if password == correct_password:
            try:
                quotation_obj = get_object_or_404(quotation_management, id=rid)
                quotation_obj.delete()
                messages.success(request, "Quotation deleted successfully.")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")
        else:
            messages.error(request, "Invalid password. Deletion failed.")
    
    return redirect('/display_quotation')
# Delete Invoice

def delete_invoice(request , rid):
    if request.method == "POST":
        password = request.POST.get("password")
        correct_password = "seva123"  # Replace with the actual password you want to use

        if password == correct_password:
            try:
                m=invoice.objects.filter(id=rid)
                m.delete()
                messages.success(request, "Record deleted successfully.")
            except invoice.DoesNotExist:
                messages.error(request, "Record not found.")
        else:
            messages.error(request, "Invalid password. Deletion failed.")

    return redirect('/display_invoice')


# Delete Inventory

# def delete_inventory(request , rid):
#     m=inventory.objects.filter(id=rid)

#     m.delete()

#     return redirect('/display_inventory')


# Delete Lead Management

def delete_lead_management(request , rid):
    if request.method == "POST":
        password = request.POST.get("password")
        correct_password = "seva123"  # Replace with the actual password you want to use

        if password == correct_password:
            try:
                m=lead_management.objects.filter(id=rid)
                m.delete()
                messages.success(request, "Record deleted successfully.")
            except lead_management.DoesNotExist:
                messages.error(request, "Record not found.")
        else:
            messages.error(request, "Invalid password. Deletion failed.")

    return redirect('/display_lead_management')



# Edit Section

# Edit Customer Details

def edit_customer(request , rid):
    if request.method =='GET':
        m=customer_details.objects.filter(id=rid)
        branches = Branch.objects.all()
        context={'branches':branches, 'data':m}
        # context['data']=m
        return render(request , 'edit_customer.html' , context)
    
    else:
        ufullname=request.POST['ufullname']
        uprimaryemail=request.POST['uprimaryemail']
        usecondaryemail=request.POST['usecondaryemail'] or ""
        uprimarycontact=request.POST['uprimarycontact']
        usecondarycontact=request.POST['usecondarycontact'] or None
        ucontactperson=request.POST['ucontactperson']
        udesignation=request.POST['udesignation']
        ushifttopartyaddress=request.POST['ushifttopartyaddress']
        ushifttopartycity=request.POST['ushifttopartycity']
        ushifttopartystate=request.POST['ushifttopartystate']
        ushifttopartypostal=request.POST['ushifttopartypostal']
        usoldtopartyaddress=request.POST['usoldtopartyaddress']
        usoldtopartycity=request.POST['usoldtopartycity']
        usoldtopartystate=request.POST['usoldtopartystate']
        usoldtopartypostal=request.POST['usoldtopartypostal']
        ucustomer_type = request.POST['ucustomer_type']
        uor_name = request.POST['uor_name']
        uor_contact = request.POST['uor_contact'] or None   
        ubranch = request.POST['branch'] 

        branch = Branch.objects.get(id = ubranch)  

        m=customer_details.objects.filter(id=rid)

        m.update(fullname=ufullname , primaryemail=uprimaryemail,  secondaryemail=usecondaryemail , primarycontact=uprimarycontact , secondarycontact=usecondarycontact , contactperson=ucontactperson , designation=udesignation , shifttopartyaddress=ushifttopartyaddress , shifttopartycity=ushifttopartycity , shifttopartystate=ushifttopartystate , shifttopartypostal=ushifttopartypostal , soldtopartyaddress=usoldtopartyaddress , soldtopartycity=usoldtopartycity , soldtopartystate=usoldtopartystate , soldtopartypostal=usoldtopartypostal, customer_type = ucustomer_type, or_name = uor_name, or_contact = uor_contact , branch = branch)

        return redirect( '/display_customer')
    



from .models import Reschedule
from datetime import datetime


@csrf_exempt
def edit_service_records(request, rid):
    service = get_object_or_404(service_management, id=rid)
    selected_products = ServiceProduct.objects.filter(service_id=service).select_related('product')
    customer = get_object_or_404(customer_details, id=service.customer_id)
    category_choices = Product.CATEGORY_CHOICES
    products = Product.objects.all()
    sales_persons = list(SalesPerson.objects.all()) + list(BranchManager.objects.all())
    frequency_choices = [str(i) for i in range(1, 13)] + ['Fortnight', 'Weekly', 'Daily']

    if request.method == "POST":
        try:
            total_price = Decimal('0.00')
            total_charges = Decimal('0.00')
            total_with_gst = Decimal('0.00')

            # --- Update existing ServiceProducts ---
            for sp in selected_products:
                price = Decimal(request.POST.get(f'price_{sp.id}', sp.price))
                quantity = Decimal(request.POST.get(f'quantity_{sp.id}', sp.quantity))
                gst = Decimal(request.POST.get(f'gst_{sp.id}', sp.gst_percentage))
                description = request.POST.get(f'description_{sp.id}', sp.description)

                line_total = price * quantity
                line_total_charges = line_total * gst / Decimal('100')
                line_total_with_gst = line_total + line_total_charges

                sp.price = price
                sp.quantity = quantity
                sp.gst_percentage = gst
                sp.description = description
                sp.total_with_gst = line_total_with_gst
                sp.save()

                total_price += line_total
                total_charges += line_total_charges
                total_with_gst += line_total_with_gst

            # --- Add new selected products ---
            selected_products_json = (request.POST.get('selected_products_json') or '[]').strip()
            try:
                selected_products = json.loads(selected_products_json)
            except json.JSONDecodeError:
                selected_products = []
                print("⚠️ Warning: Invalid JSON in selected_products_json, defaulting to empty list.")

            print("Selected Products:", selected_products)


            for item in selected_products:
                product_id = item.get('p_id')
                price = item.get('price')
                quantity = item.get('quantity')
                gst_percentage = item.get('gst', 0)
                description = item.get('description', '')

                if not price or not quantity or not gst_percentage:
                    continue

                try:
                    product = Product.objects.get(product_id=product_id)
                    price = Decimal(price)
                    quantity = Decimal(quantity)
                    gst = Decimal(gst_percentage)

                    line_total = price * quantity
                    line_total_charges = line_total * gst / Decimal('100')
                    line_total_with_gst = line_total + line_total_charges

                    ServiceProduct.objects.create(
                        service=service,
                        product=product,
                        price=price,
                        quantity=quantity,
                        gst_percentage=gst,
                        total_with_gst=line_total_with_gst,
                        description=description,
                    )

                    total_price += line_total
                    total_with_gst += line_total_with_gst

                except Product.DoesNotExist:
                    print(f"Product with ID {product_id} does not exist.")
                    continue

            # --- Save updated totals to service ---
            service.total_price = total_price
            service.total_charges = total_charges
            service.total_price_with_gst = total_with_gst
            service.contract_type = request.POST.get('contract_type')
            service.contract_status = request.POST.get('contract_status')
            service.property_type = request.POST.get('property_type')
            service.warranty_period = request.POST.get('warranty_period')
            service.state = request.POST.get('state')
            service.city = request.POST.get('city')
            service.address = request.POST.get('address')
            service.pincode = request.POST.get('pincode')
            service.gps_location = request.POST.get('gps_location')
            service.frequency_count = request.POST.get('frequency_count')
            service.sales_person_name = request.POST.get('sales_person_name')
            service.sales_person_contact_no = request.POST.get('sales_person_contact_no')
            service.lead_date = request.POST.get('lead_date')
            service.delivery_time = request.POST.get('delivery_time')
            service.service_date = request.POST.get('service_date')
            service.save()

            # messages.success(request, "Service record updated successfully.")
            return redirect('display_service_management')

        except Exception as e:
            print("Error in edit_service_records:", str(e))
            traceback.print_exc()
            messages.error(request, f"Error updating service record: {str(e)}")

    context = {
        'service': service,
        'selected_products': selected_products,
        'customer': customer,
        'category_choices': category_choices,
        'products': products,
        'sales_persons': sales_persons,
        'frequency_choices': frequency_choices,
    }
    return render(request, 'edit_service_records.html', context)

def delete_service_records(request, rid):
    service_management.objects.get(id = rid).delete()
    return redirect('display_service_management')

# This is for the delete product inside the edit field 

@csrf_exempt
def delete_service_product(request, pid):
    if request.method == 'POST':
        try:
            product = get_object_or_404(ServiceProduct, id=pid)
            product.delete()
            return JsonResponse({'success': True, 'message': 'Product deleted'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


from .signals import service_scheduled
# This function is for edit service management in reshdule 
def edit_service_management(request, rid):
    if request.method == 'GET':
        service_obj = get_object_or_404(service_management, id=rid)
        previous_reschedules = Reschedule.objects.filter(service_id=rid)
        all_technicians = TechnicianProfile.objects.all()

        # Get users already assigned as technicians (via TechWorkList)
        allocated_technicians = User.objects.filter(techworklist__service=service_obj).distinct()
   
        context = {
            'data': [service_obj],
            'previous_reschedules': previous_reschedules,
            'technicians': all_technicians,
            'selected_technicians': allocated_technicians,
        }

        return render(request, 'edit_service_management.html', context)

    else:
        service_obj = get_object_or_404(service_management, id=rid)
        technician_ids = request.POST.getlist('technicians')
        print("technician_ids :", technician_ids)
        ucustomer_id = request.POST.get('ucustomer')
        uaddress = request.POST.get('uaddress')
        utotal_price = float(request.POST['utotal_price'])
        utotal_price_with_gst = float(request.POST['utotal_price_with_gst'])
        ucontract_type = request.POST.get('ucontract_type')
        ucontract_status = request.POST.get('ucontract_status')
        uproperty_type = request.POST.get('uproperty_type')
        uwarranty_period = request.POST.get('uwarranty_period')
        ustate = request.POST.get('ustate')
        ucity = request.POST.get('ucity')
        upincode = request.POST.get('upincode')
        ugps_location = request.POST.get('ugps_location')
        ufrequency_count = request.POST.get('ufrequency_count')
        upayment_terms = request.POST.get('upayment_terms')
        usales_person_name = request.POST.get('usales_person_name')
        usales_person_contact_no = request.POST.get('usales_person_contact_no')
        udelivery_time = request.POST.get('udelivery_time')
        ulead_date = request.POST['ulead_date']
        uservice_date = request.POST['uservice_date']
        udescription = request.POST.get('work_description')
        upayment_status = request.POST.get('customer_payment_status')

        # Step 1: Ensure WorkAllocation exists for the service
        work_allocation, created = WorkAllocation.objects.get_or_create(
            service=service_obj,
            defaults={
                'fullname': service_obj.customer.fullname,
                'customer_contact': service_obj.customer.primarycontact,
                'customer_address': service_obj.address,
                'work_description': udescription or "Updated work",
                'customer_payment_status': upayment_status or "Pending",
                'payment_amount': service_obj.total_price_with_gst,
                'gps_location': service_obj.gps_location,
            }
        )

        # Step 2: Delete old TechWorkList entries
        TechWorkList.objects.filter(service=service_obj).delete()
        # Step 3: Assign new technicians
        technician_profiles = TechnicianProfile.objects.filter(id__in=technician_ids)
        work_allocation.technician.set(technician_profiles)  # update technicians for WorkAllocation
        
        for tech_profile in technician_profiles:
            tech_user = tech_profile.user  # convert to User model
            tech_work = TechWorkList.objects.create(
                technician=tech_user,
                service=service_obj,
            )
            tech_work.work.add(work_allocation)


        # Step 4: Update service fields
        try:
            customer = customer_details.objects.get(id=ucustomer_id)
        except customer_details.DoesNotExist:
            return HttpResponse("Customer not found")

        service_obj.customer = customer
        service_obj.address = uaddress
        service_obj.total_price = utotal_price
        service_obj.total_price_with_gst = utotal_price_with_gst
        service_obj.contract_type = ucontract_type
        service_obj.contract_status = ucontract_status
        service_obj.property_type = uproperty_type
        service_obj.warranty_period = uwarranty_period
        service_obj.state = ustate
        service_obj.city = ucity
        service_obj.pincode = upincode
        service_obj.gps_location = ugps_location
        service_obj.frequency_count = ufrequency_count
        service_obj.payment_terms = upayment_terms
        service_obj.sales_person_name = usales_person_name
        service_obj.sales_person_contact_no = usales_person_contact_no
        service_obj.delivery_time = udelivery_time
        service_obj.lead_date = ulead_date
        service_obj.service_date = uservice_date
        service_obj.technicians.set(technician_ids)
        service_obj.save()
        work_allocation.save()
        service_scheduled.send(sender=WorkAllocation, service_id=service_obj.id, created=False)
        return redirect('/display_allocation')

# Edit Quotation
import json
from .models import quotation_management, QuotationTerm, Product  # adjust as needed

def edit_quotation(request, rid):
    quotation = quotation_management.objects.get(id=rid)
    sales_person_list = list(SalesPerson.objects.all())+ list(BranchManager.objects.all())
    thank_notes = quotation_management.objects.values_list('thank_u_note', flat=True).distinct()
    products = Product.objects.all()
    customer = quotation.customer
    if request.method == "POST":
        delete_ids = request.POST.getlist('delete_product_ids[]')
        existing_products = quotation.product_details_json or []

        updated_products = []
        existing_product_ids = []

        # Get form fields
        customer_full_name = request.POST.get('customer_full_name')
        contact_no = request.POST.get('contact_no')
        secondary_contact_no = request.POST.get('secondary_contact_no') or None
        customer_email = request.POST.get('customer_email')
        secondary_email = request.POST.get('secondary_email')
        contact_by = request.POST.get('contact_by')
        contact_by_no = request.POST.get('contact_by_no')
        address = request.POST.get('address')
        subject = request.POST.get('subject')
        branch_id = request.POST.get('branch_id')
        selected_term_ids = request.POST.getlist('terms_and_conditions[]')  # get selected terms
        ordered_term_ids_str = request.POST.get('terms_and_conditions_ordered', '')
        custom_terms = request.POST.get('add_terms_conditions')
        customer_type = request.POST.get('customer_type')
        or_name = request.POST.get('or_name')
        or_contact = request.POST.get('or_contact') 
        thank_u_note = request.POST.get('thank_u_note')
        quotation_date = request.POST.get('quotation_date')
        customer.primarycontact = contact_no
        customer.fullname = customer_full_name
        customer.primaryemail = customer_email
        customer.secondaryemail = secondary_email
        customer.secondarycontact = secondary_contact_no
        customer.customer_type = customer_type
        customer.or_name = or_name
        try:
            customer.or_contact = int(or_contact) if or_contact not in (None, "", "None") else None
        except (TypeError, ValueError):
            customer.or_contact = None

         # Save changes
        customer.save(update_fields=[
            "primarycontact",
            "fullname",
            "secondarycontact",
            "primaryemail",
            "secondaryemail",
            "customer_type",
            "or_name",
            "or_contact",
        ])
        
        for counter , product in enumerate(existing_products, start=1):
            product_id = str(product['id'])
            if product_id in delete_ids:
                continue
            existing_product_ids.append(product_id)

            try:
                price = float(request.POST.get(f'product_price_{product_id}', product.get('price', 0)))
                quantity = float(request.POST.get(f'product_quantity_{product_id}', product.get('quantity', 0)))
                gst = float(request.POST.get(f'product_gst_{product_id}', product.get('gst', 0)))
                description = request.POST.get(f'product_description_{product_id}', product.get('description', ''))
                unit = request.POST.get(f'product_unit_{product_id}', product.get('unit', ''))
                hsn_code =  request.POST.get(f'product_hsn_{product_id}',product.get('hsn_code',''))
            except (TypeError, ValueError):
                continue

            updated_products.append({
                'id': counter,
                'p_id': product_id,
                'name': product.get('name'),
                'price': price,
                'quantity': quantity,
                'hsn_code':hsn_code,
                'gst': gst,
                'description': description,
                'unit': unit
            })

        # New products
        try:
            new_products = json.loads(request.POST.get("product_details_json", "[]"))
        except json.JSONDecodeError:
            new_products = []

        for new_product in new_products:
            if str(new_product.get('id')):
                updated_products.append(new_product)

        # Totals
        try:
            total_without_gst = float(request.POST.get("grand_total_without_gst", 0))
            total_gst = float(request.POST.get("grand_total_gst", 0))
        except ValueError:
            total_without_gst = 0
            total_gst = 0

        enable_gst = request.POST.get('enable_gst') == 'on'
        gst_type = request.POST.get('gst_type', 'cgst_sgst')

        if enable_gst:
            gst_status = 'GST'
            cgst = total_gst / 2 if gst_type == 'cgst_sgst' else 0
            sgst = total_gst / 2 if gst_type == 'cgst_sgst' else 0
            igst = total_gst if gst_type == 'igst' else 0
        else:
            gst_status = 'NON-GST'
            cgst = sgst = igst = total_gst = 0

        total_with_gst = total_without_gst + total_gst

        # Save quotation fields
        # quotation.customer_full_name = customer_full_name
        # quotation.contact_no = contact_no
        # quotation.secondary_contact_no = secondary_contact_no
        # quotation.customer_email = customer_email
        # quotation.secondary_email = secondary_email
        quotation.contact_by = contact_by
        quotation.contact_by_no = contact_by_no
        quotation.address = address
        quotation.subject = subject
        quotation.branch_id = branch_id
        quotation.quotation_date = quotation_date
        quotation.product_details_json = updated_products
        quotation.total_price = total_without_gst
        quotation.gst_total = total_gst
        quotation.total_price_with_gst = total_with_gst
        quotation.cgst = cgst
        quotation.sgst = sgst
        quotation.igst = igst
        quotation.gst_status = gst_status
        quotation.apply_gst = enable_gst
        quotation.custom_terms = custom_terms
        quotation.or_name = or_name
        try:
            quotation.or_contact = int(or_contact) if or_contact not in (None, "", "None") else None
        except (TypeError, ValueError):
            quotation.or_contact = None
        quotation.thank_u_note = thank_u_note
        quotation.save()

        # Save terms
        # quotation.terms_and_conditions.set(selected_term_ids)
        # Parse and apply ordered term IDs
        ordered_term_ids = [int(tid) for tid in ordered_term_ids_str.split(',') if tid.isdigit()]
        print('term', ordered_term_ids)
        quotation.terms_and_conditions.set(ordered_term_ids)

        quotation.terms_order = ordered_term_ids
        quotation.save()
        return redirect('display_quotation')

    else:
        terms = QuotationTerm.objects.all()
        terms_order = quotation.terms_order or []  # fallback to empty list if None

        ordered_terms = [term for tid in terms_order for term in terms if term.id == tid]
        remaining_terms = [term for term in terms if term.id not in terms_order]
        all_terms = ordered_terms + remaining_terms

        branches = Branch.objects.all()
        branch = Branch.objects.get(id = quotation.branch_id)
        
        try:
            product_details = json.loads(quotation.product_details_json)
        except Exception:
            product_details = []

        product_ids = [item['id'] for item in product_details if 'id' in item]
        selected_categories = list(
            Product.objects.filter(product_id__in=product_ids)
            .values_list('category', flat=True)
            .distinct()
        )

        return render(request, 'edit_quotation.html', {
            'quotation': quotation,
            'all_terms': all_terms,
            'branch':branch,
            'branches': branches,
            'product_details': json.dumps(product_details),
            'selected_categories': selected_categories,
            'category_choices': Product.CATEGORY_CHOICES,
            "sales_person_list": sales_person_list,
            "thank_notes": list(thank_notes),
            'products': products,
        })

# Edit Invoice



def edit_invoice(request , rid):

    if request.method =='GET':

        m=invoice.objects.filter(id=rid)

        context={}
        context['data']=m
    
        return render(request , 'edit_invoice.html' , context)
    
    else: 
        umodeofpayment = request.POST['umodeofpayment']
        udispatchedthrough = request.POST['udispatchedthrough']
        utermofdelivery = request.POST['utermofdelivery']
        utermsandcondition = request.POST['utermsandcondition']
        ucompany_name = request.POST['ucompany_name']
        ucompany_email = request.POST['ucompany_email']
        ucompany_contact_no = request.POST['ucompany_contact_no']
        # uinvoice_no = request.POST['uinvoice_no']
       
        udescription_of_goods = request.POST['udescription_of_goods']
        uhsn_sac_code = request.POST['uhsn_sac_code']
        uquantity = int(request.POST['uquantity'])
        uprice = float(request.POST['uprice'])
        udiscount =  float(request.POST['udiscount']) if request.POST['udiscount'] else None
        ugst_checkbox = True if 'ugst_checkbox' in request.POST else False
        utotal_amount = request.POST['utotal_amount']
        utotal_amount_with_gst = request.POST['utotal_amount_with_gst']
        utotal_amount_in_words = request.POST['utotal_amount_in_words']
        upan_card_no = request.POST['upan_card_no']
        uaccount_no = request.POST['uaccount_no']
        ubranch = request.POST['ubranch']
        uifsc_code = request.POST['uifsc_code']
        udispatched_date_str = request.POST['udelivery_date']
        udelivery_date_str = request.POST['udispatched_date']

        try:
            udispatched_date = datetime.strptime(udispatched_date_str, '%Y-%m-%d').date() if udispatched_date_str else timezone.now().date()
        except ValueError:
            udispatched_date = None  # Handle invalid date format
        
        try:
            udelivery_date = datetime.strptime(udelivery_date_str, '%Y-%m-%d').date() if udelivery_date_str else timezone.now().date()
        except ValueError:
            udelivery_date = None  # Handle invalid date format

        utotal_amount = uquantity * uprice
        
        udiscounted_amount = utotal_amount - (utotal_amount * (udiscount / 100))
        utotal_amount_with_gst = udiscounted_amount * 1.18 if ugst_checkbox else udiscounted_amount

       
        

        m=invoice.objects.filter(id=rid)

        m.update(modeofpayment=umodeofpayment, dispatchedthrough=udispatchedthrough , termofdelivery=utermofdelivery,  termsandcondition=utermsandcondition , company_name=ucompany_name , company_email=ucompany_email , company_contact_no=ucompany_contact_no  , dispatched_date=udispatched_date, description_of_goods=udescription_of_goods,hsn_sac_code=uhsn_sac_code ,quantity =uquantity,price=uprice, discount=udiscount, gst_checkbox=ugst_checkbox,total_amount=utotal_amount ,total_amount_with_gst =utotal_amount_with_gst, total_amount_in_words=utotal_amount_in_words ,pan_card_no=upan_card_no , account_no=uaccount_no , branch=ubranch ,ifsc_code=uifsc_code ,delivery_date =udelivery_date )

       
        return redirect( '/display_invoice')


# Edit Lead Management


from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from datetime import datetime
from .models import lead_management

def edit_lead_management(request, rid):
    if request.method =='GET':

        m=lead_management.objects.filter(id=rid)
        salesperson = SalesPerson.objects.all()
        context={'salespersons':salesperson}
        context['data']=m
    
        return render(request , 'edit_lead_management.html' , context)
    
    elif request.method == 'POST':
        role = request.user.userprofile.role
        usourceoflead = request.POST.get('usourceoflead', '')
        # usalesperson = request.POST.get('usalesperson')
        ucustomername = request.POST.get('ucustomername', '')
        ucustomersegment = request.POST.get('ucustomersegment', '')
        utypeoflead = request.POST.get('utypeoflead', '')
        ucontactedby = request.POST.get('ucontactedby', '')
        uenquirydate = request.POST.get('uenquirydate', '')
        ucustomer_type = request.POST.get('ucustomer_type','')
        uor_name = request.POST.get('uor_name') or None
        uor_contact = request.POST.get('uor_contact') or None


        try:
            uenquirydate = datetime.strptime(uenquirydate, '%Y-%m-%d').date() if uenquirydate else timezone.now().date()
        except ValueError:
            uenquirydate = None  # Handle invalid date format

        umaincategory = request.POST.get('umaincategory', '')
        usubcategory = request.POST.get('usubcategory', '')
        uprimarycontact = request.POST.get('uprimarycontact', '')
        usecondarycontact_raw = request.POST.get('usecondarycontact')
        usecondarycontact = int(usecondarycontact_raw) if usecondarycontact_raw else None
        ucustomeremail = request.POST.get('ucustomeremail', '')
        ucustomeraddress = request.POST.get('ucustomeraddress', '')
        ulocation = request.POST.get('ulocation', '')       
        ucity = request.POST.get('ucity', '')
        ufirstfollowupdate = request.POST.get('ufirstfollowupdate', '')

        try:
            ufirstfollowupdate = datetime.strptime(ufirstfollowupdate, '%Y-%m-%d').date() if ufirstfollowupdate else timezone.now().date()
        except ValueError:
            ufirstfollowupdate = None  # Handle invalid date format

        m=lead_management.objects.filter(id=rid)
        sp = None
        bm = None
        ad = None
        if role == "sales":
            sp = SalesPerson.objects.get(mobile_no = request.user.username)
        elif role == "branch_manager":
            bm = BranchManager.objects.get(mobile_no = request.user.username)
        else: 
            ad = request.user

        m.update(
            sourceoflead = usourceoflead,
            salesperson=sp,
            branch_manager = bm,
            admin = ad,
            customername = ucustomername,
            customersegment = ucustomersegment,
            typeoflead = utypeoflead,
            contactedby = ucontactedby,
            enquirydate = uenquirydate,
            maincategory = umaincategory,
            subcategory = usubcategory,
            primarycontact = uprimarycontact,
            secondarycontact = usecondarycontact,
            customeremail = ucustomeremail,
            customeraddress = ucustomeraddress,
            location = ulocation,
            city = ucity,
            firstfollowupdate=ufirstfollowupdate,
            customer_type = ucustomer_type,
            or_name = uor_name,
            or_contact = uor_contact,
        )
        
       
        return redirect('/display_lead_management')



 
def search_inventory(request):
    query = request.GET.get('q', '')
    if query:
        results = Inventory_summary.objects.filter(
            Q(customerid__icontains=query) | Q(customername__icontains=query)
        )
    else:
        results = Inventory_summary.objects.all()
    
    return render(request, 'search_inventory.html', {'results': results})



def search(request):
    search_query = request.GET.get('q', '').strip()
    sort_field = request.GET.get('sort', 'firstname')  # Default sort by 'firstname'
    
    # Determine the sort order
    if sort_field.startswith('-'):
        sort_order = 'desc'
        sort_field = sort_field[1:]  # Remove the '-' to get the actual field name
    else:
        sort_order = 'asc'

    results = customer_details.objects.all()  # Start with all records

    if search_query:
        results = results.filter(
            Q(firstname__icontains=search_query) |
            Q(lastname__icontains=search_query) |
            Q(customerid__icontains=search_query)
        )
    
    # Apply sorting
    results = results.order_by(f"{'-' if sort_order == 'desc' else ''}{sort_field}")

    # Pagination
    paginator = Paginator(results, 10)  # Show 10 results per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'search_query': search_query,
        'page_obj': page_obj,
        'sort_order': sort_order,
        'sort_field': sort_field,
    }

    if not results.exists():
        context['message'] = "No search results found."
    
    return render(request, 'search.html', context)

# In crmapp/views.py
from django.shortcuts import render
from .forms import InventoryServiceForm , CustomerImportForm
from .models import customer_details, Product, Inventory_summary

def inventory_service(request):
    if request.method == 'POST':
        form = InventoryServiceForm(request.POST)
        if form.is_valid():
            customer_id = form.cleaned_data['customer_id']
            customer_name = form.cleaned_data['customer_name']
            product_quantities = {
                form.cleaned_data['p1']: form.cleaned_data['p1_quantity'],
                form.cleaned_data.get('p2'): form.cleaned_data.get('p2_quantity'),
                form.cleaned_data.get('p3'): form.cleaned_data.get('p3_quantity'),
                form.cleaned_data.get('p4'): form.cleaned_data.get('p4_quantity'),
            }

            inventory_entries = []
            total_price = 0

            for product, quantity in product_quantities.items():
                if product and quantity:
                    product_instance = Product.objects.get(product_id=product.product_id)
                    if product_instance.quantity >= quantity:
                        product_instance.quantity -= quantity
                        product_instance.save()

                        inventory_entry = Inventory_summary.objects.create(
                            customer_id=customer_id,
                            customer_name=customer_name,
                            product=product_instance,
                            quantity=quantity,
                            total_price=product_instance.price * quantity
                        )
                        inventory_entries.append(inventory_entry)
                        total_price += inventory_entry.total_price
                    else:
                        return render(request, 'inventory_service.html', {'form': form, 'error': f'Not enough {product_instance.product_name} in stock.'})

            return render(request, 'inventory_summary_result.html', {'customer_id': customer_id, 'customer_name': customer_name, 'inventory_entries': inventory_entries, 'total_price': total_price})
    else:
        form = InventoryServiceForm()
    
    return render(request, 'inventory_service.html', {'form': form})

from django.shortcuts import render
from .models import Inventory_summary, customer_details

def inventory_summary(request):
    customers = customer_details.objects.all()
    inventory_summary = []

    for customer in customers:
        # Filter entries by the correct customer ID
        entries = Inventory_summary.objects.filter(customer_id=customer.customerid)
        
        # Check if entries exist and calculate total price
        if entries.exists():
            total_price = sum(entry.total_price for entry in entries)
        else:
            total_price = 0

        inventory_summary.append({
            'customer': customer,
            'entries': entries,
            'total_price': total_price
        })

    context = {
        'inventory_summary': inventory_summary,
    }
    
    return render(request, 'inventory_summary.html', context)


# @login_required
# def add_product(request):
#     from_quotation = request.GET.get('from_quotation') == 'true'

#     if request.method == 'POST':
#         form = AddProductForm(request.POST)
#         if form.is_valid():
#             form.save()

#             if from_quotation:
#                 # Redirect back to quotation form (retain user input via session or back button)
#                 return redirect(reverse('create_quotation'))
#             else:
#                 return render(request, 'add_product_success.html')
#     else:
#         form = AddProductForm()

#     # Render either as a modal or full page based on access
#     return render(request, 'add_product.html', {
#         'form': form,
#         'from_quotation': from_quotation
#     })

@login_required
def add_product(request):
    from_quotation = request.GET.get('from_quotation') == 'true'

    if request.method == 'POST':
        form = AddProductForm(request.POST)
        if form.is_valid():
            product = form.save()

            # If AJAX -> return JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'product_id': product.id, 'product_name': product.name})

            if from_quotation:
                return redirect(reverse('create_quotation'))
            else:
                return render(request, 'add_product_success.html')
    else:
        form = AddProductForm()

    # If AJAX -> return only the form HTML
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'add_product_form.html', {'form': form})

    # Otherwise, render full page
    return render(request, 'add_product.html', {'form': form, 'from_quotation': from_quotation})



@login_required
def update_product(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)
    form = UpdateProductForm(request.POST or None, instance=product)

    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('/products/')  

    return render(request, 'update_product.html', {'form': form})

@login_required
@role_required(['admin','sales','branch_manager','operation_person'])
def create_technician_profile(request):
    if not request.user.is_staff:
        return redirect('not_authorized')

    if request.method == 'POST':
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        contact_number = request.POST.get('contact_number')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        postal_code = request.POST.get('postal_code')
        date_of_joining = request.POST.get('date_of_joining')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(id = int(branch_id))
        errors = {}

        if TechnicianProfile.objects.filter(contact_number=contact_number).exists():
            errors['contact_number'] = 'Contact number already exists'

        if TechnicianProfile.objects.filter(email=email).exists():
            errors['email'] = 'Email already exists'

        if User.objects.filter(username=contact_number).exists():
            errors['contact_number'] = 'Contact number (username) already exists'

        if password != confirm_password:
            errors['password'] = 'Passwords do not match'

        if errors:
            return render(request, 'create_technician_profile.html', {
                'errors': errors,
                'form_data': request.POST
            })

        # ✅ Create Django User
        username = contact_number
        user = User.objects.create_user(
            username=username,
            password=password,
            email=email,
            first_name=first_name,
            last_name=last_name
        )

        # ✅ Assign role to user via UserProfile
       # update the auto-created profile
        user_profile = user.userprofile  
        user_profile.role = 'technician'
        user_profile.phone = contact_number
        user_profile.save()


        # ✅ Create the TechnicianProfile linked to the user
        TechnicianProfile.objects.create(
            user=user,
            first_name=first_name,
            last_name=last_name,
            email=email,
            contact_number=contact_number,
            address=address,
            city=city,
            state=state,
            postal_code=postal_code,
            date_of_joining=date_of_joining,
            branch = branch
        )

        return redirect('/technicians')  # Replace with your success page or named URL
    branches = Branch.objects.all()
    return render(request, 'create_technician_profile.html',{'branches':branches})




from django.shortcuts import render, get_object_or_404, redirect
from .models import TechnicianProfile
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.hashers import make_password

from django.shortcuts import render, get_object_or_404, redirect
from .models import TechnicianProfile

# Display technicians with demo passwords
@login_required
@role_required(['admin','branch_manager','operation_person'])
def display_technician(request):
    if request.user.userprofile.role == 'admin':
        technicians = TechnicianProfile.objects.all()
    elif request.user.userprofile.role == 'branch_manager':
        branch_manager = BranchManager.objects.get(mobile_no = request.user.username)
        technicians = TechnicianProfile.objects.filter(branch = branch_manager.branch)
    elif request.user.userprofile.role == 'operation_person':
        operation_person = OperationPerson.objects.get(user=request.user)
        technicians = TechnicianProfile.objects.filter(branch=operation_person.branch)
    # Manual dictionary to simulate decrypted passwords for each technician
    # This is just for demonstration/testing purposes — do not use in production!
    password_map = {
        tech.id: f"password{tech.id}"  # Replace with actual known plain passwords if you have them
        for tech in technicians
    }

    # Combine technician info with their 'plain' passwords
    technician_data = []
    for tech in technicians:
        technician_data.append({
            'id': tech.id,
            'first_name': tech.first_name,
            'last_name': tech.last_name,
            'email': tech.email,
            'contact_number': tech.contact_number,
            'city': tech.city,
            'date_of_joining': tech.date_of_joining,
            'password': password_map.get(tech.id, 'Unknown')
        })

    return render(request, 'display_technician.html', {'technician_data': technician_data})


# Edit technician
@login_required
@role_required(['admin','sales','branch_manager','operation_person'])
def edit_technician(request, technician_id):
    technician = get_object_or_404(TechnicianProfile, id=technician_id)
    user = technician.user
    branches = Branch.objects.all()
    if request.method == 'POST':
        # Collect POST data
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        contact_number = request.POST.get('contact_number')
        address = request.POST.get('address')
        city = request.POST.get('city')
        state = request.POST.get('state')
        postal_code = request.POST.get('postal_code')
        date_of_joining = parse_date(request.POST.get('date_of_joining'))
        password = request.POST.get('password')
        branch_id = request.POST.get('branch')
        branch = Branch.objects.get(id = int(branch_id))
        # ✅ Create or update User
        if user:
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.username = contact_number
            if password:
                user.set_password(password)
            user.save()
        else:
            user = User.objects.create_user(
                username=contact_number,
                password=password or User.objects.make_random_password(),
                email=email,
                first_name=first_name,
                last_name=last_name
            )
            technician.user = user  # Link new user to technician

        # ✅ Create or update UserProfile
        user_profile, created = UserProfile.objects.get_or_create(user=user)
        user_profile.role = 'technician'
        user_profile.phone = contact_number
        user_profile.save()

        # ✅ Update TechnicianProfile
        technician.first_name = first_name
        technician.last_name = last_name
        technician.email = email
        technician.contact_number = contact_number
        technician.address = address
        technician.city = city
        technician.state = state
        technician.postal_code = postal_code
        technician.date_of_joining = date_of_joining
        technician.branch = branch
        technician.save()

        messages.success(request, "Technician updated successfully.")
        return redirect('display_technician')

    return render(request, 'edit_technician.html', {'technician': technician,'branches':branches})
# Delete technician
@role_required(['admin','sales','branch_manager','operation_person'])
def delete_technician(request, technician_id):
    technician = get_object_or_404(TechnicianProfile, id=technician_id)
    user = technician.user
    technician.delete()
    user.delete()
    messages.success(request, "Technician deleted successfully.")
    return redirect('display_technician')


def not_authorized(request):
    return render(request, 'not_authorized.html')


# def technician_login(request):
#     if request.method == 'POST':
#         contact_number = request.POST.get('contact_number')
#         password = request.POST.get('password')
        
#         user = authenticate(request, username=contact_number, password=password)
#         if user is not None:
#             login(request, user)
#             return redirect('technician_dashboard')
#         else:
#             return render(request, 'technician_login.html', {'error': 'Invalid login credentials'})

#     return render(request, 'technician_login.html')

def technician_login(request):
    if request.method == 'POST':
        contact_number = request.POST.get('contact_number')
        password = request.POST.get('password')
        
        user = authenticate(request, username=contact_number, password=password)

        if user is not None:
            try:
                if user.userprofile.role == 'technician':
                    login(request, user)
                    return redirect('technician_dashboard')
                else:
                    error = 'Access denied: Not a technician.'
            except UserProfile.DoesNotExist:
                error = 'User profile not found.'
        else:
            error = 'Invalid login credentials.'
        
        return render(request, 'technician_login.html', {'error': error})

    return render(request, 'technician_login.html')


# @login_required
# def technician_dashboard(request):
#     user = request.user
#     try:
#         technician_profile = TechnicianProfile.objects.get(user=user)
#     except TechnicianProfile.DoesNotExist:
#         technician_profile = None

#     context = {
#         'user': user,
#         'technician_profile': technician_profile,
#     }

#     works = WorkAllocation.objects.all()


#     print('works: ',works)
#     for work in works:
#         if work.status == 'Pending':
#             work.status = 'workdesk'
#             work.save()
#             TechWorkList.objects.create(technician=request.user, work=work)
    
#     return render(request, 'technician_dashboard.html', context)
import json
from django.utils.timezone import now, timedelta
from calendar import monthrange
from django.db.models import Count
from django.utils.timezone import make_aware
from datetime import datetime


# def technician_dashboard(request):
#     user = request.user
#     try:
#         technician_profile = TechnicianProfile.objects.get(user=user)
#     except TechnicianProfile.DoesNotExist:
#         technician_profile = None


#     techworklistobj = TechWorkList.objects.all()
#     for i in techworklistobj:
#         print('tech',i.technician.first_name,'completion time',i.completion_datetime )


#     works = WorkAllocation.objects.all()
#     for work in works:
#         if work.status == 'Pending':
#             work.status = 'workdesk'
#             work.save()
#             TechWorkList.objects.create(technician=request.user, work=work)

#     # Get the current date or use the month and year from query parameters
#     query_month = request.GET.get('month')
#     query_year = request.GET.get('year')

#     if query_month and query_year:
#         selected_month = int(query_month)
#         selected_year = int(query_year)
#     else:
#         today = now()
#         selected_month = today.month
#         selected_year = today.year

#     # Get the start and end dates for the selected month
#     from django.utils.timezone import make_aware

#     print('selected month: ', selected_month)
#     print('selected year',selected_year)
#     # Make the start and end dates timezone-aware
#     first_day_of_month = make_aware(datetime(selected_year, selected_month, 1))
#     last_day_of_month = make_aware(datetime(selected_year, selected_month, monthrange(selected_year, selected_month)[1], 23, 59, 59))

#     print("first day",first_day_of_month)
#     print('last day of month', last_day_of_month)


#     # Filter works for the selected month
#     print('Filter Range:', first_day_of_month, '-', last_day_of_month)

#     completed_works = TechWorkList.objects.filter(
#         technician=user,
#         status='Completed',
#         completion_datetime__range=[first_day_of_month, last_day_of_month]
#     )

#     print('Completed Works:', completed_works)

#     # Group works by week
#     weekly_work_counts = {}
#     current_date = first_day_of_month
#     while current_date <= last_day_of_month:
#         week_start = current_date
#         week_end = week_start + timedelta(days=6)

#         week_label = f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b')}"
#         count = completed_works.filter(
#             completion_datetime__range=[week_start, min(week_end, last_day_of_month)]
#         ).count()

#         weekly_work_counts[week_label] = count
#         print(f"Week: {week_label}, Count: {count}")  # Debugging output
#         current_date = week_end + timedelta(days=1)

#     # Prepare data for the chart
#     labels = list(weekly_work_counts.keys())
#     data = list(weekly_work_counts.values())


#     chart_data = {
#         'labels': labels,
#         'data': data
#     }
#     chart_data_json = json.dumps(chart_data)

#     # Determine previous and next months for navigation
#     previous_month = first_day_of_month - timedelta(days=1)
#     next_month = last_day_of_month + timedelta(days=1)
#     if "notifications" not in request.session:
#          request.session["notifications"] = [
#              {"title": "Work Assigned", "message": "You have a new job today", "timestamp": now().strftime("%d %b %Y %H:%M")},
#              {"title": "Reminder", "message": "Check your completed jobs for this week", "timestamp": now().strftime("%d %b %Y %H:%M")},
#          ]

#     notifications = request.session.get("notifications", [])

#     context = {
#         'user': user,
#         'technician_profile': technician_profile,
#         'chart_data_json': chart_data_json,  # Pass chart data to the template
#         'selected_month': selected_month,
#         'selected_year': selected_year,
#         'previous_month': previous_month,
#         'next_month': next_month,
#         "notifications": notifications,
#         "notifications_count": len(notifications),
#     }


#     return render(request, 'technician_dashboard.html', context)

@login_required
@role_required(['technician'])
def technician_dashboard(request):
    user = request.user
    try:
        technician_profile = TechnicianProfile.objects.get(user=user)
    except TechnicianProfile.DoesNotExist:
        technician_profile = None

    # 👇 Don’t re-create TechWorkList here blindly
    # Just display existing work assignments

    # Get current month/year from query params or today
    query_month = request.GET.get('month')
    query_year = request.GET.get('year')

    if query_month and query_year:
        selected_month = int(query_month)
        selected_year = int(query_year)
    else:
        today = now()
        selected_month = today.month
        selected_year = today.year

    # Make timezone-aware dates
    from django.utils.timezone import make_aware
    first_day_of_month = make_aware(datetime(selected_year, selected_month, 1))
    last_day_of_month = make_aware(datetime(
        selected_year,
        selected_month,
        monthrange(selected_year, selected_month)[1],
        23, 59, 59
    ))

    # ✅ Completed works this month
    completed_works = TechWorkList.objects.filter(
        technician=user,
        status='Completed',
        completion_datetime__range=[first_day_of_month, last_day_of_month]
    )

    # ✅ Weekly counts
    weekly_work_counts = {}
    current_date = first_day_of_month
    while current_date <= last_day_of_month:
        week_start = current_date
        week_end = week_start + timedelta(days=6)
        week_label = f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b')}"
        count = completed_works.filter(
            completion_datetime__range=[week_start, min(week_end, last_day_of_month)]
        ).count()
        weekly_work_counts[week_label] = count
        current_date = week_end + timedelta(days=1)

    # ✅ Chart data
    chart_data = {
        'labels': list(weekly_work_counts.keys()),
        'data': list(weekly_work_counts.values())
    }
    chart_data_json = json.dumps(chart_data)

    # ✅ Notifications directly from TechWorkList
    notifications = TechWorkList.objects.filter(
        technician=user,
        is_notified=True
    )
    notifications_count = notifications.count()

    context = {
        'user': user,
        'technician_profile': technician_profile,
        'chart_data_json': chart_data_json,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'previous_month': first_day_of_month - timedelta(days=1),
        'next_month': last_day_of_month + timedelta(days=1),
        "notifications": notifications,
        "notifications_count": notifications_count,
    }
    

    return render(request, 'technician_dashboard.html', context)

@csrf_exempt
def clear_notifications(request):
    if request.method == "POST":
        TechWorkList.objects.filter(
            technician=request.user,
            is_notified=True
        ).update(is_notified=False)
        return JsonResponse({"status": "cleared"})
    return JsonResponse({"status": "invalid request"}, status=400)


def create_superadmin(request):
    # List of superadmin details
    superadmins = [
        {
            'username': 'admin1',
            'email': 'superadmin1@example.com',
            'password': 'admin1'
        },
        {
            'username': 'admin2',
            'email': 'superadmin2@example.com',
            'password': 'admin2'
        }
    ]
    
    for admin in superadmins:
        # Check if the superadmin already exists
        if not User.objects.filter(username=admin['username']).exists():
            # Create a superuser with hardcoded values
            User.objects.create_superuser(
                username=admin['username'],
                email=admin['email'],
                password=admin['password']
            )
    
    return HttpResponse('Superadmins created successfully.')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import TechnicianProfile, WorkAllocation, service_management
# from .models import service_management, TechnicianProfile, WorkAllocation, TechWorkList
@login_required
def allocate_work(request, service_id):
    service_object = get_object_or_404(service_management, id=service_id)

    customer_fullname = service_object.customer.fullname
    customer_address = service_object.address
    customer_city = service_object.city
    customer_state = service_object.state
    customer_pincode = service_object.pincode
    customer_contact = service_object.customer.primarycontact
    payment_amount = service_object.total_price_with_gst
    gps_location = service_object.gps_location
    
    customer_address = f"{customer_address}, {customer_city}, {customer_state} - {customer_pincode}"

    
    if request.method == 'POST':
        technician_ids = request.POST.getlist('technicians')
        customer_address = request.POST.get('customer_address')
        work_description = request.POST.get('work_description')
        customer_payment_status = request.POST.get('customer_payment_status')

        # Create the work allocation
        work_allocation = WorkAllocation.objects.create(
            service=service_object,
            fullname=customer_fullname,
            customer_contact=customer_contact,
            customer_address=customer_address,
            work_description=work_description,
            customer_payment_status=customer_payment_status,
            payment_amount=payment_amount,
            gps_location=gps_location,
        )
   
        # Add all technicians to the WorkAllocation
        technicians = TechnicianProfile.objects.filter(id__in=technician_ids)
        work_allocation.technician.set(technicians)

        # Create TechWorkList for each technician
        for tech in technicians:
            tech_worklist = TechWorkList.objects.create(
                technician=tech.user,  
                service=service_object,
            )
            tech_worklist.work.add(work_allocation)
       
        return redirect('work_allocation_success')

    technicians = TechnicianProfile.objects.all()
    context = {
        'technicians': technicians,
        'customer_fullname': customer_fullname,
        'customer_address' : customer_address,
        'customer_contact': customer_contact,
        'payment_amount': payment_amount,
        'gps_location': gps_location,
    }
    return render(request, 'allocate_work.html', context)





from crmapp.models import Reschedule, service_management

def reschedule_create(request, service_id):
    service = get_object_or_404(service_management, id=service_id)
    if request.method == "POST":
        # Retrieve data from POST request
        reason = request.POST.get('reason', '').strip()
        if not reason:
            return render(request, 'reschedule_reason.html', {
                'error': 'Please provide a reason for rescheduling.',
                'service': service
            })

        # Log the reason in RescheduleLog
        Reschedule.objects.create(
            service=service,
            old_service_date=service.service_date,
            old_delivery_time=service.delivery_time,
            reason=reason
        )

        # Redirect to the edit form
        return redirect('edit_service_management', rid=service.id)

    return render(request, 'reschedule.html', {'service': service})


from django.core.paginator import Paginator

def technician_work_list(request):
    if not request.user.is_staff:
        return HttpResponse("Not authorized", status=403)

    search = request.GET.get("search",'').strip()
    payment_status = request.GET.get('payment_status','')
    branch = request.GET.get('branch','')
    work_status = request.GET.get('work_status')
    technician = request.GET.get('technician','')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    # Fetch work allocations
    user_profile = request.user.userprofile
    if user_profile.role == "admin":
        work_allocations = WorkAllocation.objects.all()

    elif user_profile.role == 'branch_manager':
        branch_manager = BranchManager.objects.get(mobile_no = request.user.username)
        work_allocations = WorkAllocation.objects.filter(service__branch = branch_manager.branch)

    elif user_profile.role == 'operation_person':
        operation_person = OperationPerson.objects.get(user = request.user)
        work_allocations = WorkAllocation.objects.filter(service__branch = operation_person.branch)

    # Apply search filter
    if search:
        work_allocations = work_allocations.filter(
            Q(service__customer__fullname__icontains=search) |
            Q(service__customer__primarycontact__icontains=search) 
        )

    if payment_status:
        work_allocations = work_allocations.filter(customer_payment_status = payment_status)
  
    if branch:
        work_allocations = work_allocations.filter(service__branch = branch )

    if work_status:
        work_allocations = work_allocations.filter(status = work_status)

    if technician :
        work_allocations = work_allocations.filter(technician = technician)

    if from_date:
        from_date_obj = parse_date(from_date)
        if from_date_obj:
            work_allocations = work_allocations.filter(created_at__gte = from_date_obj)
    if to_date:
        to_date_obj = parse_date(to_date)
        if to_date_obj:
            # Add +1 day so the filter includes the whole 'to_date'
            work_allocations = work_allocations.filter(created_at__lt=to_date_obj + timedelta(days=1))
    # Set up pagination (10 items per page)
    paginator = Paginator(work_allocations, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    payment_status_choices  = WorkAllocation.objects.values_list('customer_payment_status', flat=True).distinct()
    w_status = WorkAllocation.objects.values_list('status', flat=True).distinct()
    branch = Branch.objects.all()
    tech = TechnicianProfile.objects.all()
    context = {'page_obj': page_obj, 
               'search': search,
                "payment_status": payment_status, 
                "payment_status_choices":payment_status_choices ,
                "branch":branch,
                "w_status":w_status,
                'tech':tech,
                "querystring": request.GET.urlencode(),
             }
    return render(request, 'technician_work_list.html', context)

from openpyxl.utils import get_column_letter
def export_technician_work_list(request):
    if not request.user.is_staff:
        return HttpResponse("Not authorized", status=403)

    search = request.GET.get("search", '').strip()
    payment_status = request.GET.get('payment_status', '')
    branch = request.GET.get('branch', '')
    work_status = request.GET.get('work_status')
    technician = request.GET.get('technician', '')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    work_allocations = WorkAllocation.objects.all()

    # Apply same filters as list view
    if search:
        work_allocations = work_allocations.filter(
            Q(service__customer__fullname__icontains=search) |
            Q(service__customer__primarycontact__icontains=search)
        )
    if payment_status:
        work_allocations = work_allocations.filter(customer_payment_status=payment_status)
    if branch:
        work_allocations = work_allocations.filter(service__branch=branch)
    if work_status:
        work_allocations = work_allocations.filter(status=work_status)
    if technician:
        work_allocations = work_allocations.filter(technician=technician)
    if from_date:
        from_date_obj = parse_date(from_date)
        if from_date_obj:
            work_allocations = work_allocations.filter(created_at__gte=from_date_obj)
    if to_date:
        to_date_obj = parse_date(to_date)
        if to_date_obj:
            work_allocations = work_allocations.filter(created_at__lt=to_date_obj + timedelta(days=1))

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Work Allocations"

    # Header row
    headers = ["Customer", "Mobile", "Branch", "Technician", "Status", "Payment Status", "Created At"]
    ws.append(headers)

    # Data rows
    for wa in work_allocations:
        ws.append([
            wa.service.customer.fullname if wa.service and wa.service.customer else "",
            wa.service.customer.primarycontact if wa.service and wa.service.customer else "",
            wa.service.branch.branch_name if wa.service and wa.service.branch else "",
            ", ".join([f"{t.first_name} {t.last_name}" for t in wa.technician.all()]) if wa.technician.exists() else "",
            wa.status,
            wa.customer_payment_status,
            wa.created_at.strftime("%Y-%m-%d %H:%M") if wa.created_at else "",
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Return response as Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="technician_work_list.xlsx"'
    wb.save(response)
    return response

def edit_work(request, work_id):
    work_allocation = get_object_or_404(WorkAllocation, id=work_id)
    
    if request.method == 'POST':
        work_allocation.customer_name = request.POST.get('customer_name')
        work_allocation.work_description = request.POST.get('work_description')
        work_allocation.customer_payment_status = request.POST.get('customer_payment_status')
        work_allocation.payment_amount = request.POST.get('payment_amount')
        work_allocation.save()
        return redirect('technician_work_list')
    
    return render(request, 'edit_work.html', {'work_allocation': work_allocation})

def delete_work(request, work_id):
    if request.method == 'POST':
        work_allocation = get_object_or_404(WorkAllocation, id=work_id)
        work_allocation.delete()
        return redirect('technician_work_list')
    return HttpResponse("Method not allowed", status=405)




@login_required
def handle_work_allocation(request, work_id):
    work_allocation = get_object_or_404(WorkAllocation, id=work_id)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        if status in ['Accepted', 'Rejected']:
            work_allocation.status = status
            work_allocation.save()
        return redirect('technician_work_list')
    
    return render(request, 'handle_work_allocation.html', {'work_allocation': work_allocation})

def work_allocation_success(request):
    return render(request, 'work_allocation_success.html')

@login_required
def pending_work(request):
    try:
        technician_profile = TechnicianProfile.objects.get(user=request.user)
    except TechnicianProfile.DoesNotExist:
        technician_profile = None

    pending_works = WorkAllocation.objects.filter(technician=technician_profile, status='Pending')

    # work_allocations = TechWorkList.objects.filter(technician=request.user)
    return render(request, 'pending_work.html', {'pending_works': pending_works })

    # return render(request, 'pending_work.html', {'pending_works': pending_works})


# views.py
# from django.shortcuts import redirect, render
# from paypalrestsdk import Payment
# from django.conf import settings
# import paypalrestsdk

# # Import the PayPal SDK configuration
# from .paypal import paypalrestsdk

# def create_paypal_payment(request):
#     if request.method == 'POST':
#         # Example payment details
#         payment = paypalrestsdk.Payment({
#             "intent": "sale",
#             "payer": {
#                 "payment_method": "paypal"
#             },
#             "redirect_urls": {
#                 "return_url": "http://localhost:8000/payment-success",
#                 "cancel_url": "http://localhost:8000/payment-cancel"
#             },
#             "transactions": [{
#                 "item_list": {
#                     "items": [{
#                         "name": "Quotation Payment",
#                         "sku": "12345",
#                         "price": "10.00",
#                         "currency": "USD",
#                         "quantity": 1
#                     }]
#                 },
#                 "amount": {
#                     "total": "10.00",
#                     "currency": "USD"
#                 },
#                 "description": "Payment for quotation."
#             }]
#         })

#         # Create the payment
#         if payment.create():
#             for link in payment.links:
#                 if link.rel == "approval_url":
#                     # Redirect the user to PayPal for authorization
#                     return redirect(link.href)
#         else:
#             print(payment.error)

#     return render(request, 'checkout.html')

# views.py
from django.shortcuts import redirect, render
from paypalrestsdk import Payment
from django.conf import settings
import paypalrestsdk

# Import the PayPal SDK configuration
from .paypal import paypalrestsdk

def create_paypal_payment(request):
    if request.method == 'POST':
        # Extracting invoice details from POST data or database
        customer_id = request.POST.get('customer_id')  # Assuming this is sent via POST
        total_amount_with_gst = request.POST.get('total_amount_with_gst')  # Total amount with GST
        company_name = request.POST.get('company_name')  # Assuming other params are also passed

        # Ensure that `total_amount_with_gst` is properly formatted
        total_amount_with_gst = format(float(total_amount_with_gst), '.2f')

        # Example payment details with dynamic data
        payment = paypalrestsdk.Payment({
            "intent": "sale",
            "payer": {
                "payment_method": "paypal"
            },
            "redirect_urls": {
                "return_url": "http://localhost:8000/payment-success",
                "cancel_url": "http://localhost:8000/payment-cancel"
            },
            "transactions": [{
                "item_list": {
                    "items": [{
                        "name": f"Invoice Payment for {company_name}",
                        "sku": customer_id,  # Using Customer ID as SKU
                        "price": total_amount_with_gst,  # Price in INR
                        "currency": "USD",  # Currency set to INR
                        "quantity": 1
                    }]
                },
                "amount": {
                    "total": total_amount_with_gst,  # Total amount from invoice
                    "currency": "USD"  # Set currency to INR
                },
                "description": f"Payment for Invoice from {company_name}."
            }]
        })

        # Create the payment
        if payment.create():
            for link in payment.links:
                if link.rel == "approval_url":
                    # Redirect the user to PayPal for authorization
                    return redirect(link.href)
        else:
            print(payment.error)

    return render(request, 'checkout.html')



# views.py
def payment_success(request):
    payment_id = request.GET.get('paymentId')
    payer_id = request.GET.get('PayerID')

    payment = paypalrestsdk.Payment.find(payment_id)

    if payment.execute({"payer_id": payer_id}):
        # Payment was successful
        return render(request, 'payment_success.html')
    else:
        # Payment failed
        return render(request, 'payment_error.html')


def payment_cancel(request):
    # Handle cancellation
    return render(request, 'payment_cancel.html')


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import WorkAllocation, TechWorkList
from django.contrib.auth.decorators import login_required

# @login_required
# def work_list(request):
    

# @login_required
# def accept_work(request, work_id):
#     work = get_object_or_404(WorkAllocation, id=work_id)
#     if request.method == 'POST':
#         work.status = 'Accepted'
#         work.save()
#         TechWorkList.objects.create(technician=request.user, work=work)
#         return redirect('work_list')
#     return render(request, 'accept_work.html', {'work': work})

# @login_required
# def reject_work(request, work_id):
#     work = get_object_or_404(WorkAllocation, id=work_id)
#     if request.method == 'POST':
#         work.status = 'Rejected'
#         work.save()
#         return redirect('pending_work')
#     return render(request, 'reject_work.html', {'work': work})

# @login_required
# def complete_work(request, work_id):
#     tech_work = get_object_or_404(TechWorkList, work_id=work_id, technician=request.user)
#     if request.method == 'POST':
#         if request.FILES.get('photo_before_service'):
#             tech_work.photo_before_service = request.FILES['photo_before_service']
#         if request.FILES.get('photo_after_service'):
#             tech_work.photo_after_service = request.FILES['photo_after_service']
#         if request.FILES.get('customer_signature_photo'):
#             tech_work.customer_signature_photo = request.FILES['customer_signature_photo']
#         if request.FILES.get('payment_photo'):
#             tech_work.payment_photo = request.FILES['payment_photo']

#         # Update payment status only if it's currently pending
#         if tech_work.work.customer_payment_status == 'Pending' and request.POST.get('customer_payment_status'):
#             tech_work.work.customer_payment_status = request.POST['customer_payment_status']

#         # Update work and tech work status
#         tech_work.status = 'Completed'
#         tech_work.work.status = 'Completed'
#         tech_work.work.save()
#         tech_work.save()
#         return redirect('completed_work_list')

#     return render(request, 'complete_work.html', {'tech_work': tech_work})

import base64
from django.core.files.base import ContentFile
from django.utils.timezone import now
from .models import TechWorkList, UploadedFile

@login_required
def complete_work(request, work_id):
    tech_work = get_object_or_404(TechWorkList, id=work_id, technician=request.user)
    print('techn_work',tech_work)
    for w in tech_work.work.all():
        print(w.payment_amount)

    if request.method == 'POST':
        # Get related work object (WorkAllocation)
        work_allocation = tech_work.work.first()  # Assuming only 1 work per tech_work

        # Handle Photos Before Service
        photos_before_service = request.FILES.getlist('photos_before_service')
        for photo in photos_before_service:
            uploaded_file = UploadedFile.objects.create(file=photo)
            tech_work.photos_before_service.add(uploaded_file)

        # Handle Photos After Service
        photos_after_service = request.FILES.getlist('photos_after_service')
        for photo in photos_after_service:
            uploaded_file = UploadedFile.objects.create(file=photo)
            tech_work.photos_after_service.add(uploaded_file)

        # Handle digital signature
        signature_data = request.POST.get('signature_data')
        if signature_data:
            format, imgstr = signature_data.split(';base64,')
            ext = format.split('/')[-1]
            signature_file = ContentFile(base64.b64decode(imgstr), name=f'xsignature.{ext}')
            tech_work.customer_signature_photo.save(f'signature_{work_allocation.id}.{ext}', signature_file)

        # Handle signature upload
        customer_signature_photo = request.FILES.get('customer_signature_photo')
        if customer_signature_photo:
            tech_work.customer_signature_photo = customer_signature_photo

        # Handle Payment Photos
        payment_photos = request.FILES.getlist('payment_photos')
        for photo in payment_photos:
            uploaded_file = UploadedFile.objects.create(file=photo)
            tech_work.payment_photos.add(uploaded_file)

        # Update customer payment status on WorkAllocation
        payment_status = request.POST.get('customer_payment_status')
        if not payment_status:
            payment_status = work_allocation.customer_payment_status or 'Pending'
            work_allocation.customer_payment_status = payment_status
            work_allocation.save()

        tech_work.payment_mode = payment_status
        print('payment_status',payment_status)
        tech_work.payment_type = request.POST.get('payment_type')
        print('payment type',request.POST.get('payment_type'))
        try:
            tech_work.remaining_amount = float(request.POST.get('remaining_balance') or 0.00)
        except ValueError:
            tech_work.remaining_amount = 0.00
        tech_work.next_due_date = request.POST.get('next_due_date') or None
        tech_work.save()
        # Update all related TechWorkList entries for this work allocation
        related_tech_works = TechWorkList.objects.filter(work=work_allocation)
        for tw in related_tech_works:
            tw.status = 'Completed'
            tw.completion_datetime = now()
            tw.save()

        # Finally, mark the WorkAllocation as Completed
        work_allocation.status = 'Completed'
        work_allocation.save()

        return redirect('completed_work_list')
    return render(request, 'complete_work.html', {'tech_work': tech_work})


@login_required
def completed_work_list(request):
    completed_works = TechWorkList.objects.filter(technician=request.user, status='Completed')
    return render(request, 'completed_work_list.html', {'completed_works': completed_works})

@login_required
def work_details(request, work_id):
    work = get_object_or_404(TechWorkList, id=work_id, technician=request.user)
    
    related_technicians = []
    for wa in work.work.all():  # Assuming this is a ManyToMany of WorkAllocation objects
        for tech in wa.technician.all():  # Assuming ManyToMany of TechnicianProfile
            if tech.user != request.user:
                related_technicians.append(tech)
    
    # Remove duplicates if any
    related_technicians = list({tech.id: tech for tech in related_technicians}.values())
    
    return render(request, 'work_details.html', {
        'work': work,
        'related_technicians': related_technicians,
    })



def view_work_details(request, work_id):
    tech_work = get_object_or_404(TechWorkList, pk=work_id)
    context = {
        'tech_work': tech_work
    }
    return render(request, 'work_details.html', context)

# views.py

from django.shortcuts import render
from django.views.generic import ListView, DetailView
from .models import TechWorkList

class AdminCompletedWorkView(ListView):
    model = TechWorkList
    template_name = 'admin_completed_work_list.html'
    paginate_by = 10  # ✅ Enable pagination

    def get_queryset(self):
        request = self.request
        user_profile = request.user.userprofile

        qs = (
            TechWorkList.objects
            .filter(status='Completed')
            .exclude(customer_signature_photo='')
            .prefetch_related(
                'work',
                'work__technician',
                'work__service',
                'work__service__customer'
            )
        )

        # ✅ Role-based access
        if user_profile.role == 'branch_manager':
            branch_manager = BranchManager.objects.get(mobile_no=request.user.username)
            qs = qs.filter(work__service__branch=branch_manager.branch)

        elif user_profile.role == 'operation_person':
            operation_person = OperationPerson.objects.get(user=request.user)
            qs = qs.filter(work__service__branch=operation_person.branch)

        elif user_profile.role != 'admin':
            return TechWorkList.objects.none()

        # ✅ Filters
        search = request.GET.get("search", "").strip()
        technician = request.GET.get("technician", "").strip()
        from_date = request.GET.get("from_date", "").strip()
        to_date = request.GET.get("to_date", "").strip()

        # Search filter
        if search:
            qs = qs.filter(
                Q(work__service__customer__fullname__icontains=search) |
                Q(work__service__customer__primarycontact__icontains=search)
            )

        # Technician dropdown
        if technician:
            qs = qs.filter(work__technician__id=technician)

        # Date range filters
        if from_date:
            qs = qs.filter(completion_datetime__date__gte=from_date)

        if to_date:
            qs = qs.filter(completion_datetime__date__lte=to_date)

        return qs.distinct()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        # ✅ Technician dropdown list
        ctx["technicians"] = TechnicianProfile.objects.all().order_by('first_name')

        # ✅ Preserve filters
        query = self.request.GET.copy()
        if "page" in query:
            query.pop("page")
        ctx["querystring"] = query.urlencode()

        # ✅ Return filter values to template
        ctx["search"] = self.request.GET.get("search", "")
        ctx["technician"] = self.request.GET.get("technician", "")
        ctx["from_date"] = self.request.GET.get("from_date", "")
        ctx["to_date"] = self.request.GET.get("to_date", "")

        return ctx

    
class AdminWorkDetailView(DetailView):
    model = TechWorkList
    template_name = 'admin_work_detail.html'
    context_object_name = 'work'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        work_instance = self.get_object()
        
        # Get all related technicians
        technicians_set = set()
        for wa in work_instance.work.all():
            for tech in wa.technician.all():
                technicians_set.add(tech)
        context['related_technicians'] = technicians_set

        # Get customer details from the first related WorkAllocation's service
        first_wa = work_instance.work.first()
        if first_wa and first_wa.service and first_wa.service.customer:
            customer = first_wa.service.customer
            context['customer'] = customer
        else:
            context['customer'] = None

        return context


def import_leads(request):
    if request.method == 'POST':
        form = LeadImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            file_type = file.name.split('.')[-1]
            if file_type == 'csv':
                handle_csv(file)
            elif file_type == 'xlsx':
                handle_xlsx(file)
            else:
                messages.error(request, 'Unsupported file format. Please upload a CSV or XLSX file.')
                return redirect('import_leads')
            
            return redirect('display_lead_management')  
    else:
        form = LeadImportForm()
    
    return render(request, 'import_leads.html', {'form': form})

import csv
from datetime import datetime

def handle_csv(file):
    decoded_file = file.read().decode('utf-8').splitlines()
    reader = csv.reader(decoded_file)
    next(reader)  

    for row in reader:
        try:
            # Handle empty or malformed rows
            if len(row):
                print(f"Skipped incomplete row: {row}")
                continue
            lead_management.objects.create(
                sourceoflead=row[1].strip(),
                salesperson=row[2].strip(),
                customername=row[3].strip(),
                customersegment=row[4].strip(),
                enquirydate=datetime.strptime(row[5], "%Y-%m-%d").date(),
                contactedby=row[6].strip(),
                maincategory=row[7].strip(),
                subcategory=row[8].strip(),
                primarycontact=row[9].strip(),
                secondarycontact=row[10] if row[10].strip().lower() != 'null' else None,
                customeremail=row[11].strip(),
                customeraddress=row[12].strip(),
                location=row[13].strip(),
                city=row[14].strip(),
                typeoflead=row[15].strip(),
                firstfollowupdate=datetime.strptime(row[16], "%Y-%m-%d").date(),
                stage=int(row[17].strip()),
                branch=row[18].strip(),
                state=row[19].strip(),
            )
        except Exception as e:
            print(f"❌ Error importing row {row}: {e}")


def handle_xlsx(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        lead_management.objects.create(
            sourceoflead=row[0],
            salesperson=row[1],
            primarycontact=row[2],
            customeraddress=row[3],
            customeremail=row[4],
            enquirydate=row[5],
            typeoflead=row[6],
            city=row[7],
            contactedby=row[8],
            customername =row[9],
            customersegment=row[10],
            location=row[11],
            maincategory=row[12],
            secondarycontact=row[13],
            subcategory=row[14],
            firstfollowupdate=row[15],
            stage=row[16],
        )
# try1









def calendar_view(request):
    return render(request, 'dashboard/dashboard.html')

def meeting_data(request):
    # Fetch all meetings
    meetings = Meeting.objects.all()
    events = []
    for meeting in meetings:
        # Calculate end time as 1 hour after start time
        start_datetime = datetime.datetime.combine(meeting.meeting_date, meeting.meeting_time)
        end_datetime = start_datetime + datetime.timedelta(hours=1)
        print("stratdate", start_datetime)
        print("enddatetiem", end_datetime)
        events.append({
            'title': f"{meeting.meeting_agenda}",
            'start': start_datetime.isoformat(),
            'end': end_datetime.isoformat(),
            'description': f"Agenda: {meeting.meeting_agenda}, Participants: {meeting.participants}"
        })
    return JsonResponse(events, safe=False)

# crmapp/views.py

import matplotlib.pyplot as plt
from io import BytesIO
import base64
from django.shortcuts import render
from crmapp.models import Product  # Your Product model

# def dashboard_view(request):
#     # Fetch the count of products per category
#     pest_control_count = Product.objects.filter(category='Pest Control').count()
#     fumigation_count = Product.objects.filter(category='Fumigation').count()
#     product_sell_count = Product.objects.filter(category='Product Sell').count()

#     # Prepare data for the bar chart
#     categories = ['Pest Control', 'Fumigation', 'Product Sell']
#     counts = [pest_control_count, fumigation_count, product_sell_count]

#     # Create the bar chart using matplotlib
#     fig, ax = plt.subplots()
#     ax.bar(categories, counts, color=['#FF6384', '#36A2EB', '#FFCE56'])
#     ax.set_title('Number of Products per Category')
#     ax.set_xlabel('Product Category')
#     ax.set_ylabel('Number of Products')

#     # Save the figure to a BytesIO object to convert it into an image for the web
#     buffer = BytesIO()
#     plt.savefig(buffer, format='png')
#     buffer.seek(0)
    
#     # Convert the image to base64
#     image_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

#     # Prepare context
#     context = {
#         'product_chart': image_base64,  # Send the chart as a base64 string to the template
#     }

#     return render(request, 'dashboard/dashboard.html', context)


@login_required
def go_towork(request, work_id):
    work = get_object_or_404(WorkAllocation, id=work_id)
    if request.method == 'POST':
        work.status = 'workdesk'
        work.save()
        TechWorkList.objects.create(technician=request.user, work=work)
        return redirect('work_list')
    return render(request, 'accept_work.html', {'work': work})

from django.shortcuts import render
from .models import WorkAllocation  # Import your model here

@login_required
def work_list_view(request):
    # Get search query
    query = request.GET.get('search', '')

    # Base queryset: only Pending work for this technician
    work_allocations = TechWorkList.objects.filter(
        technician=request.user,
        status="Pending"
    )

    # Apply search filter if query is provided
    if query:
        work_allocations = work_allocations.filter(
            Q(work__customer_contact__icontains=query)
        )

    # Debug (Optional)
    for work in work_allocations:
        print("work:", work, "| status:", work.status)

    return render(request, 'work_list.html', {'work_allocations': work_allocations})

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from crmapp.models import TechWorkList
from django.templatetags.static import static

def view_work_pdf(request, work_id):
    work = get_object_or_404(TechWorkList, pk=work_id, technician=request.user)
    logo_path = request.build_absolute_uri(static('images/Logo.png'))

    context = {'work': work,'logo_path': logo_path}
    html = render_to_string('work_pdf_template.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="work_{work_id}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating the PDF', status=400)
    
    return response

def download_work_pdf(request, work_id):
    work = get_object_or_404(TechWorkList, pk=work_id, technician=request.user)
    logo_path = request.build_absolute_uri(static('images/logo.png'))

    context = {'work': work,'logo_path': logo_path}
    html = render_to_string('work_pdf_template.html', context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_{work_id}.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating the PDF', status=400)
    
    return response


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse
import urllib.parse

def generate_pdf_link(request, work_id):
    work = get_object_or_404(TechWorkList, id=work_id)
    pdf_url = request.build_absolute_uri(reverse('download_work_pdf', args=[work_id]))
    whatsapp_message = f"Here is your service report: {pdf_url}"
    encoded_message = urllib.parse.quote(whatsapp_message)
    whatsapp_url = f"https://wa.me/{work.work.customer_contact}?text={encoded_message}"
    return redirect(whatsapp_url)




from django.shortcuts import render, redirect
from .models import Branch

from .models import Branch

def create_branch(request):
    if request.method == 'POST':
        branch_name = request.POST.get('branch_name')
        contact_1 = request.POST.get('contact_1')
        contact_2 = request.POST.get('contact_2') or None
        email_1 = request.POST.get('email_1')
        email_2 = request.POST.get('email_2') or None
        gst_number = request.POST.get('gst_number')
        pan_number = request.POST.get('pan_number')
        state = request.POST.get('state')
        code = request.POST.get('code')
        shortcut = request.POST.get('shortcut')
        full_address = request.POST.get('full_address')
        is_head_office = request.POST.get("is_head_office") == "on"
        print(is_head_office)
        # Only pass state; model will set shortcut and code
        branch = Branch(
            branch_name=branch_name,
            contact_1=contact_1,
            contact_2=contact_2,
            email_1=email_1,
            email_2=email_2,
            gst_number=gst_number,
            pan_number=pan_number,
            state=state,
            code = code,
            shortcut = shortcut,
            full_address=full_address,
            is_head_office = is_head_office

        )
        branch.save()  
        return redirect('branch_list')
   
    return render(request, 'create_branch.html' , {"state_map":state_map})


def branch_list(request):
    branches = Branch.objects.all().order_by('-created_at')
    return render(request, 'branch_list.html', {'branches': branches})

def delete_branch(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    branch.delete()
    return redirect('branch_list')

def edit_branch(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)

    if request.method == 'POST':
        branch.branch_name = request.POST.get('branch_name')
        branch.contact_1 = request.POST.get('contact_1')
        branch.contact_2 = request.POST.get('contact_2') or None
        branch.email_1 = request.POST.get('email_1')
        branch.email_2 = request.POST.get('email_2') or None
        branch.gst_number = request.POST.get('gst_number')
        branch.pan_number = request.POST.get('pan_number')
        branch.state = request.POST.get('state')
        branch.code = request.POST.get('code')
        branch.shortcut = request.POST.get('shortcut')
        branch.full_address = request.POST.get('full_address')
        branch.is_head_office = request.POST.get("is_head_office") == "on"

        branch.save()
        return redirect('branch_list')

    return render(request, 'edit_branch.html', {'branch': branch, "state_map":state_map})


def get_branch_details(request, branch_id):
    try:
        branch = Branch.objects.get(id=branch_id)
        data = {
            'contact_1': branch.contact_1,
            'contact_2': branch.contact_2 or '',
            'email_1': branch.email_1,
            'email_2': branch.email_2 or '',
            'gst_number': branch.gst_number,
            'pan_number': branch.pan_number,
            'full_address': branch.full_address,
            'state': branch.state,
            'code': branch.code,
            'shortcut': branch.shortcut,
        }
        return JsonResponse(data)
    except Branch.DoesNotExist:
        return JsonResponse({'error': 'Branch not found'}, status=404)    


def get_customer_details(request):
    if 'contact_no' in request.GET:
        contact_no = request.GET.get('contact_no')
        try:
            customer = customer_details.objects.get(primarycontact=contact_no)
            data = {
                'customer_id': customer.id,
                'customer_full_name': customer.fullname,
                'secondary_contact_no':customer.secondarycontact,
                'customer_email': customer.primaryemail,
                'secondary_email' : customer.secondaryemail,
                'contactperson':customer.contactperson,
                'shifttopartyaddress': customer.shifttopartyaddress,
                'city': customer.shifttopartycity,
                'state': customer.shifttopartystate,
                'pincode':customer.shifttopartypostal,
                'soldtopartyaddress': customer.soldtopartyaddress,
                'sold_city': customer.soldtopartycity,
                'sold_state': customer.soldtopartystate,
                'sold_pincode':customer.soldtopartypostal,
                'customer_type': customer.customer_type,
                'or_name':customer.or_name,
                'or_contact':customer.or_contact,  
            }
           
            return JsonResponse(data)
        
        except customer_details.DoesNotExist:
            return JsonResponse({'error': 'Customer not found'}, status=404)
    return JsonResponse({'error': 'No contact number provided'}, status=400)


from .models import BankAccounts   
from django.http import JsonResponse

from .models import BankAccounts  # Make sure the model is correctly imported
from django.views.decorators.http import require_POST

@login_required
def create_bank_account(request):
    if request.method == 'POST':
        bank_name = request.POST.get('bank_name')
        account_number = request.POST.get('bank_account_number')
        ifs_code = request.POST.get('ifs_code')
        branch = request.POST.get('branch')


        if not all([bank_name, account_number, ifs_code, branch]):
            return JsonResponse({'error': 'All fields are required'}, status=400)
        
     
        if BankAccounts.objects.filter(account_number=account_number).exists():
            return JsonResponse({'error': 'Bank account already exists'}, status=409)

        BankAccounts.objects.create(
            bank_name=bank_name,
            account_number=account_number,
            ifs_code=ifs_code,
            branch=branch
        )
        return redirect('list_bank_accounts') 

    return render(request, 'create_bank_account.html')

@login_required
def list_bank_accounts(request):
    accounts = BankAccounts.objects.all()
    return render(request, 'list_bank_accounts.html', {'bank_accounts': accounts})

@login_required
def edit_bank_account(request, account_id):
    bank_account = get_object_or_404(BankAccounts, id=account_id)

    if request.method == 'POST':
        bank_name = request.POST.get('bank_name')
        account_number = request.POST.get('bank_account_number')
        ifs_code = request.POST.get('ifs_code')
        branch = request.POST.get('branch')

        if not all([bank_name, account_number, ifs_code, branch]):
            return render(request, 'edit_bank_account.html', {
                'error': 'All fields are required',
                'bank': bank_account
            })

        # Optional: Check if another account exists with the same number
        if BankAccounts.objects.filter(account_number=account_number).exclude(id=bank_account.id).exists():
            return render(request, 'edit_bank_account.html', {
                'error': 'Another account with this number already exists',
                'bank': bank_account
            })

        # Update fields
        bank_account.bank_name = bank_name
        bank_account.account_number = account_number
        bank_account.ifs_code = ifs_code
        bank_account.branch = branch
        bank_account.save()

        return redirect('list_bank_accounts')

    return render(request, 'edit_bank_account.html', {'bank': bank_account})


@login_required
@require_POST
def delete_bank_account(request, account_id):
    bank_account = get_object_or_404(BankAccounts, id=account_id)
    bank_account.delete()
    return redirect('list_bank_accounts')



from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from .models import TaxInvoice, TaxInvoiceItem, quotation_management, customer_details, BankAccounts
import json
from datetime import datetime


def create_tax_invoice(request):
    use_quotation = request.GET.get("use_quotation") == "true"
    if  request.method == "POST":
        try:
            if use_quotation:
                # 1. Get quotation and related data
                quotation_no = request.POST.get("quotation_no")
                quotation = get_object_or_404(quotation_management, quotation_no=quotation_no)
                customer = get_object_or_404(customer_details, primarycontact=quotation.customer.primarycontact)
                branch = get_object_or_404(Branch, id = quotation.branch_id)
                gst_enabled = quotation.apply_gst
                if quotation.igst > 0:
                    gst_type = "IGST"
                else:
                    gst_type = "CGST + SGST"

                # 2. Get product data
                product_data = request.POST.get("product_data", "[]")
                items = json.loads(product_data)
                shifttopartystate=request.POST.get('shifttopartystate')
                shifttopartystatecode=request.POST.get('shifttopartystatecode')
                print("shift", shifttopartystate, shifttopartystatecode)
                soldtopartystate=request.POST.get('soldtopartystate')
                soldtopartystatecode=request.POST.get('soldtopartystatecode')
                print("sold", soldtopartystate, soldtopartystatecode)
            else:
                quotation = None
                branch = get_object_or_404(Branch, id = request.POST.get('branch_id'))
                customer = get_object_or_404(customer_details, primarycontact=request.POST.get('contact_no'))
                # Get the raw value from POST (e.g., "Maharashtra-27")
                shifttopartystate_raw = request.POST.get('shifttopartystate', '')

                # Extract state name and code
                if '-' in shifttopartystate_raw:
                    shifttopartystate, shifttopartystatecode = shifttopartystate_raw.split('-', 1)
                else:
                    shifttopartystate = shifttopartystate_raw
                    shifttopartystatecode = ''

                # Same logic for sold_to
                soldtopartystate_raw = request.POST.get('soldtopartystate', '')
                if '-' in soldtopartystate_raw:
                    soldtopartystate, soldtopartystatecode = soldtopartystate_raw.split('-', 1)
                else:
                    soldtopartystate = soldtopartystate_raw
                    soldtopartystatecode = ''
                gst_enabled = request.POST.get("gst_enabled")
                gst_type = request.POST.get("gst_type") 

                selected_products_json = request.POST.get('selected_products_json','[]')
                # print(selected_products_json)
                print("Selected Products JSON:", selected_products_json)
                items = json.loads(selected_products_json)
            bank_id = request.POST.get("bank_id")
            bank = get_object_or_404(BankAccounts, id=bank_id)

            # Update customer details
            customer.primarycontact = request.POST.get('contact_no')
            customer.fullname = request.POST.get('customer_full_name')
            customer.primaryemail = request.POST.get('customer_email')
            customer.customer_type = request.POST.get('customer_type')
            customer.or_name = request.POST.get('or_name', '')
            customer.or_contact = request.POST.get('or_contact') or None
            customer.save()

            # Create TaxInvoice with grand_total
            invoice = TaxInvoice.objects.create(
                quotation=quotation,
                customer=customer,
                bank=bank,
                branch = branch,
                referance_no_and_date=request.POST.get("referance_no_and_date"),
                other_referance=request.POST.get("other_references"),
                delivery_note=request.POST.get("delivery_note"),
                modern_terms_of_payment=request.POST.get("mode_terms_of_payment"),
                buyers_order_no=request.POST.get("buyer_order_no"),
                dated=parse_date_or_none(request.POST.get("dated")),
                dispatch_doc_no=request.POST.get("dispatch_doc_no"),
                delivery_note_date=parse_date_or_none(request.POST.get("delivery_note_date")),
                dispatched_through=request.POST.get("dispatched_through"),
                destination=request.POST.get("destination"),
                service_titel=request.POST.get('service_titel'),
                shift_gstin_uin=request.POST.get('shift_gstin_uin'),
                shift_pan_number = request.POST.get('shift_pan'),
                shifttopartystate=shifttopartystate,
                shifttopartystatecode=shifttopartystatecode,
                sold_gstin_uin=request.POST.get('sold_gstin_uin'),
                sold_pan_number = request.POST.get('sold_pan'), 
                soldtopartystate=soldtopartystate,
                soldtopartystatecode=soldtopartystatecode,
                remarks = request.POST.get('remarks'),
                terms_of_delivery = request.POST.get('terms_of_delivery'),
                ship_to_address = request.POST.get('ship_to_address'),
                bill_to_address = request.POST.get('bill_to_address'),
            
                gst_type=gst_type if gst_enabled else "No GST"
            )

            # 5. Save products       
            grand_total = 0
            for item in items:
                price = float(item.get('price', 0))
                quantity = float(item.get('quantity', 0))
                gst_percent = float(item.get('gst', 0))
       
                total = price * quantity 
                if gst_enabled:  
                    gst_amount = round((total * gst_percent) / 100, 2)
                else:
                    gst_amount = 0

                grand_total += total + gst_amount

                TaxInvoiceItem.objects.create(
                    tax_invoice=invoice,
                    product_name=item.get('name'),
                    hsn_code=item.get('hsn'),
                    quantity=quantity,
                    description = item.get('description'),
                    unit=item.get('unit'),
                    price=price,
                    gst_percent=gst_percent,
                    gst_amount=gst_amount,
                    total=total
                )
            
            invoice.grand_total = grand_total
            invoice.save()



            return redirect("display_tax_invoice")

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    banks = BankAccounts.objects.all()
    branches = Branch.objects.all()
    category_choices = Product.CATEGORY_CHOICES
    product = Product.objects.all()
    selected_state = request.POST.get('soldtopartystate', '')
    context = {'banks': banks , 'use_quotation':use_quotation,
               'branches':branches , 'state_map':state_map,
               'selected_state':selected_state,
               'category_choices':category_choices,
               'product':product} 
    return render(request, "create_tax_invoice.html", context)

def parse_date_or_none(date_str):
    """Utility to parse date from string or return None."""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
    except:
        return None



def display_tax_invoice(request):
    query = request.GET.get('search', '')
    sort_by = request.GET.get('sort_by', '')
    sort_order = request.GET.get('order', 'asc')
    
    if request.user.userprofile.role == 'admin':
        m = TaxInvoice.objects.all()
    elif request.user.userprofile.role == 'branch_manager':
        branch = BranchManager.objects.get(mobile_no=request.user.username).branch
        m = TaxInvoice.objects.filter(branch=branch)
    elif request.user.userprofile.role == 'sales':
        sales = SalesPerson.objects.get(mobile_no=request.user.username)
        if sales.co_ordinator:
            m = TaxInvoice.objects.all()
        else:
            m = TaxInvoice.objects.filter(created_by_no=sales.mobile_no)
    else:
        m = TaxInvoice.objects.none()

    if query:
        m = m.filter(
            Q(customer__customerid__icontains=query) |
            Q(tax_invoice_no__icontains=query)
        )
   
    
    if sort_by == 'name':
        sort_field = 'customer__fullname'
    elif sort_by == 'invoice_no':
        sort_field = 'tax_invoice_no'
    else:
        sort_field = 'id'  # default sorting if no valid sort_by is provided

    # Apply ordering based on direction
    if sort_order == 'desc':
        m = m.order_by(f'-{sort_field}')
    else:
        m = m.order_by(sort_field)

    paginator = Paginator(m, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    start_index = (page_obj.number - 1) * paginator.per_page

    context = {
        'current_order': sort_order,
        'current_sort_by': sort_by,
        'query': query,
        'page_obj': page_obj,
        'start_index': start_index,
    }
    return render(request, 'display_invoice.html', context)


def edit_tax_invoice(request, id):
    invoice = get_object_or_404(TaxInvoice, id=id)
    items = invoice.items.all()
    branches = Branch.objects.all()
    banks = BankAccounts.objects.all()
    category_choices = Product.CATEGORY_CHOICES
    product = Product.objects.all()
    customer = invoice.customer
    if request.method == "POST":

        customer.fullname = request.POST.get('customer_full_name')
        customer.primarycontact = request.POST.get('contact_no')
        customer.primaryemail = request.POST.get('customer_email')
        customer.customer_type = request.POST.get('customer_type')
        customer.or_contact = request.POST.get('or_contact')
        customer.or_name = request.POST.get('or_name')
        customer.save()

        invoice.branch_id = request.POST.get('branch_id')
        invoice.customer = customer
        invoice.bill_to_address = request.POST.get('bill_to_address')
        invoice.ship_to_address = request.POST.get('ship_to_address')
        invoice.shift_gstin_uin = request.POST.get('shift_gstin_uin')
        invoice.shift_pan_number = request.POST.get('shift_pan')
        invoice.sold_gstin_uin = request.POST.get('sold_gstin_uin')
        invoice.sold_pan_number = request.POST.get('sold_pan')
        invoice.buyers_order_no = request.POST.get('buyer_order_no')
        invoice.dispatch_doc_no = request.POST.get('dispatch_doc_no')
        invoice.dated = parse_date(request.POST.get('dated'))
        invoice.referance_no_and_date = request.POST.get('referance_no_and_date')
        invoice.dispatched_through = request.POST.get('dispatched_through')
        invoice.destination = request.POST.get('destination')
        invoice.other_referance =  request.POST.get('other_references')
        invoice.modern_terms_of_payment = request.POST.get('mode_terms_of_payment')
        invoice.delivery_note = request.POST.get('delivery_note')
        invoice.delivery_note_date = parse_date(request.POST.get('delivery_note_date'))
        invoice.remarks = request.POST.get('remarks')
        invoice.terms_of_delivery = request.POST.get('terms_of_delivery')
        invoice.bank_id = request.POST.get('bank_id')
        shifttopartystate_raw = request.POST.get('shifttopartystate', '')
        # Extract state name and code
        if '-' in shifttopartystate_raw:
            invoice.shifttopartystate, invoice.shifttopartystatecode = shifttopartystate_raw.split('-', 1)
    
        # Same logic for sold_to
        soldtopartystate_raw = request.POST.get('soldtopartystate', '')
        if '-' in soldtopartystate_raw:
            invoice.soldtopartystate, invoice.soldtopartystatecode = soldtopartystate_raw.split('-', 1)
        invoice.save()
        gst_enabled = request.POST.get("gst_enabled")
        gst_type = request.POST.get("gst_type") 
        deleted_ids = request.POST.get("deleted_items", "")
        if deleted_ids:
            ids = [int(x) for x in deleted_ids.split(",") if x.isdigit()]
            TaxInvoiceItem.objects.filter(id__in=ids, tax_invoice=invoice).delete()

        total_items = int(request.POST.get("total_items", 0))
        products = []
        for i in range(1, total_items + 1):
            product_data = {
                "hsn_code": request.POST.get(f"old_hsn_code_{i}"),
                "price": request.POST.get(f"old_price_{i}"),
                "quantity": request.POST.get(f"old_quantity_{i}"),
                "description": request.POST.get(f"old_description_{i}"),
                "unit": request.POST.get(f"old_unit_{i}"),
                "gst_percent": request.POST.get(f"old_gst_percent_{i}"),
            }
            products.append(product_data)

        # 1. Update existing items
        for product in products:  # products from old_* form fields
            row_id = product.get("row_id")
            if not row_id:
                continue
            
            try:
                item = TaxInvoiceItem.objects.get(id=row_id, tax_invoice=invoice)
            except TaxInvoiceItem.DoesNotExist:
                continue
            
            price = float(product.get("price", 0))
            quantity = float(product.get("quantity", 0))
            gst_percent = float(product.get("gst_percent", 0))
            total = price * quantity

            gst_amount = round((total * gst_percent) / 100, 2) if gst_enabled else 0

            item.hsn_code = product.get("hsn_code")
            item.price = price
            item.quantity = quantity
            item.description = product.get("description")
            item.unit = product.get("unit")
            item.gst_percent = gst_percent
            item.gst_amount = gst_amount
            item.total = total
            item.save()


        # 2. Add new items (from JSON)
        selected_products_json = request.POST.get('selected_products_json','[]')
        # print(selected_products_json)
        print("Selected Products JSON:", selected_products_json)
        try:
            new_items = json.loads(selected_products_json) if selected_products_json else []
        except json.JSONDecodeError:
            new_items = []
        grand_total = 0
        for item in new_items:  # items = new products from JSON
            price = Decimal(item.get('price', 0) or 0)
            quantity = Decimal(item.get('quantity', 0) or 0)
            gst_percent = Decimal(item.get('gst', 0) or 0)

            total = price * quantity
            gst_amount = round((total * gst_percent) / 100, 2) if gst_enabled else 0

            grand_total += total + gst_amount

            TaxInvoiceItem.objects.create(
                tax_invoice=invoice,
                product_name=item.get('name'),
                hsn_code=item.get('hsn'),
                quantity=quantity,
                description=item.get('description'),
                unit=item.get('unit'),
                price=price,
                gst_percent=gst_percent,
                gst_amount=gst_amount,
                total=total
            )

        # 3. Update invoice total
        invoice.grand_total = grand_total + sum(x.total + x.gst_amount for x in invoice.items.all())
        
        invoice.save(update_fields=['grand_total'])

        return redirect('display_tax_invoice')

    data = {
        "invoice":invoice,
        "items":items,
        "branches":branches,
        'state_map':state_map,
        'banks': banks,
        'category_choices':category_choices,
        'product':product,
    }
    return render(request, 'edit_tax_invoice.html', context= data)

def tax_invoice_pdf(request, id):
    invoice = get_object_or_404(TaxInvoice, id=id)
    items = invoice.items.all()
    cgst = sgst = igst = 0 
    branch = invoice.branch
    if invoice.quotation:
        cgst = invoice.quotation.cgst or 0
        sgst = invoice.quotation.sgst or 0
        igst = invoice.quotation.igst or 0
    else:
        gst = items.aggregate(total_gst=Sum('gst_amount'))['total_gst'] or 0
        if invoice.gst_type == 'CGST + SGST':
            cgst = gst / 2
            sgst = gst / 2
        else:
            igst  = gst

    amount_in_words = price_in_words(invoice.grand_total)
    # print("data" ,invoice.quotation.igst )
    context = {
        'invoice': invoice,
        'items': items,
        'cgst_total': cgst,
        'sgst_total': sgst,
        'igst_total':igst,
        'amount_in_words':amount_in_words,
        'today': datetime.now(),
    }

    
    # Render the template
    template_path = 'tax_invoice_pdf.html'
    template = get_template(template_path)
    html_string = template.render(context)

    # Generate PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)
    if pdf.err:
        return HttpResponse('Error generating PDF', status=500)
    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    # Check if 'download=true' is passed in query params
    if request.GET.get('download') == 'true':
        response['Content-Disposition'] = f'attachment; filename="TaxInvoice_{invoice.id}.pdf"'
    else:
        response['Content-Disposition'] = f'inline; filename="TaxInvoice_{invoice.id}.pdf"'

    return response


def delete_invoice(request, invoice_id):
    invoice = get_object_or_404(TaxInvoice, id = invoice_id)
    invoice.delete()
    return redirect('/display_tax_invoice')



import csv
from django.http import HttpResponse
from .models import TaxInvoice, TaxInvoiceItem , PaymentsRecord
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError

def export_invoice_excel(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="invoice_list.csv"'

    writer = csv.writer(response)

    # Header row
    writer.writerow([
        'Invoice No', 'Quotation No', 'Customer Name', 'Bank Name', 'Created At', 'Grand Total',
        'Product Name', 'HSN Code', 'Quantity', 'Unit', 'Price', 'GST %', 'GST Amount', 'Total'
    ])

    # Fetch invoices and write rows with their items
    invoices = TaxInvoice.objects.select_related('quotation', 'customer', 'bank').prefetch_related('items')

    for invoice in invoices:
        for item in invoice.items.all():
            writer.writerow([
                invoice.tax_invoice_no,
                invoice.quotation.quotation_no if invoice.quotation else '',
                invoice.customer.fullname if invoice.customer else '',
                invoice.bank.bank_name if invoice.bank else '',
                invoice.created_at.strftime('%Y-%m-%d %H:%M'),
                invoice.grand_total,
                item.product_name,
                item.hsn_code,
                item.quantity,
                item.unit,
                item.price,
                item.gst_percent,
                item.gst_amount,
                item.total
            ])

    return response


#  Payments Record Section


def create_payment_record(request):
    payment_choices = PaymentsRecord.PAYMENT_MODE_CHOICES
    payment_ratings = [(i, f"{i} Star") for i in range(1, 6)]
    

    if request.method == "GET":
        return render(request, "payment_records_create.html", {
            'payment_choices': payment_choices,
            'payment_ratings': payment_ratings
        })

    if request.method == "POST":
        invoice_no = request.POST.get("main_invoice")
        try:
            invoice = TaxInvoice.objects.get(tax_invoice_no=invoice_no)

            record = PaymentsRecord(
                main_invoice=invoice,
                amount_paid=request.POST.get("amount_paid"),
                payment_date=request.POST.get("payment_date"),
                next_due_date=request.POST.get("next_due_date") or None,
                previous_due_date=request.POST.get("previous_due_date") or None,
                work_type=request.POST.get("work_type"),
                payment_details=request.POST.get("Payment_details"),
                payment_mode=request.POST.get("payment_mode"),
                payment_rating=request.POST.get("payment_rating") or None,
                remarks=request.POST.get("remarks"),
                attachment=request.FILES.get("payment_attachment")
            )

            
            try:
                record.full_clean()  # This triggers the clean() method
                record.save()
                messages.success(request, "Payment record created successfully.")
                return redirect("payment_record_lists")
            except ValidationError as ve:
                for msg in ve.messages:
                    messages.error(request, msg)

        except TaxInvoice.DoesNotExist:
            messages.error(request, "Invoice not found.")

        except ValidationError as ve:
            messages.error(request, f"Validation error: {ve}")

        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")


    return render(request, "payment_records_create.html", {
        'payment_choices': payment_choices,
        'payment_ratings': payment_ratings
    })


@csrf_exempt 
def fetch_invoice_details(request):
    if request.method == "POST":
        invoice_no = request.POST.get("invoice_no") 
        try:
            invoice = TaxInvoice.objects.get(tax_invoice_no=invoice_no)
            customer = invoice.customer

            last_payment = PaymentsRecord.objects.filter(main_invoice=invoice).order_by('-id').first()
            amount_remaining = last_payment.amount_remaining if last_payment else invoice.grand_total 

            data = {
                "success": True,
                "fullname": customer.fullname,
                "mobile": customer.primarycontact,
                "email": customer.primaryemail,
                "total_amount": invoice.grand_total,
                "amount_remaining":float(amount_remaining),
            }
        except TaxInvoice.DoesNotExist:
            data = {"success": False, "error": "Invoice not found"}

        return JsonResponse(data)
    

def fetch_tax_invoice_details(request, id):
    invoice = TaxInvoice.objects.get(id=id)
    return render(request, 'tax_invoice_details.html', {'invoice': invoice})

    
from django.db.models import Max, Subquery, OuterRef
from collections import OrderedDict

def payment_records_list(request):
    search_query = request.GET.get('search', '')
    ageing_filter = request.GET.get('ageing', '')
    due_order = request.GET.get('due_order', '')
    remain_amount = request.GET.get('remain_amount','')
    print("ageing_filter", ageing_filter)

    latest_ids = PaymentsRecord.objects.values('main_invoice_id') \
        .annotate(latest_id=Max('id')) \
        .values_list('latest_id', flat=True)
    
    payments = PaymentsRecord.objects.filter(id__in=latest_ids).select_related('main_invoice')

    if search_query:
        payments = payments.filter(payment_invoice_no__icontains=search_query)

    if ageing_filter:
        payments = [
            p for p in payments 
            if p.ageing and p.ageing.replace('–', '-').replace(' Days', '') == ageing_filter
        ]

    if due_order == "asc":
        payments = sorted(payments, key=lambda x: x.next_due_date or x.payment_date)
    elif due_order == "desc":
        payments = sorted(payments, key=lambda x: x.next_due_date or x.payment_date, reverse=True)

    if remain_amount == "asc":
        payments = sorted(payments, key=lambda x: x.amount_remaining)
    elif remain_amount == "desc":
        payments = sorted(payments, key=lambda x: x.amount_remaining, reverse=True)

    context = {
        "payments": payments,
        "search_query": search_query,
        "ageing_selected": ageing_filter,
        "due_order": due_order,
        "remain_amount":remain_amount,
    }
    return render(request, "payment_records_list.html", context)



def payment_records_details(request , pk):
    search_query = request.GET.get('search', '')
    ageing_filter = request.GET.get('ageing', '')
    due_order = request.GET.get('due_order', '')
    remain_amount = request.GET.get('remain_amount','')

    main_invoice = get_object_or_404(TaxInvoice, id=pk)
    payments = PaymentsRecord.objects.filter(main_invoice_id = main_invoice)
 
    if search_query:
        payments = payments.filter(payment_invoice_no__icontains=search_query)

    if ageing_filter:
        payments = [
            p for p in payments 
            if p.ageing and p.ageing.replace('–', '-').replace(' Days', '') == ageing_filter
        ]

    if due_order == "asc":
        payments = sorted(payments, key=lambda x: x.next_due_date or x.payment_date)
    elif due_order == "desc":
        payments = sorted(payments, key=lambda x: x.next_due_date or x.payment_date, reverse=True)

    if remain_amount == "asc":
        payments = sorted(payments, key=lambda x: x.amount_remaining)
    elif remain_amount == "desc":
        payments = sorted(payments, key=lambda x: x.amount_remaining, reverse=True)

    context = {
        "payments": payments,
        "main_invoice": main_invoice,
        "search_query": search_query,
        "ageing_selected": ageing_filter,
        "due_order": due_order,
        "remain_amount":remain_amount,
    }
    return render(request, "payment_records_details.html", context)


from reportlab.lib.pagesizes import A4
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from django.templatetags.static import static
from django.http import HttpResponse
import io
from .models import quotation_management, MessageTemplates
from .custom_filters import price_in_words  
from reportlab.lib.colors import HexColor   
from reportlab.lib.enums import TA_RIGHT, TA_CENTER


def indian_currency(number):
    """
    Format number in Indian style: 1,23,456.00
    Safe against invalid input (returns 0.00)
    """
    try:
        number = float(number)
    except (ValueError, TypeError):
        number = 0.0
        
    s = f"{number:.2f}"
    integer, decimal = s.split(".")
    
    if len(integer) > 3:
        last3 = integer[-3:]
        rest = integer[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        integer = ",".join(parts) + "," + last3
            
    return f"{integer}.{decimal}"


def draw_footer_and_logo(canvas, doc, logo_path, footer_path, branch):
    # --- HEADER ---
    try:
        logo = ImageReader(logo_path)
        canvas.drawImage(
            logo, 20 * mm, A4[1] - 40 * mm,
            width=30 * mm, height=30 * mm,
            mask='auto'
        )
    except Exception as e:
        print("Logo load failed:", e)

    branch_font = ("Helvetica-Bold", 10)
    sfs_font = ("Helvetica-Bold", 15)

    branch_text = 'Seva Facility Services Pvt Ltd'
    sfs_text = "SFS PEST CONTROL"

    branch_width = canvas.stringWidth(branch_text, *branch_font)
    sfs_width = canvas.stringWidth(sfs_text, *sfs_font)

    right_margin = A4[0] - 20 * mm

    # --- Branch name ---
    canvas.setFont(*branch_font)
    canvas.drawRightString(right_margin, A4[1] - 20 * mm, branch_text)

    # --- SFS centered above ---
    branch_center_x = right_margin - (branch_width / 2)
    sfs_x = branch_center_x - (sfs_width / 2)
    canvas.setFont(*sfs_font)
    canvas.drawString(sfs_x, A4[1] - 15 * mm, sfs_text)

    # --- Address (dynamic height) ---
    style = ParagraphStyle(
        "right_address",
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        alignment=TA_RIGHT
    )

    current_y = A4[1] - 22 * mm  # start below branch name

    if branch.full_address:
        addr_para = Paragraph(branch.full_address, style)
        addr_w, addr_h = addr_para.wrap(branch_width, 100*mm)
        addr_para.drawOn(canvas, right_margin - addr_w, current_y - addr_h)
        current_y -= addr_h + 10  # move down based on actual height

    # --- Email(s) ---
    canvas.setFont("Helvetica", 8.5)
    if branch.email_1 or branch.email_2:
        canvas.drawRightString(right_margin, current_y, f"{branch.email_1}, {branch.email_2}")
        current_y -= 12

    # --- Contact ---
    if branch.contact_1:
        canvas.drawRightString(right_margin, current_y, branch.contact_1)
        current_y -= 12

    # --- GST + PAN ---
    gst_pan_text = f"GSTIN: {branch.gst_number} | PAN No: {branch.pan_number}"
    canvas.drawRightString(right_margin, current_y, gst_pan_text)
    current_y -= 12

    # --- Horizontal Line ---
    canvas.setLineWidth(0.8)
    canvas.setStrokeColorRGB(0, 0, 0)
    canvas.line(20 * mm, current_y, A4[0] - 20 * mm, current_y)

    # --- Footer Image ---
    try:
        footer = ImageReader(footer_path)
        iw, ih = footer.getSize()
        aspect = ih / float(iw)
        new_width = A4[0]
        new_height = new_width * aspect
        bottom_margin = 3 * mm
        canvas.drawImage(
            footer,
            0, bottom_margin,
            width=new_width, height=new_height,
            preserveAspectRatio=True,
            mask='auto'
        )
    except Exception as e:
        print("Footer load failed:", e)


        
import io
from reportlab.lib.pagesizes import A4
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Assuming these are already defined in your file:
# - draw_footer_and_logo(canvas, doc_obj, logo_path, footer_path, branch)
# - indian_currency(value)
# - price_in_words(value)
# - quotation_management model
# - static('images/Logo.png') and static('images/NewFooter.png')

def safe_float(value):
    try:
        return float(str(value).strip())
    except (ValueError, TypeError, AttributeError):
        return 0.0

def reportlab_quotation_pdf(request, id):
    quotation = quotation_management.objects.get(id=id)
    branch = quotation.branch
    logo_path = request.build_absolute_uri(static('images/Logo.png'))
    footer_path = request.build_absolute_uri(static('images/NewFooter.png'))
    light_blue = HexColor("#0070C0")

    buffer = io.BytesIO()
    doc = BaseDocTemplate(buffer, pagesize=A4,
                          leftMargin=20 * mm, rightMargin=20 * mm,
                          topMargin=50 * mm, bottomMargin=35 * mm)
    doc.quotation_no = quotation.quotation_no

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')

    def _header_footer(canvas, doc_obj):
        draw_footer_and_logo(canvas, doc_obj, logo_path, footer_path, branch)

    doc.addPageTemplates([PageTemplate(id='quotation_template', frames=frame, onPage=_header_footer)])

    styles = getSampleStyleSheet()
    normal = styles['Normal']
    bold = ParagraphStyle(name='bold', parent=normal, fontName='Helvetica-Bold')
    small = ParagraphStyle(name='small', parent=normal, fontSize=9)

    full_width = ParagraphStyle(name='full_width', parent=normal, leftIndent=0, firstLineIndent=0,
                                rightIndent=0, spaceBefore=0, spaceAfter=0, fontSize=10, leading=9)

    elements = []

    # ── Customer + Quotation Details ────────────────────────────────────────
    left_style = ParagraphStyle(name='left', fontSize=10, leading=10)
    right_style = ParagraphStyle(name='right', fontSize=10, alignment=TA_RIGHT, leading=12)
    address_style = ParagraphStyle('address', parent=left_style, leading=14)

    customer_details = [
        Paragraph(f"<b>Name :</b> {quotation.customer.fullname}", left_style),
        Paragraph(f"<b>Phone :</b> {quotation.customer.primarycontact}", left_style),
        Paragraph(f"<b>Email :</b> {quotation.customer.primaryemail}", left_style),
        Paragraph(f"<b>Address :</b> {quotation.address}", address_style),
    ]
    if quotation.or_name:
        customer_details.append(
            Paragraph(f"<b>Person :</b> {quotation.or_name} - {quotation.or_contact}", left_style)
        )

    left_table = Table([[item] for item in customer_details], colWidths=[95 * mm])
    left_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('LINEBEFORE', (0, 0), (0, -1), 3, light_blue),
    ]))

    quotation_details = [
        Paragraph("<b>Quotation</b>", ParagraphStyle(name='title', fontSize=12, alignment=TA_RIGHT, leading=12)),
        Paragraph(f"<b>{quotation.quotation_no}</b>", ParagraphStyle(name="temp_right_bold", parent=right_style, fontSize=14, fontName='Helvetica-Bold')),
        Paragraph(f"<b>Date:</b> {quotation.quotation_date.strftime('%d %B %Y')}", right_style)
    ]

    right_table = Table([[item] for item in quotation_details], colWidths=[85 * mm])

    combined_table = Table([[left_table, right_table]], colWidths=[95 * mm, 85 * mm])
    combined_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))

    elements.append(combined_table)
    elements.append(Spacer(1, 5))

    # ── Subject and Intro ───────────────────────────────────────────────────
    thank_u_note_style = ParagraphStyle(name="thank_u_note_style", parent=full_width, fontSize=10, leading=14)

    elements.append(Paragraph(f"<b>Subject:</b> {quotation.subject}", full_width))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"<b>{quotation.thank_u_note}:</b>", thank_u_note_style))
    elements.append(Spacer(1, 10))

    # ── Product Table ───────────────────────────────────────────────────────
    product_data = [["Sr. No.", "Product / Service", "HSN", "Rate (Rs)", "Qty", "Total (Rs)"]]

    calculated_total = 0.0

    for idx, item in enumerate(quotation.product_details_json, start=1):
        hsn = item.get('hsn_code', '')

        price = safe_float(item.get('price'))
        quantity = safe_float(item.get('quantity'))
        total = price * quantity

        calculated_total += total

        description = item.get('description', '').replace('\n', '<br/>')

        product_data.append([
            str(idx),
            Paragraph(
                f"<b>{item.get('name', '')}</b><br/><font size='8'><i>{description}</i></font>",
                small
            ),
            Paragraph(hsn, ParagraphStyle(name="hsn_style", parent=small, alignment=TA_CENTER)),
            Paragraph(indian_currency(price), ParagraphStyle(name="right", parent=small, alignment=TA_RIGHT, wordWrap='CJK')),
            Paragraph(f"{quantity:.2f}<br/>{item.get('unit', '')}", ParagraphStyle(name="right", parent=small, alignment=TA_RIGHT)),
            indian_currency(total)
        ])

    col_widths = [13 * mm, 66 * mm, 20 * mm, 25 * mm, 21 * mm, 25 * mm]
    total_width = sum(col_widths)

    product_table = Table(product_data, colWidths=col_widths)
    product_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D9D9D9')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (2, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))

    elements.append(product_table)

    hr = Table([['']], colWidths=[total_width], rowHeights=16)
    hr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#D9D9D9')),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
    ]))
    elements.append(hr)

    # ── GST & Totals Calculation ────────────────────────────────────────────
    gst_amount = cgst = sgst = igst = 0.0

    if quotation.apply_gst:
        gst_rate = 18  # ← you can make this dynamic later if needed

        if quotation.igst:
            igst = (calculated_total * gst_rate) / 100
            gst_amount = igst
        else:
            cgst = (calculated_total * gst_rate) / 200
            sgst = (calculated_total * gst_rate) / 200
            gst_amount = cgst + sgst

    grand_total = calculated_total + gst_amount

    # ── Totals Section ──────────────────────────────────────────────────────
    totals_data = []

    # Total
    totals_data.append([
        Paragraph("<b>Total :</b>", right_style),
        Paragraph(f"<b>{indian_currency(calculated_total)}</b>", right_style)
    ])

    # GST breakdown
    if quotation.apply_gst:
        if quotation.igst:
            totals_data.append([
                Paragraph("<b>IGST :</b>", right_style),
                Paragraph(f"<b>{indian_currency(igst)}</b>", right_style)
            ])
        else:
            totals_data.append([
                Paragraph("<b>CGST :</b>", right_style),
                Paragraph(f"<b>{indian_currency(cgst)}</b>", right_style)
            ])
            totals_data.append([
                Paragraph("<b>SGST :</b>", right_style),
                Paragraph(f"<b>{indian_currency(sgst)}</b>", right_style)
            ])
            totals_data.append([
                Paragraph("<b>Total Tax :</b>", right_style),
                Paragraph(f"<b>{indian_currency(gst_amount)}</b>", right_style)
            ])

    # Grand Total
    totals_data.append([
        Paragraph("<b>Grand Total :</b>", right_style),
        Paragraph(f"<b>{indian_currency(grand_total)}</b>", right_style)
    ])

    # Amount in Words
    amount_words = price_in_words(grand_total)
    totals_data.append([
        Paragraph(f"<b>Total in Words : {amount_words}</b>", small),
        ""
    ])

    totals_table = Table(totals_data, colWidths=[140 * mm, 30 * mm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -2), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('SPAN', (0, -1), (1, -1)),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.gray),
    ]))

    totals_wrapper = Table([[totals_table]], colWidths=[total_width])
    totals_wrapper.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    elements.append(totals_wrapper)

    # ── Terms & Conditions ──────────────────────────────────────────────────
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("<b>Terms & Conditions</b>", bold))

    ordered_terms = []
    terms_by_id = {t.id: t for t in quotation.terms_and_conditions.all()}
    for tid in quotation.terms_order or []:
        if tid in terms_by_id:
            ordered_terms.append(terms_by_id[tid])

    custom_terms = []
    if quotation.custom_terms:
        custom_terms = [t.strip() for t in quotation.custom_terms.strip().split('\n') if t.strip()]

    terms_paragraphs = []
    idx = 1
    for t in ordered_terms:
        terms_paragraphs.append([Paragraph(f"{idx}. {t.description}", small)])
        idx += 1
    for ct in custom_terms:
        terms_paragraphs.append([Paragraph(f"{idx}. {ct}", small)])
        idx += 1

    if not terms_paragraphs:
        terms_paragraphs = [[Paragraph("No Terms & Conditions.", small)]]

    terms_table = Table(terms_paragraphs, colWidths=[doc.width])
    terms_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEBEFORE', (0, 0), (0, -1), 3, light_blue),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))

    elements.append(terms_table)

    if quotation.apply_gst:
        elements.append(Paragraph(
            "All above material and services will be attracted to GST extra as per product or service applicable.",
            small
        ))
        elements.append(Spacer(1, 5))

    elements.append(Paragraph(
        "We thank you for the opportunity given to serve you & look forward to adding you to our family of customers.",
        small
    ))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("<b>Thanking You,</b>", small))
    elements.append(Spacer(1, 2))
    elements.append(Paragraph("<b>Seva Facility Services Pvt Ltd</b>", small))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<b>SFS Representative:</b> {quotation.contact_by} - {quotation.contact_by_no}", small))

    doc.build(elements)
    buffer.seek(0)

    download = request.GET.get("download", "false").lower() == "true"
    response = HttpResponse(buffer, content_type='application/pdf')
    if download:
        response['Content-Disposition'] = f'attachment; filename="Quotation_{quotation.quotation_no}.pdf"'
    else:
        response['Content-Disposition'] = f'inline; filename="Quotation_{quotation.quotation_no}.pdf"'

    return response

def get_message_templates(request):
    templates = MessageTemplates.objects.all()
    data = {
        'templates': templates
    }
    return render(request, 'message_templates.html', context=data)


def create_message_template(request):
    messages_type_choices = MessageTemplates.MESSAGE_TYPE_CHOICE
    category_choices = MessageTemplates.CATEGORY_CHOICES
    lead_status_choices = MessageTemplates.LEAD_STATUS_CHOICES

    if request.method == "POST":
        name = request.POST.get('name')
        message_type = request.POST.get('message_type')
        category = request.POST.get('category')
        lead_status = request.POST.get('lead_status')
        subject = request.POST.get('subject') if message_type == 'email' else None
        body = request.POST.get('body')
        attachment = request.FILES.get('attachment',None)  
        # Save to DB
        MessageTemplates.objects.create(
            name=name,
            message_type=message_type,
            category=category,
            lead_status = lead_status,
            subject=subject,
            body=body,
            attachment=attachment,
            is_active=True
        )

        return redirect('message_templates')  # redirect to your templates list page

    context = {
        'messages_type_choices': messages_type_choices,
        'category_choices': category_choices,
        'lead_status_choices':lead_status_choices,
    }
    return render(request, 'create_message_template.html', context)


def edit_message_template(request, id):
    template = MessageTemplates.objects.get(id=id)

    if request.method == "POST":
        body = request.POST.get('body', '')

        # Always update body
        template.body = body  

        # Only handle subject + attachment if Email
        if template.message_type == "email":
            subject = request.POST.get('subject', '')
            template.subject = subject

        attachment = request.FILES.get('attachment')
        if attachment:
            template.attachment = attachment

        template.save()
        return redirect('message_templates')

    return render(request, 'edit_message_template.html', {"templates": template})


from django.apps import apps
from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import redirect
from .tasks import send_email_task, send_whatsapp_task

def send_lead_email(request, pk):
    lead = get_object_or_404(lead_management, pk=pk)

    if not lead.customeremail or not lead.customeremail.strip():
        messages.error(request, f"{lead.customername} has no email address.")
        return redirect("display_lead_management")
    template = MessageTemplates.objects.filter(
        message_type='email',
        category='lead',
        lead_status__iexact=lead.typeoflead.strip().lower()
    ).first()
    print("template",template)
    print("lead",lead.typeoflead)
    if not template:
        messages.error(request, f"No template found for {lead.typeoflead} leads.")
        return redirect("display_lead_management")

    # Prepare placeholders
    placeholders = {
        "customername": lead.customername,
        "typeoflead": lead.typeoflead,
        "primarycontact": lead.primarycontact,
        # Add more fields if needed
    }

    body = template.body
    subject = template.subject
    for key, value in placeholders.items():
        body = body.replace(f"{{{key}}}", str(value))
        subject = subject.replace(f"{{{key}}}", str(value))

    attachment_path = None
    attachment_name = None
    if template.attachment:
        attachment_path = template.attachment.path  # full file path
        attachment_name = os.path.basename(template.attachment.name)  # file name


    send_email_task.delay(subject, body, recipient=lead.customeremail,attachment_path=attachment_path, attachment_name=attachment_name)
    messages.success(request, f"Email sent to {lead.customername}")
    # return redirect("display_lead_management")
    return redirect(request.META.get("HTTP_REFERER", "display_lead_management"))


def send_group_lead_email(request, lead_type):
    # Get only leads in 'lead' category with the given type
    leads = lead_management.objects.filter(typeoflead__iexact=lead_type)

    sent_count = 0
    skipped_count = 0

    for lead in leads:
        # Get latest follow-up
        latest_followup = main_followup.objects.filter(lead=lead).order_by('-created_at').first()

        # Skip if Close Win
        if latest_followup and latest_followup.order_status == 'Close Win':
            skipped_count += 1
            continue

        # Skip if no email
        if not lead.customeremail or not lead.customeremail.strip():
            skipped_count += 1
            continue

        # Fetch template
        template = MessageTemplates.objects.filter(
            message_type='email',
            category='lead',
            lead_status__iexact=lead_type.strip()
        ).first()
        if not template:
            skipped_count += 1
            continue

        # Prepare placeholders
        placeholders = {
            "customername": lead.customername,
            "typeoflead": lead.typeoflead,
            "primarycontact": lead.primarycontact,
        }

        body = template.body
        subject = template.subject
        for key, value in placeholders.items():
            body = body.replace(f"{{{key}}}", str(value))
            subject = subject.replace(f"{{{key}}}", str(value))

        attachment_path = None
        attachment_name = None
        if template.attachment:
            attachment_path = template.attachment.path  # full file path
            attachment_name = os.path.basename(template.attachment.name)  # file name


        send_email_task.delay(subject, body, recipient=lead.customeremail,attachment_path=attachment_path, attachment_name=attachment_name)
        # Send email asynchronously
        # send_email_task.delay(subject, body, recipient=lead.customeremail)
        sent_count += 1

    messages.success(request, f"✅ Emails sent: {sent_count}, ⏭️ Skipped: {skipped_count}")
    # return redirect("display_lead_management")
    return redirect(request.META.get("HTTP_REFERER", "display_lead_management"))



def send_lead_whatsapp(request, pk):
    lead = get_object_or_404(lead_management, pk=pk)

    if not lead.primarycontact:
        messages.error(request, f"{lead.customername} has no phone number.")
        return redirect(request.META.get("HTTP_REFERER", "display_lead_management"))
    
    # Fetch WhatsApp template for this lead type
    template = MessageTemplates.objects.filter(
        message_type='whatsapp',
        category='lead',
        lead_status__iexact=lead.typeoflead.strip().lower()
    ).first()

    if not template:
        messages.error(request, f"No WhatsApp template found for {lead.typeoflead}.")
        return redirect(request.META.get("HTTP_REFERER", "display_lead_management"))

    # Prepare placeholders
    placeholders = {
        "customername": lead.customername,
        "typeoflead": lead.typeoflead,
        "primarycontact": lead.primarycontact,
        # Add more fields if needed
    }

    # Replace placeholders in template body
    msg = template.body
    for key, value in placeholders.items():
        msg = msg.replace(f"{{{key}}}", str(value))

    # Handle attachment if present
    attachment_url = None
    attachment_name = None
    if template.attachment:
        # Use .url (relative to MEDIA_URL)
        relative_url = template.attachment.url  # e.g. /media/message_attachments/ape1_hbHE7TU.jpg

        # Build absolute URL
        site_url = getattr(settings, "SITE_URL", "https://www.teimcrm.com")  
        attachment_url = f"{site_url}{relative_url}"
        attachment_name = os.path.basename(template.attachment.name)
    
    print("url: ", attachment_url)
    print("attachment_name :",attachment_name)
    # Queue the Celery task
    mobile = f"91{lead.primarycontact}"
    send_whatsapp_task.delay(mobile, msg, attachment_url, attachment_name)

    messages.success(request, f"WhatsApp message queued for {lead.customername}")
    return redirect(request.META.get("HTTP_REFERER", "display_lead_management"))


def send_group_lead_whatsapp(request, lead_type):
    leads = lead_management.objects.filter(typeoflead__iexact=lead_type)
    sent_count = 0
    skipped_count = 0

    site_url = getattr(settings, "SITE_URL", "https://www.teimcrm.com")

    for lead in leads:
        latest_followup = main_followup.objects.filter(lead=lead).order_by('-created_at').first()

        # Skip Closed/No contact
        if latest_followup and latest_followup.order_status == 'Close Win':
            skipped_count += 1
            continue
        if not lead.primarycontact:
            skipped_count += 1
            continue

        # Fetch template
        template = MessageTemplates.objects.filter(
            message_type='whatsapp',
            category='lead',
            lead_status__iexact=lead_type.strip()
        ).first()
        if not template:
            skipped_count += 1
            continue

        # Prepare placeholders
        placeholders = {
            "customername": lead.customername,
            "typeoflead": lead.typeoflead,
            "primarycontact": lead.primarycontact,
        }

        body = template.body
        for key, value in placeholders.items():
            body = body.replace(f"{{{key}}}", str(value))

        # Prepare attachment URL if exists
        attachment_url = None
        attachment_name = None
        if template.attachment:
            relative_url = template.attachment.url
            attachment_url = f"{site_url}{relative_url}"
            attachment_name = os.path.basename(template.attachment.name)

        # Mobile number
        mobile = f"91{lead.primarycontact}"

        # Only send msg if no attachment
        # msg = None if attachment_url else body
        msg = body
        send_whatsapp_task.delay(mobile, msg, attachment_url, attachment_name)

        sent_count += 1

    messages.success(request, f"✅ WhatsApp sent: {sent_count}, ⏭️ Skipped: {skipped_count}")
    return redirect(request.META.get("HTTP_REFERER", "display_lead_management"))


def send_quotation_pdf_on_whatsapp(request, id):
    quotation = get_object_or_404(quotation_management, id=id)

    mobile = f"91{quotation.customer.primarycontact}"  # adjust if field name is different

        # Fetch WhatsApp template for this lead type
    template = MessageTemplates.objects.filter(
        message_type='whatsapp',
        category='quotation',
    ).first()

    if not template:
        messages.error(request, f"No WhatsApp template found for quotation.")
        return redirect(request.META.get("HTTP_REFERER", "display_quotation"))

    # Prepare placeholders
    placeholders = {
        "customername": quotation.customer.fullname,

    }

    # Replace placeholders in template body
    msg = template.body
    for key, value in placeholders.items():
        msg = msg.replace(f"{{{key}}}", str(value))

    # Handle attachment if present
    attachment_url = f"https://www.teimcrm.com/generate_quotation/quotation/pdf/{quotation.id}/view?download=True"
    attachment_name = f"quotation_{quotation.id}.pdf"

    # Trigger Celery task
    send_whatsapp_task.delay(
        mobile=mobile,
        msg=msg,
        attachment_path=attachment_url,   # must be accessible URL
        attachment_name=attachment_name
    )

    return redirect(request.META.get("HTTP_REFERER", "display_quotation")) 

def send_quotation_email(request, id):
    quotation = get_object_or_404(quotation_management, id=id)

    recipient = quotation.customer.primaryemail
    # subject = f"Quotation #{quotation.id}"
    template = MessageTemplates.objects.filter(
            message_type='email',
            category='quotation').first()
    
    if not template:
        messages.error(request, f"No template found for quotation.")
        return redirect("display_quotation")
    
    placeholders = {
        "customername": quotation.customer.fullname,
    }

    body = template.body
    subject = template.subject
    for key, value in placeholders.items():
        body = body.replace(f"{{{key}}}", str(value))
        subject = subject.replace(f"{{{key}}}", str(value))
    
    attachment_path = f"https://www.teimcrm.com/generate_quotation/quotation/pdf/{quotation.id}/view?download=True"
    attachment_name = f"quotation_{quotation.id}.pdf"
    
    send_email_task.delay(
        subject=subject,
        message=body,
        recipient=recipient,
        attachment_path=attachment_path,
        attachment_name=attachment_name
    )

    return redirect(request.META.get("HTTP_REFERER", "display_quotation"))


def send_invoice_email(request, id):
    invoice = get_object_or_404(TaxInvoice, id=id)

    recipient = invoice.customer.primaryemail
    # subject = f"Quotation #{quotation.id}"
    template = MessageTemplates.objects.filter(
            message_type='email',
            category='invoice').first()
    
    if not template:
        messages.error(request, f"No template found for Invoice.")
        return redirect("display_tax_invoice")
    
    placeholders = {
        "customername": invoice.customer.fullname,
    }

    body = template.body
    subject = template.subject
    for key, value in placeholders.items():
        body = body.replace(f"{{{key}}}", str(value))
        subject = subject.replace(f"{{{key}}}", str(value))
    
    attachment_path = f"https://www.teimcrm.com/tax-invoice/pdf/{invoice.id}/?download=true"
    attachment_name = f"Invoice_{invoice.customer.fullname}.pdf"
    
    send_email_task.delay(
        subject=subject,
        message=body,
        recipient=recipient,
        attachment_path=attachment_path,
        attachment_name=attachment_name
    )

    return redirect(request.META.get("HTTP_REFERER", "display_tax_invoice"))

def send_invoice_pdf_on_whatsapp(request, id):
    invoice = get_object_or_404(TaxInvoice, id=id)

    mobile = f"91{invoice.customer.primarycontact}"  

        # Fetch WhatsApp template for this lead type
    template = MessageTemplates.objects.filter(
        message_type='whatsapp',
        category='invoice',
    ).first()

    if not template:
        messages.error(request, f"No WhatsApp template found for quotation.")
        return redirect(request.META.get("HTTP_REFERER", "display_tax_invoice"))

    # Prepare placeholders
    placeholders = {
        "customername": invoice.customer.fullname,

    }

    # Replace placeholders in template body
    msg = template.body
    for key, value in placeholders.items():
        msg = msg.replace(f"{{{key}}}", str(value))

    # Handle attachment if present
    attachment_url = f"https://www.teimcrm.com/tax-invoice/pdf/{invoice.id}/?download=true"
    attachment_name = f"Invoice_{invoice.customer.fullname}.pdf"

    # Trigger Celery task
    send_whatsapp_task.delay(
        mobile=mobile,
        msg=msg,
        attachment_path=attachment_url,   # must be accessible URL
        attachment_name=attachment_name
    )

    return redirect(request.META.get("HTTP_REFERER", "display_tax_invoice")) 







def add_product_items(request):
    categories = Product.CATEGORY_CHOICES
    products = None
    selected_category = request.GET.get('category')
    selected_product = request.GET.get('product')
    selected_product_items = None

    # Filter products by category
    if selected_category:
        products = Product.objects.filter(category=selected_category)

    # Show existing items
    if selected_product:
        selected_product_items = ProductRequiredItem.objects.filter(product_id=selected_product)

    # Save items
    if request.method == "POST":
        product_id = request.POST.get('product')
        item_names = request.POST.getlist('item_name[]')

        for name in item_names:
            if name.strip():
                ProductRequiredItem.objects.create(
                    product_id=product_id,
                    item_name=name.strip()
                )

        return redirect(f"/product-items?category={selected_category}&product={product_id}")

    context = {
        'categories': categories,
        'products': products,
        'selected_category': selected_category,
        'selected_product': selected_product,
        'selected_product_items': selected_product_items
    }

    return render(request, 'product_items.html', context)


from django.shortcuts import get_object_or_404

def delete_product_item(request, item_id):
    item = get_object_or_404(ProductRequiredItem, id=item_id)

    product_id = item.product_id
    category = item.product.category

    item.delete()

    return redirect(f"/product-items?category={category}&product={product_id}")   





def get_amc_or_normal_services(service):

    # ❌ if not AMC → normal service
    if service.contract_type != "AMC":
        return [
            {
                "name": sp.product.product_name,
                "items": [
                    {
                        "name": sp.product.product_name,
                        "qty": sp.quantity
                    }
                ]
            }
            for sp in service.service_products.all()
        ]

    # ✅ AMC SERVICE → get AMC of this customer
    amc = AMCContract.objects.filter(
        customer=service.customer
    ).order_by('-id').first()

    if not amc:
        return []

    schedules = AMCServiceSchedule.objects.filter(
        amc=amc
    ).order_by("service_date")

    return [
        {
            "name": "AMC Visit",
            "items": [
                {
                    "name": sch.service_date,
                    "qty": "Completed" if sch.is_completed else "Pending"
                }
            ]
        }
        for sch in schedules
    ]




from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch

from crmapp.models import service_management
from amc.models import AMCContract, AMCServiceSchedule


@login_required
def service_list_with_amc(request):

    services = service_management.objects.all().prefetch_related(
        Prefetch(
            "amccontract_set__service_schedules",
            queryset=AMCServiceSchedule.objects.all().order_by("service_date")
        )
    )

    final_data = []

    for service in services:

        # 🔹 Get AMC (if exists)
        amc = service.amccontract_set.first()

        if amc:
            schedules = amc.service_schedules.all()

            sub_services = [
                {
                    "date": s.service_date,
                    "status": "Completed" if s.is_completed else "Pending",
                }
                for s in schedules
            ]
        else:
            sub_services = []

        final_data.append({
            "service": service,
            "amc": amc,
            "sub_services": sub_services
        })

    return render(request, "bharat.html", {
        "final_data": final_data
    })